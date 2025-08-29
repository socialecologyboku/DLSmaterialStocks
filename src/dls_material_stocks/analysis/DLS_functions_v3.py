import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import Patch
from scipy.optimize import curve_fit
from scipy.optimize import fsolve
from matplotlib.lines import Line2D
import matplotlib.lines as mlines
from matplotlib import gridspec
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import load_workbook

"""
@author: jstreeck

"""

"""############################################################

       Description:
           
        This script specifies the functions used in bottom-up
        calculations, and to produce plots.
        
   ############################################################"""


# write dictionary to excel file
def write_dict_to_excel(dictionary, file_name):
    print(f"Writing results to file: {file_name}")
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        for key, value in dictionary.items():
            if isinstance(value, pd.DataFrame):
                value.to_excel(writer, sheet_name=str(key))
            elif isinstance(value, pd.Series):
                # If the Series has a multi-index, reset it to a DataFrame
                if isinstance(value.index, pd.MultiIndex):
                    temp_df = value.reset_index()
                else:
                    # Convert Series to DataFrame
                    temp_df = value.to_frame(name="Value")
                temp_df.to_excel(writer, sheet_name=str(key))
            else:
                # Handle different types of values
                if isinstance(value, (list, tuple, np.ndarray)):
                    # Convert lists, tuples, and numpy arrays to DataFrame
                    if np.ndim(value) == 1:
                        temp_df = pd.DataFrame(value, columns=["Value"])
                    else:
                        temp_df = pd.DataFrame(value)
                elif isinstance(value, dict):
                    # Convert dictionaries to DataFrame with two columns: keys and values
                    temp_df = pd.DataFrame(
                        list(value.items()), columns=["Key", "Value"]
                    )
                else:
                    # Handle single scalar values (including numpy floats, strings, etc.)
                    temp_df = pd.DataFrame([value], columns=["Value"])

                temp_df.to_excel(writer, sheet_name=str(key))


    wb = load_workbook(file_name)
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                cell.alignment = Alignment(horizontal="left")
                try:
                    val = str(cell.value)
                    if val:
                        max_length = max(max_length, len(val))
                except:
                    # what the hell?
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2
    wb.save(file_name)

# Read all sheets of excel file into a dictionary of dataframes
def read_excel_into_dict(file_path):
    xls = pd.ExcelFile(file_path)
    sheet_dict = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}
    return sheet_dict


# multiply function demand with indirect stock intensity and save in dictionary
def create_nested_dict(initial_dict, second_df):
    nested_dict = {}

    # Iterate over the rows of the second dataframe
    for (region, year), row in second_df.iterrows():
        # Create a sub-dictionary for each region and year
        nested_dict[(region, year)] = {}

        # Iterate over the keys (column names) in the initial dictionary
        for key in initial_dict.keys():
            # Check if key exists as a column in the second dataframe
            if key in second_df.columns:
                # Multiply the dataframe from initial_dict by the corresponding value in second_df
                nested_dict[(region, year)][key] = initial_dict[key] * row[key]

    return nested_dict


# expand dictionary from function 'create_nested_dict' to dataframe
def expand_nested_dict_to_df(d, parent_key=(), results=None):
    if results is None:
        results = []

    for key, value in d.items():
        new_key = parent_key + (key,)
        if isinstance(value, dict):
            expand_nested_dict_to_df(value, new_key, results)
        elif isinstance(value, pd.DataFrame):
            # Create a MultiIndex combining the new_key and DataFrame's index and columns
            tuples = [
                new_key + (row_index, col)
                for row_index in value.index
                for col in value.columns
            ]
            multi_index = pd.MultiIndex.from_tuples(tuples)
            # Flatten the DataFrame values and append them to the results
            flattened_values = value.to_numpy().flatten().astype(float)
            expanded_df = pd.DataFrame(flattened_values, index=multi_index)
            results.append(expanded_df)

    if parent_key == ():
        # Concatenate all DataFrames in the results
        return pd.concat(results)


# prepares dataframe output from 'expand_nested_dict_to_df' to concat with these outputs from other DLS dimensions (aligns index)
def prepare_df_for_concat(
    df, unit, stock_type, dimension, stock, year=None, product=None
):
    # input: dataframe
    # unit: specified under unit input variable (only kg, tons allowed)
    # df.index.names = ['region,year','product', 'material', 'unit']
    if "unit" not in df.index.names:
        df.insert(1, "unit", unit)
    if "stock_type" not in df.index.names:
        df.insert(1, "stock_type", stock_type)
    if "dimension" not in df.index.names:
        df.insert(1, "dimension", dimension)
    if "stock" not in df.index.names:
        df.insert(1, "stock", stock)
    if "year" not in df.index.names:
        df.insert(1, "year", year)
    if "product" not in df.index.names:
        df.insert(1, "product", product)
    # set column header of column with values to 'value
    # check unit and convert to tons if kg
    df.rename(columns={0: "value", "value_region_year": "value"}, inplace=True)
    if unit == "kg":
        df.value = df.value / 1000
    df.reset_index(inplace=True)
    df.unit = "tons"
    # if region,year are tuple, split to individual columns
    if all(name in df.columns.tolist() for name in ["region,year"]):
        tuples = df["region,year"]
        split_tuples = [list(t) for t in tuples]
        df[["region", "year"]] = split_tuples
    # set index and only keep columns with values
    df.set_index(
        [
            "region",
            "year",
            "dimension",
            "stock_type",
            "product",
            "stock",
            "material",
            "unit",
        ],
        inplace=True,
    )
    df = df[["value"]]
    return df


# calculate DLS functions from per capita to at scale (=function_demand/region, e.g., total_pkm/region)
def calc_functions_scale(direct_cap, MISO2_population_2015_subset):
    if direct_cap.index.names != ["region", "variable", "unit"]:
        direct_cap.set_index(["region", "variable", "unit"], inplace=True)
    
    direct_cap_reset = direct_cap.reset_index()
    pop_data = MISO2_population_2015_subset.reset_index()
    
    merged = direct_cap_reset.merge(pop_data, on='region', how='left')
    
    for col in direct_cap.columns:
        merged[col] = merged[col] * merged['2015'] * 1000
    
    merged = merged.drop('2015', axis=1)
    DLS_region = merged.set_index(['region', 'variable', 'unit'])
    
    return DLS_region


# calculate direct flows [kcal] of food per food item based on diet shares and kcal level per capita and day
def calc_food_flows_kcal_cap(diet_shares, kcal_level):
    diets_combined = diet_shares.merge(kcal_level[["region", "value"]], on="region")
    for column in diets_combined.columns.drop(["region", "value"]):
        diets_combined[column] = diets_combined[column] * diets_combined["value"]
    diets_combined.drop("value", axis=1, inplace=True)
    diets_combined.insert(1, "year", diet_shares.reset_index().year)
    return diets_combined


# convert direct flows of food [kcal/cap/day] to [kg/cap/day]
def conv_food_flows_kg_cap(food_flows_kcal, calorific_values):
    food_flows_kg = food_flows_kcal.set_index(["region", "year"]).div(
        calorific_values * 10, axis=1
    )
    return food_flows_kg


# calculate direct flows of food at scale [tons/cap/year/region]
def calc_food_flows_kg_scale(food_flows_kg, MISO2_population_2015_subset):
    food_flows_kg_reset = food_flows_kg.reset_index()
    pop_data = MISO2_population_2015_subset.reset_index()
    
    merged = food_flows_kg_reset.merge(pop_data, on='region', how='left')
    
    food_cols = [col for col in food_flows_kg.columns]
    # 365 days in a year
    merged[food_cols] = merged[food_cols].multiply(merged['2015'] * 365, axis=0)
    
    DLS_food_region = merged[['region', 'year'] + food_cols].set_index(['region', 'year'])
    return DLS_food_region


# multiply by indirect stock intensities per food product
def calc_indir_stock_intens_cap(
    food_kg_cap, diet_ecoinv_dict, nutrition_indstocks_dict
):
    food_kg_cap.rename(columns=diet_ecoinv_dict, inplace=True)
    # for key in nutrition_indstocks_dict:
    #     nutrition_indstocks_dict.get(key).set_index('ISIC_stocks',inplace=True)
    nutrition_indirect_stocks_intens = create_nested_dict(
        nutrition_indstocks_dict, food_kg_cap
    )

    nutrition_indirect_stocks_intens_df = expand_nested_dict_to_df(
        nutrition_indirect_stocks_intens
    )
    nutrition_indirect_stocks_intens_df.index.names = [
        "region,year",
        "product",
        "stock",
        "material",
    ]
    nutrition_indirect_stocks_intens_df = nutrition_indirect_stocks_intens_df[
        nutrition_indirect_stocks_intens_df[0] != 0
    ]
    nutrition_indirect_stocks_intens_df.rename(
        columns={0: "value_cap_day"}, inplace=True
    )
    # Extract the tuples from the desired level
    tuples = nutrition_indirect_stocks_intens_df.index.get_level_values(0)
    # Split the tuples into separate lists
    split_tuples = [list(t) for t in tuples]
    nutrition_indirect_stocks_intens_df[["region", "year"]] = split_tuples
    nutrition_indirect_stocks_intens_df = (
        nutrition_indirect_stocks_intens_df.reset_index()
        .set_index(["region", "year", "product", "stock", "material"])
        .drop(columns=["region,year"])
    )
    return nutrition_indirect_stocks_intens_df


# nutrition_indirect_stocks_intens_df as stock per kg/cap/day of food, MISO2_pop in 1000's
# resulting unit: tons/region/stock_category/material
def calc_indir_stock_scale(food_ind_stock_intens_cap, population):
    # Reset index of df_multi for the join operation and keep the original index
    df_multi_reset = food_ind_stock_intens_cap.reset_index()
    df_joined = df_multi_reset.join(population, on="region")
    df_joined["value_region_year"] = (
        df_joined["value_cap_day"] * df_joined["2015"] * 365
    )
    # Set the original index back to df_joined
    df_final = df_joined.set_index(food_ind_stock_intens_cap.index.names)
    df_final.drop(["value_cap_day", "2015"], axis=1, inplace=True)
    return df_final


# split material demand to regions that map to ecoinvent process regional distinction
def regionalize_material_demand_ecoinvent(
    materials, material_demand, path_correspondence
):
    # Iterate through each material
    material_demand_reg = material_demand.copy()

    for material in materials:
        # Load the correspondence table for the current material
        correspondence_df = pd.read_excel(
            path_correspondence, sheet_name=material, index_col="region"
        )

        # Create a dictionary to map regions to new column names
        region_to_column = correspondence_df["ecoinvent_process"].to_dict()

        # Add new columns to the DataFrame for each unique ecoinvent process
        for new_column in region_to_column.values():
            if new_column not in material_demand_reg.columns:
                material_demand_reg[new_column] = np.nan

        # Iterate through each region and update the DataFrame
        for region, new_column in region_to_column.items():
            if material in material_demand_reg.columns:
                material_demand_reg.loc[
                    material_demand_reg.index == region, new_column
                ] = material_demand_reg.loc[
                    material_demand_reg.index == region, material
                ].astype(float)

        # Function to validate the sum before and after the transformation
        def validate_sums(original_df, transformed_df, material, region_to_column):
            original_sum = original_df[material].sum()
            new_columns = list(set(list(region_to_column.values())))
            transformed_sum = transformed_df[new_columns].sum().sum()
            return original_sum, transformed_sum

        # Validate the sums before dropping the original columns
        original_sum, transformed_sum = validate_sums(
            material_demand_reg, material_demand_reg, material, region_to_column
        )
        # print(
        #     f"Validation for {material}: Original Sum = {original_sum}, Transformed Sum = {transformed_sum}"
        # )
        assert original_sum - transformed_sum < 0.01, (
            f"Sum mismatch for {material}: {original_sum} != {transformed_sum}"
        )

    # Drop the original material columns
    material_demand_reg.drop(columns=materials, inplace=True)
    return material_demand_reg


# split material demand to regions that map to ecoinvent process regional distinction - specific for coal
def regionalize_material_demand_ecoinvent_coal(
    materials, material_demand, path_correspondence
):
    # Iterate through each material
    material_demand_reg = material_demand.copy()

    for material in materials:
        # Load the correspondence table for the current material
        correspondence_df = pd.read_excel(
            path_correspondence, sheet_name=material, index_col="region"
        )

        # Create a dictionary to map regions to new column names
        region_to_column = correspondence_df["ecoinvent_process"].to_dict()

        # Add new columns to the DataFrame for each unique ecoinvent process
        for new_column in region_to_column.values():
            if new_column not in material_demand_reg.columns:
                material_demand_reg[new_column] = np.nan

        # Iterate through each region and update the DataFrame
        for region, new_column in region_to_column.items():
            if material in material_demand_reg.columns:
                material_demand_reg.loc[
                    material_demand_reg.index.get_level_values(0) == region, new_column
                ] = material_demand_reg.loc[
                    material_demand_reg.index.get_level_values(0) == region, material
                ]

        # Function to validate the sum before and after the transformation
        def validate_sums(original_df, transformed_df, material, region_to_column):
            original_sum = original_df[material].sum()
            new_columns = list(set(list(region_to_column.values())))
            transformed_sum = transformed_df[new_columns].sum().sum()
            return original_sum, transformed_sum

        # Validate the sums before dropping the original columns
        original_sum, transformed_sum = validate_sums(
            material_demand_reg, material_demand_reg, material, region_to_column
        )
        print(
            f"Validation for {material}: Original Sum = {original_sum}, Transformed Sum = {transformed_sum}"
        )
        assert original_sum - transformed_sum < 0.01, (
            f"Sum mismatch for {material}: {original_sum} != {transformed_sum}"
        )

    # Drop the original material columns
    material_demand_reg.drop(columns=materials, inplace=True)
    return material_demand_reg


# prepare DLS indicator df for upscaling
def prep_DLS_func_for_scaling_single(DLS_func, function_name):
    DLS_func_prep = DLS_func.reset_index()[["region", "value"]]
    DLS_func_prep.insert(0, "year", 2015)
    DLS_func_prep.set_index(["region", "year"], inplace=True)
    DLS_func_prep.rename(columns={"value": function_name}, inplace=True)
    return DLS_func_prep


# to direct stocks dataframe, add indices of indirect stock df to enable subtraction on all indices of the latter
def align_indices_zero(df_target, df_indirect):
    df_indirect_aligned, df_target_aligned = df_indirect.align(df_target, join="outer")
    df_target_aligned.fillna(0, inplace=True)
    return df_target_aligned


# split DLS indicators to different accounts based on share data (e.g. housing archetypes)
def calc_function_split(functions, split, function_split_label):
    # merge functions and mat_intens to prepare for multiplication
    combined = functions.merge(split.reset_index(), on="region")
    # multiply each function (columns, except 'region') by the archetype split
    combined[function_split_label] = combined["functions"] * combined["split"]
    return combined


# calculate from MI per m² to MI per capita (housing)
def calc_direct_cap(functions_split, mat_intens, function_split_label):
    # merge functions by archetype and mat_intens to prepare for multiplication
    combined = functions_split.merge(mat_intens, on=["region", "stock type"])
    combined.set_index(["region", "stock type"], inplace=True)
    # multiply each category share (columns, except 'region') by fuctions
    for column in combined.columns:
        if column != function_split_label:
            new_label = column.replace("kg/m²", "kg/cap")
            combined[new_label] = combined[column] * combined[function_split_label]
    return combined


# calculate direct stocks at scale from direct stocks per capita
def calc_direct_scale(direct_cap, MISO2_population_2015_subset):
    DLS_region = pd.DataFrame(columns=direct_cap.columns, index=direct_cap.index)
    
    for index, row in direct_cap.iterrows():
        region, year = index
        matching_row = MISO2_population_2015_subset.reset_index()[
            MISO2_population_2015_subset.reset_index()["region"] == region
        ]
        if not matching_row.empty:
            population = matching_row.iloc[0]["2015"]
            DLS_region.loc[index] = row * population
    DLS_region.columns = DLS_region.columns.str.replace(
        r"\(kg/cap\)", "", regex=True
    ).str.strip()
    DLS_region.insert(0, "unit", "tons")
    DLS_region = DLS_region.set_index(["unit"], append=True)
    return DLS_region


# multiply a single specified column with all other columns
def mult_singleCol_otherCols(dataframe_input, single_col_header):
    # multiply each category share (columns, except 'region') by fuctions
    for column in dataframe_input.columns:
        if column != single_col_header:
            new_label = column.replace("kg/m²", "kg")
            dataframe_input[new_label] = (
                dataframe_input[column] * dataframe_input[single_col_header]
            )
    return dataframe_input


# scale up from capita to national level values
def scale_kg_cap(df_kgcap, pop):
    df_scale = df_kgcap.copy()
    df_scale = df_scale.join(pop.rename(columns={"2015": "pop"}), on="region")
    df_scale["value_scale"] = df_scale["value"] * df_scale["pop"]
    df_scale = df_scale.rename({"kg/cap": "tons"})
    df_scale = (
        df_scale.reset_index()
        .set_index(
            [
                "region",
                "year",
                "dimension",
                "stock_type",
                "product",
                "stock",
                "material",
                "unit",
            ]
        )
        .drop(columns=["value", "pop"])
    )
    return df_scale


# multiply a dictionary with a dataframe and save as nested dictionary - specific for energy dimension
def create_nested_dict_energy(dict_dfs, df_multipliers):
    # Filter df_multipliers to include only the regions present in dict_dfs
    filtered_multipliers = df_multipliers[
        df_multipliers.index.get_level_values("region").isin(dict_dfs.keys())
    ]

    # Similarly, filter dict_dfs to include only the regions present in df_multipliers
    filtered_dict_dfs = {
        region: df
        for region, df in dict_dfs.items()
        if region in filtered_multipliers.index.get_level_values("region").unique()
    }

    # Initialize the resulting nested dictionary
    nested_dict_results = {}

    # Iterate over each row in the filtered df_multipliers to perform the multiplication
    for idx, multiplier_row in filtered_multipliers.iterrows():
        region, variable, elec, unit_energy = idx

        # At this point, the region is guaranteed to be in filtered_dict_dfs
        df_to_modify = filtered_dict_dfs[region].copy()

        # Iterate through each multiplier column
        for col in multiplier_row.index:
            modified_df = df_to_modify.copy()

            # Apply each multiplier to the entire DataFrame
            # This assumes a simplistic model where each column in df_multipliers
            # represents a distinct multiplier to be applied across all columns of the matched DataFrame
            # Adjust as necessary for your specific application
            modified_df = modified_df.multiply(multiplier_row[col], axis="index")

            # Use a nested key that includes the column name (multiplier name) for clarity
            nested_key = idx + (col,)
            nested_dict_results[nested_key] = modified_df

    return nested_dict_results


# adjust index for fuel use df
def adjust_index_fuel(df, index_names, tuple_pos, result_index, op_con):
    df.index.names = index_names
    df.rename(columns={0: "value"}, inplace=True)
    df = df[df.value != 0]
    tupled_index = df.index.get_level_values(tuple_pos)
    split_tuples = [list(t) for t in tupled_index]
    df.loc[:, result_index] = split_tuples
    df.loc[:, "unit"] = "kg/cap"
    df.insert(1, "year", "2015")
    df.insert(1, "stock_type", "indirect")
    # substitute dimension details with higher level labels
    subst_dim = {
        "Appliance|clean_cooking_fuel": "nutrition_i",
        "Appliance|television": "communication_i",
        "Appliance|mobile_telephone": "communication_i",
        "Appliance|refrigerator": "nutrition_i",
        "Water": "water_i",
        "Sanitation": "sanitation_i",
        "Nutrition": "nutrition_i",
        "Clothing|clothing": "clothing_i",
        "Clothing|footwear": "clothing_i",
        "Health care": "health_i",
        "Education|primary": "education_i",
        "Education|lower_secondary": "education_i",
        "Housing|total": "housing_i",
        "Roads": "transport_i",
        "Hot Water OP|total": "housing_i",
        "Transport": "transport_i",
        "Heating": "housing_i",
        "Cooling": "housing_i",
    }
    df = df.replace(subst_dim)
    df.reset_index(inplace=True)
    df["product"] = df["product"] + "_" + op_con
    df = df.set_index(
        [
            "region",
            "year",
            "dimension",
            "stock_type",
            "product",
            "stock",
            "material",
            "unit",
        ]
    ).drop(columns=["tuple", "fuel"])
    df = df.groupby(
        [
            "region",
            "year",
            "dimension",
            "stock_type",
            "product",
            "stock",
            "material",
            "unit",
        ]
    ).sum()
    df = df[df.value != 0]
    return df


# multiply a dictionary with a dataframe and save as nested dictionary - specific to fuels
# does not save dataframes that only have nans or zeros
def create_nested_dict_fuel(initial_dict, second_df):
    nested_dict = {}

    # Iterate over the rows of the second dataframe
    for (region, variable, elec, unit), row in second_df.iterrows():
        # Create a sub-dictionary for each region and year
        nested_dict[(region, variable, elec, unit)] = {}

        # Iterate over the keys (column names) in the initial dictionary
        for key in initial_dict.keys():
            # Check if key exists as a column in the second dataframe
            if key in second_df.columns:
                # Multiply the dataframe from initial_dict by the corresponding value in second_df
                result_df = initial_dict[key] * row[key]
                # Check if the resulting DataFrame is not all NaN or empty
                if not result_df.isna().all().all() and not result_df.empty:
                    nested_dict[(region, variable, elec, unit)][key] = result_df

        # Remove the entry if the sub-dictionary is empty
        if not nested_dict[(region, variable, elec, unit)]:
            del nested_dict[(region, variable, elec, unit)]

    return nested_dict


"""  #########################

              PLOTS
            
     ######################### """


### REGRESSION functions for use in plots


## SEMI-LOGARITHMIC REGRESSION
def log_func(x, a, b):
    return a * np.log(x) + b


## LOGISTIC / SATURATION REGRESSION
def logistic_func(x, a, b, c):
    return a / (1 + b * np.exp(-c * x))


## POLYNOMIAL REGRESSION
def poly_func(x, a, b, c):
    return a * x**2 + b * x + c


# split the title into two lines if longer than max_length
def split_title(title, max_length=15):
    if len(title) > max_length:
        # Find the nearest space to the middle of the string to split
        middle = len(title) // 2
        left = title.rfind(" ", 0, middle)
        right = title.find(" ", middle)

        if right != -1 and (middle - left > right - middle or left == -1):
            split_point = right
        else:
            split_point = left

        if split_point != -1:
            # Split the title at the split_point
            return title[:split_point] + "\n" + title[split_point + 1 :]
    return title


def plot_geo_multiplot(DLS_stock_reached, columns, titles):
    # Load the world map data
    world = DLS_stock_reached
    # Remove Antarctica
    world = world[world["continent"] != "Antarctica"]

    # Create figure and axes for the 2x2 subplot layout
    fig, axes = plt.subplots(3, 2, figsize=(20, 15))  # 2 rows, 2 columns
    axes = axes.flatten()  # Flatten the 2D array of axes to make iteration easier

    # Loop through each material column and corresponding title
    for i, (column, title) in enumerate(zip(columns, titles)):
        ax = axes[i]

        # Plot the world map for the current material column
        world.plot(
            column=column,
            ax=ax,
            cmap="coolwarm",
            legend=False,
            edgecolor="black",
            zorder=0,
        )

        # Remove plot frame (axis spines)
        for spine in ax.spines.values():
            spine.set_visible(False)

        # Remove x and y labels from the map plot
        ax.set_xticks([])
        ax.set_yticks([])

        # Set the title for each subplot based on the material
        ax.set_title(title, fontsize=20, loc="left")

    # Add color bar for the entire figure (shared by all subplots)
    sm = plt.cm.ScalarMappable(cmap="coolwarm", norm=plt.Normalize(vmin=0, vmax=1))
    cbar_ax = fig.add_axes(
        [0.14, 0.08, 0.76, 0.02]
    )  # Adjust position and size as needed
    cbar = fig.colorbar(sm, cax=cbar_ax, orientation="horizontal")
    cbar.ax.tick_params(labelsize=13)
    cbar.ax.text(
        0,
        2,
        "No DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=14,
    )
    cbar.ax.text(
        0.85,
        2,
        "DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=14,
    )

    axes[5].axis("off")
    # Adjust layout for better spacing
    plt.tight_layout(rect=[0, 0.1, 1, 1])  # Leave space for color bar at the bottom
    fig.suptitle(
        "Relative achievement of DLS stocks at country-level & across material groups (2015)",
        fontsize=32,
        y=1.05,
    )
    plt.show()


def plot_stacked_bars_sub_geo(df1, df2, subcategory, title, DLS_stock_reached):
    # subcategories is the label of the second index columns (besides 'region)
    grouped_df1 = df1.groupby(["region", subcategory]).sum()
    grouped_df2 = df2.groupby(["region", subcategory]).sum()

    grouped_df1 = pd.DataFrame(grouped_df1.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df2 = pd.DataFrame(grouped_df2.sum(axis=1)).rename(columns={0: "stock"})

    # Creating a single dataframe for plotting

    combined_df = pd.concat([grouped_df1, grouped_df2], axis=1, keys=["df1", "df2"])
    combined_df.columns = combined_df.columns.map("_".join)
    combined_df = combined_df / 1e3  # t/cap now

    # Load a world map
    # world = gpd.read_file(gpd.datasets.get_path('naturalearth_lowres'))
    world = DLS_stock_reached
    # Remove Antarctica
    world = world[world["continent"] != "Antarctica"]

    # Create figure and axis for the map
    fig, ax = plt.subplots(figsize=(20, 12))

    import seaborn as sns
    from matplotlib.colors import LinearSegmentedColormap

    # Get a light blue palette from seaborn
    # light_blue_palette = sns.color_palette("Blues", 8)
    light_blue_palette = sns.color_palette("coolwarm", 3)[::-1]
    light_blue_cmap = LinearSegmentedColormap.from_list(
        "light_blue", light_blue_palette
    )

    light_blue_cmap = sns.dark_palette("#69d", as_cmap=True)
    light_blue_cmap = sns.color_palette("ch:s=-.2,r=.6", as_cmap=True)
    light_blue_cmap = sns.color_palette("ch:start=.1,rot=-.15", as_cmap=True)

    # Plot the world map
    world.plot(
        column="value_cap",
        ax=ax,
        cmap=light_blue_cmap,
        legend=False,
        edgecolor="black",
        zorder=0,
    )
    # cmap='coolwarm'
    # Remove plot frame (axis spines)
    for spine in ax.spines.values():
        spine.set_visible(False)

    # Remove x and y labels from the map plot
    ax.set_xticks([])
    ax.set_yticks([])

    region_locations = {
        "AFR": (-27, -70),
        "CPA": (175, 10),
        "EEU": (18, 85 + 5),
        "FSU": (85, 80 + 5),
        "LAM": (-160, -5),
        "MEA": (-50, 8),
        "NAM": (-155, 72 + 5),
        "PAO": (165, 72 + 5),
        "PAS": (165, -55),
        "SAS": (60, -65),
        "WEU": (-45, 80 + 5),
    }

    region_full_names = {
        "AFR": "Sub-Saharan Africa",
        "CPA": "Centrally planned Asia",
        "EEU": "Eastern Europe",
        "FSU": "Former Soviet Union",
        "LAM": "Latin America",
        "MEA": "N. Africa & Middle East",
        "NAM": "North America",
        "PAO": "Japan, Australia, New Zealand",
        "PAS": "Pacific Asia",
        "SAS": "South Asia",
        "WEU": "Western Europe",
    }

    region_connections = {
        "AFR": (30, 0),
        "CPA": (95, 35),
        "EEU": (23, 55),
        "FSU": (95, 70),
        "LAM": (-80, -10),
        "MEA": (25, 25),
        "NAM": (-100, 60),
        "PAO": (145, 55),
        "PAS": (115, 5),
        "SAS": (75, 30),
        "WEU": (5, 50),
    }

    # Unique regions and dimensions for plotting
    regions = combined_df.index.get_level_values(0).unique()
    dimensions = combined_df.index.get_level_values(1).unique()

    # Colors for df1 using a distinct colormap
    colors_df1 = plt.cm.tab20(np.linspace(0, 1, len(dimensions)))
    colors_df2 = plt.cm.Reds(np.linspace(0.5, 1, len(dimensions)))

    # Legend handles
    legend_handles = [
        plt.Rectangle((0, 0), 1, 1, facecolor=color, edgecolor="black")
        for color in colors_df1
    ] + [plt.Rectangle((0, 0), 1, 1, facecolor="white", edgecolor="black")]

    # Width of each bar
    width = 0.35  # Adjusted for two bars side by side

    # Transform coordinates to figure space
    trans = ax.transData.transform
    inv = fig.transFigure.inverted().transform

    # Plot bar charts at each region's location
    for region in regions:
        if region in region_locations and region in region_connections:
            # Transform region location to figure space
            x_fig, y_fig = inv(trans(region_locations[region]))

            # Create a new axis at this location
            region_ax = fig.add_axes(
                [x_fig, y_fig, 0.06, 0.125]
            )  # Adjust size as needed

            # Add a title to each region's bar plot
            region_ax.set_title(
                split_title(region_full_names.get(region)), fontsize=13, loc="center"
            )

            # Plotting
            for j, dimension in enumerate(dimensions):
                df1_value = combined_df.loc[(region, dimension), "df1_stock"]
                df2_value = combined_df.loc[(region, dimension), "df2_stock"]

                # set negatives to zero (negatives occur in case the DLS stock threshold with practice changes is lower than DLS stocks in 2015)
                df2_value = df2_value[df2_value > 0]

                # Plot df1 value
                region_ax.bar(
                    j, df1_value, color=colors_df1[j], edgecolor="black", zorder=2
                )

                # Plot df2 value as part of the stack on top of df1
                region_ax.bar(
                    j,
                    df2_value,
                    bottom=df1_value,
                    color="white",
                    edgecolor="black",
                    zorder=2,
                )

            # Correct y-ticks for each region's bar plot
            max_value = combined_df.loc[region, :].sum(axis=1).max()
            region_ax.set_yticks(
                np.round(np.linspace(0, max_value, 5))
            )  # Adjust number of ticks as needed
            region_ax.set_ylabel("tons / capita", fontsize=14)
            region_ax.yaxis.set_tick_params(labelsize=12)

            region_ax.set_xticks([])

            # Get connection line end points
            line_end = trans(
                region_connections[region]
            )  # Country location in figure space
            line_end_fig = inv(line_end)

            # Draw first part of the line on the main map (ax)
            con_line_main = plt.Line2D(
                [line_end_fig[0], x_fig + 0.05],
                [line_end_fig[1], y_fig + 0.05],
                transform=fig.transFigure,
                color="black",
                zorder=0,  # Set zorder to 0
            )
            ax.add_line(con_line_main)  # Add the line to the main map (ax)

            # Draw second part of the line on the region axis (region_ax)
            con_line_region = plt.Line2D(
                [0, 0.05],
                [0, 0.05],
                transform=region_ax.transAxes,
                color="black",
                zorder=1,  # Set zorder to -1
            )
            region_ax.add_line(
                con_line_region
            )  # Add the line to the region axis (region_ax)

    # Corresponding labels for the legend handles
    legend_labels = [
        "..existing, health",
        "..existing, mobility",
        "..existing, nutrition",
        "..existing, shelter",
        "..existing, socialization",
    ]  # Add labels for the Rectangle objects
    legend_labels.extend(["..to close DLS gaps"])  # Add labels for the Line2D objects

    # Create the legend with both Rectangle and Line2D entries
    legend = ax.legend(
        handles=legend_handles,
        labels=legend_labels,
        title="DLS material stocks..",
        loc="lower left",
        bbox_to_anchor=(0.05, -0.06),
        fontsize=14,
        frameon=False,
    )
    plt.setp(legend.get_title(), fontsize="14", color="black", weight="normal")
    ax.set_title(title, fontsize=30, pad=160, x=0.52)
    # Add color bar for country values under the map
    sm = plt.cm.ScalarMappable(cmap=light_blue_cmap, norm=plt.Normalize(vmin=0, vmax=1))
    cbar_ax = fig.add_axes(
        [0.15, 0.15, 0.76, 0.02]
    )  # Adjust position and size as needed
    cbar = fig.colorbar(sm, cax=cbar_ax, orientation="horizontal")
    cbar.ax.tick_params(labelsize=15)
    cbar.ax.text(
        0,
        2,
        "No DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
    )
    cbar.ax.text(
        0.85,
        2,
        "DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
    )
    cbar.set_ticklabels([f"{int(100 * x)}%" for x in np.linspace(0, 1, 6)])
    plt.show()
    # fig.savefig('Fig1a_DLSgapMap.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    return combined_df


def plot_stacked_bars_sub_geo_converge(
    df1, df2, converge, subcategory, title, DLS_stock_reached
):
    # subcategories is the label of the second index columns (besides 'region)
    grouped_df1 = df1.groupby(["region", subcategory]).sum()
    grouped_df2 = df2.groupby(["region", subcategory]).sum()

    grouped_df1 = pd.DataFrame(grouped_df1.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df2 = pd.DataFrame(grouped_df2.sum(axis=1)).rename(columns={0: "stock"})

    # Creating a single dataframe for plotting
    combined_df = pd.concat(
        [grouped_df1, grouped_df2, converge], axis=1, keys=["df1", "df2", "converge"]
    )
    combined_df.columns = combined_df.columns.map("_".join)
    combined_df = combined_df / 1e3  # t/cap now

    # Load a world map
    world = DLS_stock_reached
    # Remove Antarctica
    world = world[world["continent"] != "Antarctica"]

    # Create figure and axis for the map
    fig, ax = plt.subplots(figsize=(20, 12))

    import seaborn as sns
    from matplotlib.colors import LinearSegmentedColormap

    # Get a light blue palette from seaborn
    # light_blue_palette = sns.color_palette("Blues", 8)
    light_blue_palette = sns.color_palette("coolwarm", 3)[::-1]
    light_blue_cmap = LinearSegmentedColormap.from_list(
        "light_blue", light_blue_palette
    )

    light_blue_cmap = sns.dark_palette("#69d", as_cmap=True)
    light_blue_cmap = sns.color_palette("ch:s=-.2,r=.6", as_cmap=True)
    light_blue_cmap = sns.color_palette("ch:start=.1,rot=-.15", as_cmap=True)

    # Plot the world map
    world.plot(
        column="value_cap",
        ax=ax,
        cmap=light_blue_cmap,
        legend=False,
        edgecolor="black",
        zorder=0,
    )

    # Remove plot frame (axis spines)
    for spine in ax.spines.values():
        spine.set_visible(False)

    # Remove x and y labels from the map plot
    ax.set_xticks([])
    ax.set_yticks([])

    region_locations = {
        "AFR": (-27, -70),
        "CPA": (175, 10),
        "EEU": (18, 85 + 5),
        "FSU": (85, 80 + 5),
        "LAM": (-160, -5),
        "MEA": (-50, 8),
        "NAM": (-155, 72 + 5),
        "PAO": (165, 72 + 5),
        "PAS": (165, -55),
        "SAS": (60, -65),
        "WEU": (-45, 80 + 5),
    }

    region_full_names = {
        "AFR": "Sub-Saharan Africa",
        "CPA": "Centrally planned Asia",
        "EEU": "Eastern Europe",
        "FSU": "Former Soviet Union",
        "LAM": "Latin America",
        "MEA": "N. Africa & Middle East",
        "NAM": "North America",
        "PAO": "Japan, Australia, New Zealand",
        "PAS": "Pacific Asia",
        "SAS": "South Asia",
        "WEU": "Western Europe",
    }

    region_connections = {
        "AFR": (30, 0),
        "CPA": (95, 35),
        "EEU": (23, 55),
        "FSU": (95, 70),
        "LAM": (-80, -10),
        "MEA": (25, 25),
        "NAM": (-100, 60),
        "PAO": (145, 55),
        "PAS": (115, 5),
        "SAS": (75, 30),
        "WEU": (5, 50),
    }

    # Unique regions and dimensions for plotting
    regions = combined_df.index.get_level_values(0).unique()
    dimensions = combined_df.index.get_level_values(1).unique()

    # Colors for df1 using a distinct colormap
    colors_df1 = plt.cm.tab20(np.linspace(0, 1, len(dimensions)))
    colors_df2 = plt.cm.Reds(np.linspace(0.5, 1, len(dimensions)))

    # Legend handles
    legend_handles = [
        plt.Rectangle((0, 0), 1, 1, facecolor=color, edgecolor="black")
        for color in colors_df1
    ] + [plt.Rectangle((0, 0), 1, 1, facecolor="white", edgecolor="black")]
    legend_handles.append(
        Line2D([0, 1], [0, 0], color="red", marker="+", markersize=8, linewidth=0)
    )
    # Width of each bar
    width = 0.35  # Adjusted for two bars side by side

    # Transform coordinates to figure space
    trans = ax.transData.transform
    inv = fig.transFigure.inverted().transform

    # Plot bar charts at each region's location
    for region in regions:
        if region in region_locations and region in region_connections:
            # Transform region location to figure space
            x_fig, y_fig = inv(trans(region_locations[region]))

            # Create a new axis at this location
            region_ax = fig.add_axes(
                [x_fig, y_fig, 0.06, 0.125]
            )  # Adjust size as needed

            # Add a title to each region's bar plot
            region_ax.set_title(
                split_title(region_full_names.get(region)), fontsize=13, loc="center"
            )

            # Plotting
            for j, dimension in enumerate(dimensions):
                df1_value = combined_df.loc[(region, dimension), "df1_stock"]
                df2_value = combined_df.loc[(region, dimension), "df2_stock"]
                df_grandf = combined_df.loc[(region, dimension), "converge_value_cap"]

                # set negatives to zero (negatives occur in case the DLS stock threshold with practice changes is lower than DLS stocks in 2015)
                df2_value = df2_value[df2_value > 0]

                # Plot df1 value
                region_ax.bar(
                    j, df1_value, color=colors_df1[j], edgecolor="black", zorder=2
                )

                # Plot df2 value as part of the stack on top of df1
                region_ax.bar(
                    j,
                    df2_value,
                    bottom=df1_value,
                    color="white",
                    edgecolor="black",
                    zorder=2,
                )

                # plot line for converge threshold
                region_ax.plot(
                    [j, j],
                    [df_grandf, df_grandf],
                    marker="+",
                    color="red",
                    linewidth=0,
                    zorder=3,
                    markeredgewidth=1,
                )
                # region_ax.plot([j - 0.35, j + 0.35], [df_grandf, df_grandf], marker= '*', color='black', linewidth=0, zorder=3)

            # Correct y-ticks for each region's bar plot
            max_value = (
                combined_df.loc[region, combined_df.columns[:2]].sum(axis=1).max()
            )
            region_ax.set_yticks(
                np.round(np.linspace(0, max_value, 5))
            )  # Adjust number of ticks as needed
            region_ax.set_ylabel("t/cap", fontsize=14)
            region_ax.yaxis.set_tick_params(labelsize=12)

            region_ax.set_ylim(0, max_value * 1.1)

            region_ax.set_xticks([])

            # Get connection line end points
            line_end = trans(
                region_connections[region]
            )  # Country location in figure space
            line_end_fig = inv(line_end)

            # Draw first part of the line on the main map (ax)
            con_line_main = plt.Line2D(
                [line_end_fig[0], x_fig + 0.05],
                [line_end_fig[1], y_fig + 0.05],
                transform=fig.transFigure,
                color="black",
                zorder=0,  # Set zorder to 0
            )
            ax.add_line(con_line_main)  # Add the line to the main map (ax)

            # Draw second part of the line on the region axis (region_ax)
            con_line_region = plt.Line2D(
                [0, 0.05],
                [0, 0.05],
                transform=region_ax.transAxes,
                color="black",
                zorder=1,  # Set zorder to -1
            )
            region_ax.add_line(
                con_line_region
            )  # Add the line to the region axis (region_ax)

    # Corresponding labels for the legend handles
    legend_labels = [
        "..existing, health",
        "..existing, mobility",
        "..existing, nutrition",
        "..existing, shelter",
        "..existing, socialization",
    ]  # Add labels for the Rectangle objects
    legend_labels.extend(
        ["..to close DLS gaps: current practices"]
    )  # Add labels for the Line2D objects
    legend_labels.extend(
        ["..to close DLS gaps: converged practices"]
    )  # Add labels for the Line2D objects
    # Create the legend with both Rectangle and Line2D entries
    legend = ax.legend(
        handles=legend_handles,
        labels=legend_labels,
        title="DLS material stocks (Ø per capita)..",
        loc="lower left",
        bbox_to_anchor=(0.05, -0.095),
        fontsize=14,
        frameon=False,
    )
    legend._legend_box.align = "left"
    plt.setp(legend.get_title(), fontsize="14", color="black", weight="normal")
    ax.set_title(title, fontsize=30, pad=160, x=0.55)
    # Add color bar for country values under the map
    sm = plt.cm.ScalarMappable(cmap=light_blue_cmap, norm=plt.Normalize(vmin=0, vmax=1))
    cbar_ax = fig.add_axes(
        [0.17, 0.145, 0.76, 0.02]
    )  # Adjust position and size as needed
    cbar = fig.colorbar(sm, cax=cbar_ax, orientation="horizontal")
    cbar.ax.tick_params(labelsize=15)
    cbar.ax.text(
        0,
        1.8,
        "No DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
    )
    cbar.ax.text(
        0.85,
        1.8,
        "DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
    )
    cbar.ax.text(
        0.26,
        0.37,
        "DLS stocks reached = (existing DLS stocks / DLS stock threshold)",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
        color="white",
    )
    cbar.set_ticklabels([f"{int(100 * x)}%" for x in np.linspace(0, 1, 6)])
    plt.show()
    # fig.savefig('Fig1a_DLSgapMap.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    return combined_df


def plot_stacked_bars_sub_geo_converge_onlyMap(
    df1, df2, converge, subcategory, title, DLS_stock_reached
):
    # subcategories is the label of the second index columns (besides 'region)
    grouped_df1 = df1.groupby(["region", subcategory]).sum()
    grouped_df2 = df2.groupby(["region", subcategory]).sum()

    grouped_df1 = pd.DataFrame(grouped_df1.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df2 = pd.DataFrame(grouped_df2.sum(axis=1)).rename(columns={0: "stock"})

    # Creating a single dataframe for plotting
    combined_df = pd.concat(
        [grouped_df1, grouped_df2, converge], axis=1, keys=["df1", "df2", "converge"]
    )
    combined_df.columns = combined_df.columns.map("_".join)
    combined_df = combined_df / 1e3  # t/cap now

    # Load a world map
    world = DLS_stock_reached
    # Remove Antarctica
    world = world[world["continent"] != "Antarctica"]

    # Create figure and axis for the map
    fig, ax = plt.subplots(figsize=(20, 12))

    import seaborn as sns
    from matplotlib.colors import LinearSegmentedColormap

    # Get a light blue palette from seaborn
    # light_blue_palette = sns.color_palette("Blues", 8)
    light_blue_palette = sns.color_palette("coolwarm", 3)[::-1]
    light_blue_cmap = LinearSegmentedColormap.from_list(
        "light_blue", light_blue_palette
    )

    light_blue_cmap = sns.dark_palette("#69d", as_cmap=True)
    light_blue_cmap = sns.color_palette("ch:s=-.2,r=.6", as_cmap=True)
    light_blue_cmap = sns.color_palette("ch:start=.1,rot=-.15", as_cmap=True)

    # Plot the world map
    world.plot(
        column="value_cap",
        ax=ax,
        cmap=light_blue_cmap,
        legend=False,
        edgecolor="black",
        zorder=0,
    )

    # Remove plot frame (axis spines)
    for spine in ax.spines.values():
        spine.set_visible(False)

    # Remove x and y labels from the map plot
    ax.set_xticks([])
    ax.set_yticks([])

    region_locations = {
        "AFR": (-27, -70),
        "CPA": (175, 10),
        "EEU": (18, 85 + 5),
        "FSU": (85, 80 + 5),
        "LAM": (-160, -5),
        "MEA": (-50, 8),
        "NAM": (-155, 72 + 5),
        "PAO": (165, 72 + 5),
        "PAS": (165, -55),
        "SAS": (60, -65),
        "WEU": (-45, 80 + 5),
    }

    region_full_names = {
        "AFR": "Sub-Saharan Africa",
        "CPA": "Centrally planned Asia",
        "EEU": "Eastern Europe",
        "FSU": "Former Soviet Union",
        "LAM": "Latin America",
        "MEA": "N. Africa & Middle East",
        "NAM": "North America",
        "PAO": "Japan, Australia, New Zealand",
        "PAS": "Pacific Asia",
        "SAS": "South Asia",
        "WEU": "Western Europe",
    }

    region_connections = {
        "AFR": (30, 0),
        "CPA": (95, 35),
        "EEU": (23, 55),
        "FSU": (95, 70),
        "LAM": (-80, -10),
        "MEA": (25, 25),
        "NAM": (-100, 60),
        "PAO": (145, 55),
        "PAS": (115, 5),
        "SAS": (75, 30),
        "WEU": (5, 50),
    }

    # Unique regions and dimensions for plotting
    regions = combined_df.index.get_level_values(0).unique()
    dimensions = combined_df.index.get_level_values(1).unique()

    # Colors for df1 using a distinct colormap
    colors_df1 = plt.cm.tab20(np.linspace(0, 1, len(dimensions)))
    colors_df2 = plt.cm.Reds(np.linspace(0.5, 1, len(dimensions)))

    # Legend handles
    legend_handles = [
        plt.Rectangle((0, 0), 1, 1, facecolor=color, edgecolor="black")
        for color in colors_df1
    ] + [plt.Rectangle((0, 0), 1, 1, facecolor="white", edgecolor="black")]
    legend_handles.append(
        Line2D([0, 1], [0, 0], color="red", marker="+", markersize=8, linewidth=0)
    )
    # Width of each bar
    width = 0.35  # Adjusted for two bars side by side

    # Transform coordinates to figure space
    trans = ax.transData.transform
    inv = fig.transFigure.inverted().transform

    ax.set_title(title, fontsize=30, pad=160, x=0.55)
    # Add color bar for country values under the map
    sm = plt.cm.ScalarMappable(cmap=light_blue_cmap, norm=plt.Normalize(vmin=0, vmax=1))
    cbar_ax = fig.add_axes(
        [0.17, 0.145, 0.76, 0.02]
    )  # Adjust position and size as needed
    cbar = fig.colorbar(sm, cax=cbar_ax, orientation="horizontal")
    cbar.ax.tick_params(labelsize=15)
    cbar.ax.text(
        0,
        1.8,
        "No DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
    )
    cbar.ax.text(
        0.85,
        1.8,
        "DLS material stocks",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
    )
    cbar.ax.text(
        0.26,
        0.37,
        "DLS stocks reached = (existing DLS stocks / DLS stock threshold)",
        verticalalignment="center",
        transform=cbar_ax.transAxes,
        fontsize=16,
        color="white",
    )
    cbar.set_ticklabels([f"{int(100 * x)}%" for x in np.linspace(0, 1, 6)])
    plt.show()
    # fig.savefig('Fig1a_DLSgapMap.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    return combined_df


def plot_country_distribution(df, items):
    # Calculate stats and populations for the first item to determine the R11 order
    first_item_stats = (
        df.groupby("R11")[items[0]]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    # replace mean, as not population weighted above
    first_item_stats["mean"] = (
        df.set_index(["region", "R11"])[items[0]]
        * df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    first_item_stats = first_item_stats.sort_values(by="mean", ascending=False)
    populations = (
        df.groupby("R11")["population_R11"]
        .first()
        .reindex(index=first_item_stats.index)
    )
    unique_R11 = first_item_stats.index
    colors = plt.cm.get_cmap("tab20", len(unique_R11))
    color_dict = {R11: colors(i) for i, R11 in enumerate(unique_R11)}

    stats_bunker = {}
    # Create a figure with 3 subplots arranged horizontally
    fig, axs = plt.subplots(1, 3, figsize=(18, 6))

    # Define x-axis formatter for billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e9)}B"

    for i, item in enumerate(items):
        ax = axs[i]  # Select the current subplot
        stats = (
            df.groupby("R11")[item]
            .agg(["mean", "min", "max", "median"])
            .reindex(index=unique_R11)
        )
        # replace mean, as not population weighted above
        stats["mean"] = (
            df.set_index(["region", "R11"])[item]
            * df.set_index(["region", "R11"])["population_countries"]
        ).groupby("R11").sum() / df.set_index(["region", "R11"])[
            "population_countries"
        ].groupby("R11").sum()
        stats = stats.sort_values(by="mean", ascending=False)
        stats_bunker.update({item: stats})

        cumulative_population = 0  # Initialize cumulative population for each subplot
        x_positions = []  # Store x positions for setting x-ticks later

        for R11 in unique_R11:
            width = populations[R11]  # Population as bar width
            mean_value, min_value, max_value, med_value = stats.loc[R11]

            # Plot bar with color for the first item, no fill for others
            facecolor = color_dict[R11] if i == 0 else "none"
            ax.bar(
                cumulative_population + width / 2,
                mean_value,
                width=width,
                edgecolor="black",
                facecolor=facecolor,
                linewidth=2,
            )

            # Plot mean as a point and range as vertical lines for all items
            if i == 0:
                ax.plot(
                    cumulative_population + width / 2,
                    med_value,
                    "o",
                    color="red",
                    markersize=6,
                )
                ax.vlines(
                    cumulative_population + width / 2,
                    min_value,
                    max_value,
                    color="black",
                    lw=2,
                )
                ax.bar(
                    cumulative_population + width / 2,
                    mean_value,
                    width=width,
                    edgecolor="black",
                    facecolor=facecolor,
                    linewidth=2,
                )
            if i == 1:
                ax.plot(
                    cumulative_population + width / 2,
                    med_value,
                    "o",
                    color="blue",
                    markersize=6,
                )
                ax.vlines(
                    cumulative_population + width / 2,
                    min_value,
                    max_value,
                    color="blue",
                    lw=2,
                )
                ax.bar(
                    cumulative_population + width / 2,
                    mean_value,
                    width=width,
                    edgecolor="blue",
                    facecolor=facecolor,
                    linewidth=2,
                )
            if i == 2:
                ax.plot(
                    cumulative_population + width / 2,
                    med_value,
                    "o",
                    color="green",
                    markersize=6,
                )
                ax.vlines(
                    cumulative_population + width / 2,
                    min_value,
                    max_value,
                    color="green",
                    lw=2,
                )
                ax.bar(
                    cumulative_population + width / 2,
                    mean_value,
                    width=width,
                    edgecolor="green",
                    facecolor=facecolor,
                    linewidth=2,
                )

            cumulative_population += width
            x_positions.append(
                cumulative_population - width / 2
            )  # Center position for x-ticks

        ax.set_title(item, fontsize=20)
        ax.xaxis.set_major_formatter(plt.FuncFormatter(billion_formatter))
        ax.set_xticklabels([f"{pop}" for pop in [0, 0, 1, 2, 3, 4, 5, 6, 7]])
        ax.set_ylim(0, 350)
        ax.yaxis.set_tick_params(labelsize=16)
        ax.xaxis.set_tick_params(labelsize=14)

        # Adjustments specific to axis visibility
        if i == 0:
            # For the first subplot, keep y-axis and x-axis
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
        else:
            # For the second and third subplots, only keep x-axis
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_visible(False)
            ax.yaxis.set_visible(False)  # Hide y-axis labels and ticks

        # Common adjustments
        ax.spines["bottom"].set_visible(
            True
        )  # Ensure bottom spine is visible if needed
        ax.xaxis.set_tick_params(
            which="both", labelbottom=True
        )  # Ensure x-axis ticks are visible

        # Only add x-axis and y-axis labels for the first subplot
        if i == 0:
            ax.set_ylabel("tons / capita", fontsize=18)
            ax.set_xlabel("Population (billions)", fontsize=18)
        else:
            ax.set_xlabel("Population (billions)", fontsize=18)

        # Only add legend for the first item
        if i == 0:
            # Mapping from abbreviations to full names
            region_full_names = {
                "AFR": "Sub-Saharan Africa",
                "CPA": "Centrally planned Asia",
                "EEU": "Eastern Europe",
                "FSU": "Former Soviet Union",
                "LAM": "Latin America",
                "MEA": "N. Africa & Middle East",
                "NAM": "North America",
                "PAO": "Japan, Australia, New Zealand",
                "PAS": "Pacific Asia",
                "SAS": "South Asia",
                "WEU": "Western Europe",
            }
            legend_handles = [
                mpatches.Patch(color=color, label=region_full_names[R11])
                for R11, color in color_dict.items()
            ]
            fig.legend(
                handles=legend_handles,
                loc="upper right",
                fontsize=16,
                ncol=4,
                bbox_to_anchor=(0.81, -0.02),
            )

    plt.tight_layout()
    plt.show()
    # fig.savefig('Fig2b_R11stockDistr.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    return stats_bunker


def plot_country_distribution_one(df, items, title):
    # Calculate stats and populations for the first item to determine the R11 order
    first_item_stats = (
        df.groupby("R11")[items[0]]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    # replace mean, as not population weighted above
    first_item_stats["mean"] = (
        df.set_index(["region", "R11"])[items[0]]
        * df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    first_item_stats = first_item_stats.sort_values(by="mean", ascending=False)
    populations = (
        df.groupby("R11")["population_R11"]
        .first()
        .reindex(index=first_item_stats.index)
    )
    unique_R11 = (
        (
            (
                df.set_index(["region", "R11"])[items[1]]
                * df.set_index(["region", "R11"])["population_countries"]
            )
            .groupby("R11")
            .sum()
            / df.set_index(["region", "R11"])["population_countries"]
            .groupby("R11")
            .sum()
        )
        .sort_values(ascending=False)
        .index
    )
    colors = plt.cm.get_cmap("tab20", len(unique_R11))
    color_dict = {R11: colors(i) for i, R11 in enumerate(unique_R11)}

    # Create a single figure and axis
    fig, ax = plt.subplots(figsize=(22, 8))

    # Define x-axis formatter for billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e3)}B"

    cumulative_population = 0  # Initialize cumulative population
    x_positions = []  # Store x positions for setting x-ticks later

    # Create the legend
    region_full_names = {
        "AFR": "Subs.\nAfrica",
        "CPA": "Centrally \nplanned \nAsia",
        "EEU": "E. Europe",
        "FSU": "F. Soviet Union",
        "LAM": "Latin \nAmerica",
        "MEA": "N. Africa \n& Middle East",
        "NAM": "North America",
        "PAO": "JP, AU, NZ",
        "PAS": "Pacific \nAsia",
        "SAS": "South \nAsia",
        "WEU": "W. Europe",
    }

    for R11 in unique_R11:
        width = populations[R11]  # Population as bar width

        # Get the mean values for both items
        mean_value_1 = first_item_stats.loc[R11, "mean"]
        stats_item2 = (
            df.groupby("R11")[items[1]]
            .agg(["mean", "min", "max", "median"])
            .reindex(index=unique_R11)
        )
        # replace mean, as not population weighted above
        stats_item2["mean"] = (
            df.set_index(["region", "R11"])[items[1]]
            * df.set_index(["region", "R11"])["population_countries"]
        ).groupby("R11").sum() / df.set_index(["region", "R11"])[
            "population_countries"
        ].groupby("R11").sum()
        stats_item2 = stats_item2.reindex(index=unique_R11)

        mean_value_2 = stats_item2.loc[R11, "mean"] - mean_value_1

        # Plot bar for the first item
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor="#3b4cc0",
            facecolor="white",
            alpha=0.9,
            linewidth=4,
        )

        # Calculate the alpha based on the proportion of mean_value_1 in the total
        total_value = mean_value_1 + mean_value_2
        alpha_value = (
            mean_value_1 / total_value if total_value != 0 else 0
        )  # Avoid division by zero

        # Set the bar with the dynamically calculated alpha for facecolor
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor=(0.23, 0.30, 0.75, 0.5),  # Edgecolor with fixed alpha
            facecolor="#3b4cc0",  # Face color
            alpha=alpha_value * 0.4,  # Dynamic alpha based on the ratio
            linewidth=4,
        )

        # Stack the bar for the second item on top of the first
        ax.bar(
            cumulative_population + width / 2,
            mean_value_2,
            width=width,
            edgecolor="#b40426",
            facecolor="white",
            alpha=0.5,
            bottom=mean_value_1,
            linewidth=4,
        )

        # Plot statistical markers for the first item
        min_value_1 = first_item_stats.loc[R11, "min"]
        max_value_1 = first_item_stats.loc[R11, "max"]
        med_value_1 = first_item_stats.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2,
            med_value_1,
            "o",
            color="#3b4cc0",
            markersize=10,
        )
        ax.vlines(
            cumulative_population + width / 2,
            min_value_1,
            max_value_1,
            color="#3b4cc0",
            lw=4,
        )

        # Plot statistical markers for the second item
        min_value_2 = stats_item2.loc[R11, "min"]
        max_value_2 = stats_item2.loc[R11, "max"]
        med_value_2 = stats_item2.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2 + (width / 8),
            med_value_2,
            "o",
            color="#b40426",
            markersize=10,
        )
        ax.vlines(
            cumulative_population + width / 2 + (width / 8),
            min_value_2,
            max_value_2,
            color="#b40426",
            lw=4,
        )

        # Add the region label on top of the stacked bar
        # total_height = mean_value_1 + mean_value_2
        # ax.text(cumulative_population + width / 2 if R11 == 'CPA' else cumulative_population + width / 2 - (width/8), total_height + 3 if R11 == 'CPA' else total_height + 2, region_full_names[R11],
        #         ha='center', va='bottom', fontsize=16, color='black', rotation=0 if R11 == 'CPA' else 90)
        ax.text(
            cumulative_population + width / 2,
            2,
            region_full_names[R11],
            ha="center",
            va="bottom",
            fontsize=16,
            color="black",
            rotation=90,
        )

        cumulative_population += width
        x_positions.append(
            cumulative_population - width / 2
        )  # Center position for x-ticks

    ax.set_title(title, fontsize=35, pad=50, x=0.5)
    ax.set_xlabel("Population (billions)", fontsize=20)
    ax.set_ylabel("tons / capita", fontsize=20)

    ax.xaxis.set_major_formatter(plt.FuncFormatter(billion_formatter))
    ax.set_xticklabels([f"{pop}" for pop in [0, 0, 1, 2, 3, 4, 5, 6, 7]])
    ax.yaxis.set_tick_params(labelsize=18)
    ax.xaxis.set_tick_params(labelsize=18)

    ax.yaxis.set_tick_params(labelsize=20)
    ax.xaxis.set_tick_params(labelsize=20)

    # Add horizontal lines for reference
    for y in [20, 40, 60]:
        ax.axhline(y=y, color="gray", linestyle="--", linewidth=1)

    # Create the custom legend elements
    white_patch_blue_line = mpatches.Patch(
        facecolor="white",
        edgecolor="#3b4cc0",
        linewidth=4,
        label="Existing material stocks providing DLS (mean)",
    )
    white_patch_red_line = mpatches.Patch(
        facecolor="white",
        edgecolor="#b40426",
        linewidth=4,
        label="DLS material stocks gap (mean)",
    )

    line_with_marker = mlines.Line2D(
        [],
        [],
        color="black",
        marker="o",
        markersize=8,
        label="country variation within each region (min,med,max)",
    )
    # ax.set_xticklabels([f'{pop}' for pop in [0,0,0, 0, 1, 2, 3, 4, 5, 6, 7]])
    # Add the legend to the plot
    ax.legend(
        handles=[white_patch_blue_line, white_patch_red_line, line_with_marker],
        loc="upper right",
        fontsize=20,
    )

    plt.tight_layout()
    plt.show()
    # fig.savefig('Fig1b_DLSthresholds.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    plt.close("all")


def plot_country_distribution_one_converge_small(df, converge_df, items, title):
    # Calculate stats and populations for the first item to determine the R11 order
    first_item_stats = (
        df.groupby("R11")[items[0]]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    converge_stats = (
        converge_df.groupby("R11")["value_cap"]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    # replace mean, as not population weighted above
    first_item_stats["mean"] = (
        df.set_index(["region", "R11"])[items[0]]
        * df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    first_item_stats = first_item_stats.sort_values(by="mean", ascending=False)
    converge_stats["mean"] = (
        converge_df.set_index(["region", "R11"])["value_cap"]
        * converge_df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / converge_df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    converge_stats = converge_stats.sort_values(by="mean", ascending=False) / 1e3
    populations = (
        df.groupby("R11")["population_R11"]
        .first()
        .reindex(index=first_item_stats.index)
    )
    unique_R11 = (
        (
            (
                df.set_index(["region", "R11"])[items[1]]
                * df.set_index(["region", "R11"])["population_countries"]
            )
            .groupby("R11")
            .sum()
            / df.set_index(["region", "R11"])["population_countries"]
            .groupby("R11")
            .sum()
        )
        .sort_values(ascending=False)
        .index
    )
    colors = plt.cm.get_cmap("tab20", len(unique_R11))
    color_dict = {R11: colors(i) for i, R11 in enumerate(unique_R11)}

    # Create a single figure and axis
    fig, ax = plt.subplots(figsize=(22, 8))

    # Define x-axis formatter for billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e3)}B"

    cumulative_population = 0  # Initialize cumulative population
    x_positions = []  # Store x positions for setting x-ticks later

    # Create the legend
    region_full_names = {
        "AFR": "Subs.\nAfrica",
        "CPA": "Centrally \nplanned \nAsia",
        "EEU": "E. Europe",
        "FSU": "F. Soviet Union",
        "LAM": "Latin \nAmerica",
        "MEA": "N. Africa \n& Middle East",
        "NAM": "North America",
        "PAO": "JP, AU, NZ",
        "PAS": "Pacific \nAsia",
        "SAS": "South \nAsia",
        "WEU": "W. Europe",
    }

    for R11 in unique_R11:
        width = populations[R11]  # Population as bar width

        # Get the mean values for both items
        mean_value_1 = first_item_stats.loc[R11, "mean"]
        stats_item2 = (
            df.groupby("R11")[items[1]]
            .agg(["mean", "min", "max", "median"])
            .reindex(index=unique_R11)
        )
        # replace mean, as not population weighted above
        stats_item2["mean"] = (
            df.set_index(["region", "R11"])[items[1]]
            * df.set_index(["region", "R11"])["population_countries"]
        ).groupby("R11").sum() / df.set_index(["region", "R11"])[
            "population_countries"
        ].groupby("R11").sum()
        stats_item2 = stats_item2.reindex(index=unique_R11)

        mean_value_2 = stats_item2.loc[R11, "mean"] - mean_value_1

        # Plot bar for the first item
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor="#3b4cc0",
            facecolor="white",
            alpha=0.9,
            linewidth=4,
        )

        # Calculate the alpha based on the proportion of mean_value_1 in the total
        total_value = mean_value_1 + mean_value_2
        alpha_value = (
            mean_value_1 / total_value if total_value != 0 else 0
        )  # Avoid division by zero

        # Set the bar with the dynamically calculated alpha for facecolor
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor=(0.23, 0.30, 0.75, 0.5),  # Edgecolor with fixed alpha
            facecolor="#3b4cc0",  # Face color
            alpha=alpha_value * 0.4,  # Dynamic alpha based on the ratio
            linewidth=4,
        )

        # Stack the bar for the second item on top of the first
        ax.bar(
            cumulative_population + width / 2,
            mean_value_2,
            width=width,
            edgecolor="black",
            facecolor="white",
            alpha=0.6,
            bottom=mean_value_1,
            linewidth=4,
        )

        # Plot statistical markers for the first item
        min_value_1 = first_item_stats.loc[R11, "min"]
        max_value_1 = first_item_stats.loc[R11, "max"]
        med_value_1 = first_item_stats.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2,
            med_value_1,
            "o",
            color="#3b4cc0",
            markersize=10,
            alpha=0.4,
        )
        ax.vlines(
            cumulative_population + width / 2,
            min_value_1,
            max_value_1,
            color="#3b4cc0",
            lw=4,
            alpha=0.4,
        )

        # Plot statistical markers for the second item
        min_value_2 = stats_item2.loc[R11, "min"]
        max_value_2 = stats_item2.loc[R11, "max"]
        med_value_2 = stats_item2.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2 + (width / 8),
            med_value_2,
            "o",
            color="black",
            markersize=10,
            alpha=0.4,
        )
        ax.vlines(
            cumulative_population + width / 2 + (width / 8),
            min_value_2,
            max_value_2,
            color="black",
            lw=4,
            alpha=0.4,
        )

        # # Plot statistical markers for the converge threshold
        min_value_3 = converge_stats.loc[R11, "min"]
        max_value_3 = converge_stats.loc[R11, "max"]
        med_value_3 = converge_stats.loc[R11, "median"]
        mean_value_3 = converge_stats.loc[R11, "mean"]
        # ax.plot(cumulative_population + width / 2 + (width/4), med_value_3 , 'o', color='grey', markersize=10)
        ax.plot(
            cumulative_population + width / 2 + (width / 4),
            mean_value_3,
            "+",
            color="red",
            markersize=12,
            markeredgewidth=3,
            alpha=0.4,
        )
        ax.vlines(
            cumulative_population + width / 2 + (width / 4),
            min_value_3,
            max_value_3,
            color="red",
            lw=4,
            alpha=0.4,
        )

        ax.text(
            cumulative_population + width / 2,
            2,
            region_full_names[R11],
            ha="center",
            va="bottom",
            fontsize=20,
            color="black",
            rotation=90,
        )

        cumulative_population += width
        x_positions.append(
            cumulative_population - width / 2
        )  # Center position for x-ticks

    ax.set_title(title, fontsize=32, pad=25, x=0.5)
    ax.set_xlabel("Population (billions)", fontsize=20)
    ax.set_ylabel("tons / capita", fontsize=20)

    ax.xaxis.set_major_formatter(plt.FuncFormatter(billion_formatter))
    ax.set_xticklabels([f"{pop}" for pop in [0, 0, 1, 2, 3, 4, 5, 6, 7]])
    ax.yaxis.set_tick_params(labelsize=18)
    ax.xaxis.set_tick_params(labelsize=18)

    ax.yaxis.set_tick_params(labelsize=20)
    ax.xaxis.set_tick_params(labelsize=20)

    # Add horizontal lines for reference
    for y in [20, 40, 60]:
        ax.axhline(y=y, color="gray", linestyle="--", linewidth=1)

    # Create the custom legend elements
    white_patch_blue_line = mpatches.Patch(
        facecolor="white",
        edgecolor="#3b4cc0",
        alpha=0.4,
        linewidth=4,
        label="Existing DLS stocks (mean)",
    )
    white_patch_red_line = mpatches.Patch(
        facecolor="white",
        edgecolor="black",
        alpha=0.4,
        linewidth=4,
        label="DLS stock gap: current practices (mean)",
    )
    # white_patch_red_line_blue_fill = mpatches.Patch(facecolor='#3b4cc080', edgecolor='#b40426', linewidth=4, label='DLS stocks threshold (mean) below existing DLS stocks')

    line_with_marker = mlines.Line2D(
        [],
        [],
        color="black",
        marker="o",
        markersize=10,
        alpha=0.4,
        label="country variation within each region (min,med,max)",
    )
    line_with_marker_grandf = mlines.Line2D(
        [],
        [],
        color="red",
        alpha=0.4,
        marker="+",
        markersize=10,
        label="DLS stock threshold: converged practices (min,mean,max)",
    )

    legend = ax.legend(
        handles=[
            white_patch_blue_line,
            white_patch_red_line,
            line_with_marker,
            line_with_marker_grandf,
        ],
        loc="upper right",
        fontsize=20,
        ncol=1,
        title="DLS material stocks, Ø per capita (scale: countries entire population)",
        title_fontsize=20,
    )  # bbox_to_anchor=(0.05, -0.095),
    legend.get_title().set_ha("left")  # Or 'right' or 'center'

    ax.set_xlim(0 - 50000, cumulative_population + 50000)
    plt.tight_layout()
    plt.show()
    # fig.savefig('Fig1b_DLSthresholds.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    plt.close("all")



    
# def plot_country_distribution_one_converge_small(df, converge_df, items, title):

#     # Calculate stats and populations for the first item to determine the R11 order
#     first_item_stats = df.groupby('R11')[items[0]].agg(['mean','min', 'max', 'median']).sort_values(by='mean', ascending=False)
#     converge_stats = converge_df.groupby('R11')['value_cap'].agg(['mean','min', 'max', 'median']).sort_values(by='mean', ascending=False)
#     #replace mean, as not population weighted above
#     first_item_stats['mean'] = (df.set_index(['region','R11'])[items[0]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()
#     first_item_stats =  first_item_stats.sort_values(by='mean', ascending=False)
#     converge_stats['mean'] = (converge_df.set_index(['region','R11'])['value_cap'] * converge_df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / converge_df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()
#     converge_stats =  converge_stats.sort_values(by='mean', ascending=False) /1e3
#     populations = df.groupby('R11')['population_R11'].first().reindex(index=first_item_stats.index)
#     unique_R11 = ((df.set_index(['region','R11'])[items[1]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()).sort_values(ascending=False).index
#     colors = plt.cm.get_cmap('tab20', len(unique_R11))
#     color_dict = {R11: colors(i) for i, R11 in enumerate(unique_R11)}
    
#     # Create a single figure and axis
#     fig, ax = plt.subplots(figsize=(22, 8))
    
#     # Define x-axis formatter for billions
#     def billion_formatter(x, pos):
#         return f'{int(x / 1e3)}B'
    
    
#     cumulative_population = 0  # Initialize cumulative population
#     x_positions = []  # Store x positions for setting x-ticks later
    
#     # Create the legend
#     region_full_names = {
#         'AFR': 'Subs.\nAfrica',
#         'CPA': 'Centrally \nplanned \nAsia',
#         'EEU': 'E. Europe',
#         'FSU': 'F. Soviet Union',
#         'LAM': 'Latin \nAmerica',
#         'MEA': 'N. Africa \n& Middle East',
#         'NAM': 'North America',
#         'PAO': 'JP, AU, NZ',
#         'PAS': 'Pacific \nAsia',
#         'SAS': 'South \nAsia',
#         'WEU': 'W. Europe'}
    

#     for R11 in unique_R11:
#         width = populations[R11]  # Population as bar width
        
#         # Get the mean values for both items
#         mean_value_1 = first_item_stats.loc[R11, 'mean']
#         stats_item2 = df.groupby('R11')[items[1]].agg(['mean', 'min', 'max', 'median']).reindex(index=unique_R11)
#         #replace mean, as not population weighted above
#         stats_item2['mean'] = (df.set_index(['region','R11'])[items[1]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()
#         stats_item2 =  stats_item2.reindex(index=unique_R11)
        
#         mean_value_2 = stats_item2.loc[R11, 'mean'] -      mean_value_1 
        
#         # Plot bar for the first item
#         ax.bar(cumulative_population + width / 2, mean_value_1, width=width, edgecolor='#3b4cc0',
#                facecolor='white', alpha=0.9, linewidth=4)

#         # Calculate the alpha based on the proportion of mean_value_1 in the total
#         total_value = mean_value_1 + mean_value_2
#         alpha_value = mean_value_1 / total_value if total_value != 0 else 0  # Avoid division by zero

#     # Set the bar with the dynamically calculated alpha for facecolor
#         ax.bar(
#     cumulative_population + width / 2,
#     mean_value_1,
#     width=width,
#     edgecolor=(0.23, 0.30, 0.75, 0.5),  # Edgecolor with fixed alpha
#     facecolor='#3b4cc0',  # Face color
#     alpha=alpha_value*0.4,  # Dynamic alpha based on the ratio
#     linewidth=4
# )
       
#         # Stack the bar for the second item on top of the first
#         ax.bar(cumulative_population + width / 2, mean_value_2, width=width, edgecolor='black',
#                facecolor='white', alpha=0.6, bottom=mean_value_1, linewidth=4)
    
        
#         # Plot statistical markers for the first item
#         min_value_1 = first_item_stats.loc[R11, 'min']
#         max_value_1 = first_item_stats.loc[R11, 'max']
#         med_value_1 = first_item_stats.loc[R11, 'median']
#         ax.plot(cumulative_population + width / 2, med_value_1, 'o', color='#3b4cc0', markersize=10, alpha=0.4)
#         ax.vlines(cumulative_population + width / 2, min_value_1, max_value_1, color='#3b4cc0', lw=4, alpha=0.4)
        
#         # Plot statistical markers for the second item
#         min_value_2 = stats_item2.loc[R11, 'min']
#         max_value_2 = stats_item2.loc[R11, 'max']
#         med_value_2 = stats_item2.loc[R11, 'median']
#         ax.plot(cumulative_population + width / 2 + (width/8), med_value_2 , 'o', color='black', markersize=10, alpha=0.4)
#         ax.vlines(cumulative_population + width / 2+ (width/8), min_value_2 , max_value_2 ,
#                   color='black', lw=4, alpha=0.4)
        

#         # # Plot statistical markers for the converge threshold
#         min_value_3 = converge_stats.loc[R11, 'min']
#         max_value_3 = converge_stats.loc[R11, 'max']
#         med_value_3 = converge_stats.loc[R11, 'median']
#         mean_value_3 = converge_stats.loc[R11, 'mean']
#         #ax.plot(cumulative_population + width / 2 + (width/4), med_value_3 , 'o', color='grey', markersize=10)
#         ax.plot(cumulative_population + width / 2 + (width/4),   mean_value_3 , '+', color='red', markersize=12, markeredgewidth=3, alpha=0.4)
#         ax.vlines(cumulative_population + width / 2+ (width/4), min_value_3 , max_value_3 ,
#                   color='red', lw=4, alpha=0.4)
        
#         ax.text(cumulative_population + width / 2, 2, region_full_names[R11], 
#                 ha='center', va='bottom', fontsize=20, color='black', rotation= 90)
        
#         cumulative_population += width
#         x_positions.append(cumulative_population - width / 2)  # Center position for x-ticks
        
        
#     ax.set_title(title, fontsize=32, pad=25, x=0.5)
#     ax.set_xlabel('Population (billions)', fontsize=20)
#     ax.set_ylabel('tons / capita', fontsize=20)
   
    
#     ax.xaxis.set_major_formatter(plt.FuncFormatter(billion_formatter))
#     ax.set_xticklabels([f'{pop}' for pop in [0, 0, 1, 2, 3, 4, 5, 6, 7]])
#     ax.yaxis.set_tick_params(labelsize=18)
#     ax.xaxis.set_tick_params(labelsize=18)
    
#     ax.yaxis.set_tick_params(labelsize=20)
#     ax.xaxis.set_tick_params(labelsize=20)
    
#     # Add horizontal lines for reference
#     for y in [20, 40, 60]:
#         ax.axhline(y=y, color='gray', linestyle='--', linewidth=1)
        
#     # Create the custom legend elements
#     white_patch_blue_line = mpatches.Patch(facecolor='blue', edgecolor='#3b4cc0', alpha=0.3,linewidth=4, label='Existing DLS stocks (regional mean)') 
#     white_patch_red_line = mpatches.Patch(facecolor='white', edgecolor='black',alpha=0.4, linewidth=4, label='DLS stock gap: current practices (regional mean)')
#     #white_patch_red_line_blue_fill = mpatches.Patch(facecolor='#3b4cc080', edgecolor='#b40426', linewidth=4, label='DLS stocks threshold (mean) below existing DLS stocks')

#     line_with_marker = mlines.Line2D([], [], color='black', marker='o', markersize=10, alpha=0.4, label='country variation within each region (min,med,max)')
#     line_with_marker_grandf = mlines.Line2D([], [], color='red', alpha=0.4, marker='+', markersize=10, label='DLS stock threshold: converged practices (min,mean,max)')
    
#     legend = ax.legend(handles=[white_patch_blue_line, white_patch_red_line, line_with_marker, line_with_marker_grandf],
#           loc='upper right', fontsize=20,  ncol=1, title = 'DLS material stocks, Ø per capita (scale: countries entire population)', 
#           title_fontsize=20) # bbox_to_anchor=(0.05, -0.095),
#     legend.get_title().set_ha('left')  # Or 'right' or 'center'
    
    
#     ax.set_xlim(0 - 50000, cumulative_population + 50000)
#     plt.tight_layout()
#     fig.savefig("Fig_2b.pdf", bbox_inches='tight')
#     plt.show()
#     plt.close('all')



def plot_country_distribution_one_converge_small_current(df, converge_df, items, title):
    # Calculate stats and populations for the first item to determine the R11 order
    first_item_stats = (
        df.groupby("R11")[items[0]]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    converge_stats = (
        converge_df.groupby("R11")["value_cap"]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    # replace mean, as not population weighted above
    first_item_stats["mean"] = (
        df.set_index(["region", "R11"])[items[0]]
        * df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    first_item_stats = first_item_stats.sort_values(by="mean", ascending=False)
    converge_stats["mean"] = (
        converge_df.set_index(["region", "R11"])["value_cap"]
        * converge_df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / converge_df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    converge_stats = converge_stats.sort_values(by="mean", ascending=False) / 1e3
    populations = (
        df.groupby("R11")["population_R11"]
        .first()
        .reindex(index=first_item_stats.index)
    )
    unique_R11 = (
        (
            (
                df.set_index(["region", "R11"])[items[1]]
                * df.set_index(["region", "R11"])["population_countries"]
            )
            .groupby("R11")
            .sum()
            / df.set_index(["region", "R11"])["population_countries"]
            .groupby("R11")
            .sum()
        )
        .sort_values(ascending=False)
        .index
    )
    colors = plt.cm.get_cmap("tab20", len(unique_R11))
    color_dict = {R11: colors(i) for i, R11 in enumerate(unique_R11)}

    # Create a single figure and axis
    fig, ax = plt.subplots(figsize=(22, 8))

    # Define x-axis formatter for billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e3)}B"

    cumulative_population = 0  # Initialize cumulative population
    x_positions = []  # Store x positions for setting x-ticks later

    # Create the legend
    region_full_names = {
        "AFR": "Subs.\nAfrica",
        "CPA": "Centrally \nplanned \nAsia",
        "EEU": "E. Europe",
        "FSU": "F. Soviet Union",
        "LAM": "Latin \nAmerica",
        "MEA": "N. Africa \n& Middle East",
        "NAM": "North America",
        "PAO": "JP, AU, NZ",
        "PAS": "Pacific \nAsia",
        "SAS": "South \nAsia",
        "WEU": "W. Europe",
    }

    for R11 in unique_R11:
        width = populations[R11]  # Population as bar width

        # Get the mean values for both items
        mean_value_1 = first_item_stats.loc[R11, "mean"]
        stats_item2 = (
            df.groupby("R11")[items[1]]
            .agg(["mean", "min", "max", "median"])
            .reindex(index=unique_R11)
        )
        # replace mean, as not population weighted above
        stats_item2["mean"] = (
            df.set_index(["region", "R11"])[items[1]]
            * df.set_index(["region", "R11"])["population_countries"]
        ).groupby("R11").sum() / df.set_index(["region", "R11"])[
            "population_countries"
        ].groupby("R11").sum()
        stats_item2 = stats_item2.reindex(index=unique_R11)

        mean_value_2 = stats_item2.loc[R11, "mean"] - mean_value_1

        # Plot bar for the first item
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor="#3b4cc0",
            facecolor="white",
            alpha=0.9,
            linewidth=4,
        )

        # Calculate the alpha based on the proportion of mean_value_1 in the total
        total_value = mean_value_1 + mean_value_2
        alpha_value = (
            mean_value_1 / total_value if total_value != 0 else 0
        )  # Avoid division by zero

        # Set the bar with the dynamically calculated alpha for facecolor
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor=(0.23, 0.30, 0.75, 0.5),  # Edgecolor with fixed alpha
            facecolor="#3b4cc0",  # Face color
            alpha=alpha_value * 0.4,  # Dynamic alpha based on the ratio
            linewidth=4,
        )

        # Stack the bar for the second item on top of the first
        ax.bar(
            cumulative_population + width / 2,
            mean_value_2,
            width=width,
            edgecolor="black",
            facecolor="white",
            alpha=0.6,
            bottom=mean_value_1,
            linewidth=4,
        )

        # Plot statistical markers for the first item
        min_value_1 = first_item_stats.loc[R11, "min"]
        max_value_1 = first_item_stats.loc[R11, "max"]
        med_value_1 = first_item_stats.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2,
            med_value_1,
            "o",
            color="#3b4cc0",
            markersize=10,
            alpha=0.4,
        )
        ax.vlines(
            cumulative_population + width / 2,
            min_value_1,
            max_value_1,
            color="#3b4cc0",
            lw=4,
            alpha=0.4,
        )

        # Plot statistical markers for the second item
        min_value_2 = stats_item2.loc[R11, "min"]
        max_value_2 = stats_item2.loc[R11, "max"]
        med_value_2 = stats_item2.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2 + (width / 8),
            med_value_2,
            "o",
            color="black",
            markersize=10,
            alpha=0.4,
        )
        ax.vlines(
            cumulative_population + width / 2 + (width / 8),
            min_value_2,
            max_value_2,
            color="black",
            lw=4,
            alpha=0.4,
        )

        ax.text(
            cumulative_population + width / 2,
            2,
            region_full_names[R11],
            ha="center",
            va="bottom",
            fontsize=20,
            color="black",
            rotation=90,
        )

        cumulative_population += width
        x_positions.append(
            cumulative_population - width / 2
        )  # Center position for x-ticks

    ax.set_title(title, fontsize=32, pad=25, x=0.5)
    ax.set_xlabel("Population (billions)", fontsize=20)
    ax.set_ylabel("tons / capita", fontsize=20)

    ax.xaxis.set_major_formatter(plt.FuncFormatter(billion_formatter))
    ax.set_xticklabels([f"{pop}" for pop in [0, 0, 1, 2, 3, 4, 5, 6, 7]])
    ax.yaxis.set_tick_params(labelsize=18)
    ax.xaxis.set_tick_params(labelsize=18)

    ax.yaxis.set_tick_params(labelsize=20)
    ax.xaxis.set_tick_params(labelsize=20)

    # Add horizontal lines for reference
    for y in [20, 40, 60]:
        ax.axhline(y=y, color="gray", linestyle="--", linewidth=1)

    # Create the custom legend elements
    white_patch_blue_line = mpatches.Patch(
        facecolor="white",
        edgecolor="#3b4cc0",
        alpha=0.4,
        linewidth=4,
        label="Existing DLS stocks (mean)",
    )
    white_patch_red_line = mpatches.Patch(
        facecolor="white",
        edgecolor="black",
        alpha=0.4,
        linewidth=4,
        label="DLS stock gap: current practices (mean)",
    )
    # white_patch_red_line_blue_fill = mpatches.Patch(facecolor='#3b4cc080', edgecolor='#b40426', linewidth=4, label='DLS stocks threshold (mean) below existing DLS stocks')

    line_with_marker = mlines.Line2D(
        [],
        [],
        color="black",
        marker="o",
        markersize=10,
        alpha=0.4,
        label="country variation within each region (min,med,max)",
    )
    # line_with_marker_grandf = mlines.Line2D([], [], color='red', alpha=0.4, marker='+', markersize=10, label='DLS stock threshold: converged practices (min,mean,max)')

    legend = ax.legend(
        handles=[white_patch_blue_line, white_patch_red_line, line_with_marker],
        loc="upper right",
        fontsize=20,
        ncol=1,
        title="DLS material stocks, Ø per capita (scale: countries entire population)",
        title_fontsize=20,
    )  # bbox_to_anchor=(0.05, -0.095),
    legend.get_title().set_ha("left")  # Or 'right' or 'center'

    ax.set_xlim(0 - 50000, cumulative_population + 50000)
    plt.tight_layout()
    plt.show()
    # fig.savefig('Fig1b_DLSthresholds.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    plt.close("all")


def plot_country_distribution_one_withTotal(df, items, converge_df):
    # Calculate stats and populations for the first item to determine the R11 order
    first_item_stats = (
        df.groupby("R11")[items[0]]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    populations = (
        df.groupby("R11")["population_R11"]
        .first()
        .reindex(index=first_item_stats.index)
    )
    # replace mean, as not population weighted above
    first_item_stats["mean"] = (
        df.set_index(["region", "R11"])[items[0]]
        * df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    first_item_stats = first_item_stats.sort_values(by="mean", ascending=False)
    populations = (
        df.groupby("R11")["population_R11"]
        .first()
        .reindex(index=first_item_stats.index)
    )

    converge_stats = (
        converge_df.groupby("R11")["value_cap"]
        .agg(["mean", "min", "max", "median"])
        .sort_values(by="mean", ascending=False)
    )
    converge_stats["mean"] = (
        converge_df.set_index(["region", "R11"])["value_cap"]
        * converge_df.set_index(["region", "R11"])["population_countries"]
    ).groupby("R11").sum() / converge_df.set_index(["region", "R11"])[
        "population_countries"
    ].groupby("R11").sum()
    converge_stats = converge_stats.sort_values(by="mean", ascending=False) / 1e3
    # sort by height
    # unique_R11 = ((df.set_index(['region','R11'])[items[2]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()).sort_values(ascending=False).index
    # sort by preset order
    unique_R11 = [
        "CPA",
        "EEU",
        "WEU",
        "FSU",
        "NAM",
        "PAO",
        "LAM",
        "AFR",
        "MEA",
        "PAS",
        "SAS",
    ]
    colors = plt.cm.get_cmap("tab20", len(unique_R11))
    color_dict = {R11: colors(i) for i, R11 in enumerate(unique_R11)}

    # Create a single figure and axis
    fig, ax = plt.subplots(figsize=(22, 8))

    # Define x-axis formatter for billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e3)}B"

    cumulative_population = 0  # Initialize cumulative population
    x_positions = []  # Store x positions for setting x-ticks later

    # Create the legend
    region_full_names = {
        "AFR": "Subs. Africa",
        "CPA": "Centrally planned Asia",
        "EEU": "E. Europe",
        "FSU": "Former Soviet Union",
        "LAM": "Latin America",
        "MEA": "N. Africa & Middle East",
        "NAM": "North America",
        "PAO": "JP, AU, NZ",
        "PAS": "Pacific Asia",
        "SAS": "South Asia",
        "WEU": "W. Europe",
    }

    for R11 in unique_R11:
        width = populations[R11]  # Population as bar width

        # Get the mean values for all three items
        mean_value_1 = first_item_stats.loc[R11, "mean"]
        stats_item2 = (
            df.groupby("R11")[items[1]]
            .agg(["mean", "min", "max", "median"])
            .reindex(index=unique_R11)
        )
        # replace mean, as not population weighted above
        stats_item2["mean"] = (
            df.set_index(["region", "R11"])[items[1]]
            * df.set_index(["region", "R11"])["population_countries"]
        ).groupby("R11").sum() / df.set_index(["region", "R11"])[
            "population_countries"
        ].groupby("R11").sum()
        stats_item2 = stats_item2.reindex(index=unique_R11)
        mean_value_2 = stats_item2.loc[R11, "mean"] - mean_value_1

        stats_item3 = (
            df.groupby("R11")[items[2]]
            .agg(["mean", "min", "max", "median"])
            .reindex(index=unique_R11)
        )
        # replace mean, as not population weighted above
        stats_item3["mean"] = (
            df.set_index(["region", "R11"])[items[2]]
            * df.set_index(["region", "R11"])["population_countries"]
        ).groupby("R11").sum() / df.set_index(["region", "R11"])[
            "population_countries"
        ].groupby("R11").sum()
        stats_item3 = stats_item3.reindex(index=unique_R11)
        mean_value_3 = stats_item3.loc[R11, "mean"] - mean_value_1 - mean_value_2

        # Plot bar for the first item
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor="#3b4cc0",
            facecolor="white",
            linewidth=4,
        )

        # Calculate the alpha based on the proportion of mean_value_1 in the total
        total_value = mean_value_1 + mean_value_2
        alpha_value = (
            mean_value_1 / total_value if total_value != 0 else 0
        )  # Avoid division by zero

        # Set the bar with the dynamically calculated alpha for facecolor
        ax.bar(
            cumulative_population + width / 2,
            mean_value_1,
            width=width,
            edgecolor=(0.23, 0.30, 0.75, 0.5),  # Edgecolor with fixed alpha
            facecolor="#3b4cc0",  # Face color
            alpha=alpha_value * 0.4,  # Dynamic alpha based on the ratio
            linewidth=4,
        )

        # Stack the bar for the second item on top of the first
        ax.bar(
            cumulative_population + width / 2,
            mean_value_2,
            width=width,
            edgecolor="black",
            facecolor="white",
            alpha=0.8,
            bottom=mean_value_1,
            linewidth=4,
        )

        # Stack the bar for the third item on top of the second
        ax.bar(
            cumulative_population + width / 2,
            mean_value_3,
            width=width,
            edgecolor="black",
            hatch=".",
            facecolor="white",
            alpha=0.6,
            bottom=mean_value_1 + mean_value_2,
            linewidth=4,
        )

        # Plot statistical markers for the first item
        min_value_1 = first_item_stats.loc[R11, "min"]
        max_value_1 = first_item_stats.loc[R11, "max"]
        med_value_1 = first_item_stats.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2,
            med_value_1,
            "o",
            color="#3b4cc0",
            markersize=10,
            alpha=0.4,
        )

        ax.vlines(
            cumulative_population + width / 2,
            min_value_1,
            max_value_1,
            color="#3b4cc0",
            lw=4,
            alpha=0.4,
        )

        # Plot statistical markers for the second item
        min_value_2 = stats_item2.loc[R11, "min"]
        max_value_2 = stats_item2.loc[R11, "max"]
        med_value_2 = stats_item2.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2 + (width / 8),
            med_value_2,
            "o",
            color="black",
            markersize=10,
            alpha=0.4,
        )
        ax.vlines(
            cumulative_population + width / 2 + (width / 8),
            min_value_2,
            max_value_2,
            color="black",
            lw=4,
            alpha=0.4,
        )

        # Plot statistical markers for the third item
        min_value_3 = stats_item3.loc[R11, "min"]
        max_value_3 = stats_item3.loc[R11, "max"]
        med_value_3 = stats_item3.loc[R11, "median"]
        ax.plot(
            cumulative_population + width / 2 + (width / 4),
            med_value_3,
            "o",
            color="black",
            markersize=10,
            alpha=0.4,
        )
        ax.vlines(
            cumulative_population + width / 2 + (width / 4),
            min_value_3,
            max_value_3,
            color="black",
            lw=4,
            linestyle=":",
            alpha=0.8,
        )

        cumulative_population += width
        x_positions.append(
            cumulative_population - width / 2
        )  # Center position for x-ticks

    # ax.set_title("Country variation of material stocks providing existing DLS-levels, to close DLS gaps & beyond-DLS", fontsize=32, pad=50, x=0.5)
    ax.set_xlabel("Population (billions)", fontsize=20)
    ax.set_ylabel("tons / capita", fontsize=20)

    ax.xaxis.set_major_formatter(plt.FuncFormatter(billion_formatter))
    ax.set_xticklabels([f"{pop}" for pop in [0, 0, 1, 2, 3, 4, 5, 6, 7]])
    ax.yaxis.set_tick_params(labelsize=18)
    ax.xaxis.set_tick_params(labelsize=18)

    ax.yaxis.set_tick_params(labelsize=20)
    ax.xaxis.set_tick_params(labelsize=20)

    # Add horizontal lines for reference
    for y in [50, 100, 150, 200, 250, 300]:
        ax.axhline(y=y, color="gray", linestyle="--", linewidth=1)

    # Create the custom legend elements
    white_patch_blue_line = mpatches.Patch(
        facecolor="white",
        edgecolor="#3b4cc0",
        linewidth=4,
        alpha=0.4,
        label="Existing DLS stocks (mean)",
    )
    white_patch_red_line = mpatches.Patch(
        facecolor="white",
        edgecolor="black",
        linewidth=4,
        alpha=0.4,
        label="DLS stock gap: current practices (mean)",
    )
    white_patch_green_line = mpatches.Patch(
        facecolor="white",
        edgecolor="black",
        linewidth=4,
        hatch=".",
        alpha=0.4,
        label="Difference to existing economy-wide stocks (mean)",
    )

    line_with_marker = mlines.Line2D(
        [],
        [],
        color="black",
        linestyle=":",
        marker="o",
        markersize=8,
        label="Country variation within each region (min, med, max)",
    )

    legend = ax.legend(
        handles=[
            white_patch_blue_line,
            white_patch_red_line,
            white_patch_green_line,
            line_with_marker,
        ],
        loc="upper right",
        fontsize=20,
        ncol=1,
        title="Material stocks, Ø per capita (scale: countries entire population)",
        title_fontsize=20,
    )

    legend.get_title().set_ha("left")  # Or 'right' or 'center'

    # ax.set_title("(c) Country variation of material stocks for DLS\n against economy-wide stocks", fontsize=28, pad=50, x=0.5)
    ax.set_title(
        "(c) Region & country variation of DLS material stocks against economy-wide stocks",
        fontsize=32,
        pad=25,
        x=0.5,
    )
    ax.set_xlim(0 - 50000, cumulative_population + 50000)
    ax.set_ylim(0, 350)
    plt.tight_layout()
    plt.show()
    plt.close("all")
    
    
    
    
    

# def plot_country_distribution_one_withTotal(df, items, converge_df):
#     # Calculate stats and populations for the first item to determine the R11 order
#     first_item_stats = df.groupby('R11')[items[0]].agg(['mean', 'min', 'max', 'median']).sort_values(by='mean', ascending=False)
#     populations = df.groupby('R11')['population_R11'].first().reindex(index=first_item_stats.index)
#     #replace mean, as not population weighted above
#     first_item_stats['mean'] = (df.set_index(['region','R11'])[items[0]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()
#     first_item_stats =  first_item_stats.sort_values(by='mean', ascending=False)
#     populations = df.groupby('R11')['population_R11'].first().reindex(index=first_item_stats.index)
    
#     converge_stats = converge_df.groupby('R11')['value_cap'].agg(['mean','min', 'max', 'median']).sort_values(by='mean', ascending=False)
#     converge_stats['mean'] = (converge_df.set_index(['region','R11'])['value_cap'] * converge_df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / converge_df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()
#     converge_stats =  converge_stats.sort_values(by='mean', ascending=False) /1e3
#     # sort by height
#     #unique_R11 = ((df.set_index(['region','R11'])[items[2]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()).sort_values(ascending=False).index
#     # sort by preset order
#     unique_R11 = ['CPA','EEU','WEU','FSU','NAM','PAO','LAM','AFR','MEA','PAS','SAS']
#     colors = plt.cm.get_cmap('tab20', len(unique_R11))
#     color_dict = {R11: colors(i) for i, R11 in enumerate(unique_R11)}
    
#     # Create a single figure and axis
#     fig, ax = plt.subplots(figsize=(22, 8))
    
#     # Define x-axis formatter for billions
#     def billion_formatter(x, pos):
#         return f'{int(x / 1e3)}B'
    
#     cumulative_population = 0  # Initialize cumulative population
#     x_positions = []  # Store x positions for setting x-ticks later
    
#     # Create the legend
#     region_full_names = {
#         'AFR': 'Subs. Africa',
#         'CPA': 'Centrally planned Asia',
#         'EEU': 'E. Europe',
#         'FSU': 'Former Soviet Union',
#         'LAM': 'Latin America',
#         'MEA': 'N. Africa & Middle East',
#         'NAM': 'North America',
#         'PAO': 'JP, AU, NZ',
#         'PAS': 'Pacific Asia',
#         'SAS': 'South Asia',
#         'WEU': 'W. Europe'
#     }

#     for R11 in unique_R11:
#         width = populations[R11]  # Population as bar width

#         # Get the mean values for all three items
#         mean_value_1 = first_item_stats.loc[R11, 'mean']
#         stats_item2 = df.groupby('R11')[items[1]].agg(['mean', 'min', 'max', 'median']).reindex(index=unique_R11)
#         #replace mean, as not population weighted above
#         stats_item2['mean'] = (df.set_index(['region','R11'])[items[1]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()
#         stats_item2 =  stats_item2.reindex(index=unique_R11)
#         mean_value_2 = stats_item2.loc[R11, 'mean'] - mean_value_1
        
#         stats_item3 = df.groupby('R11')[items[2]].agg(['mean', 'min', 'max', 'median']).reindex(index=unique_R11)
#         #replace mean, as not population weighted above
#         stats_item3['mean'] = (df.set_index(['region','R11'])[items[2]] * df.set_index(['region','R11'])['population_countries']).groupby('R11').sum() / df.set_index(['region','R11'])['population_countries'].groupby('R11').sum()
#         stats_item3 =  stats_item3.reindex(index=unique_R11)
#         mean_value_3 = stats_item3.loc[R11, 'mean'] - mean_value_1 - mean_value_2
        
#         # Plot bar for the first item
#         ax.bar(cumulative_population + width / 2, mean_value_1, width=width, edgecolor='#3b4cc0',
#                facecolor='white', linewidth=4)
        
#         # Calculate the alpha based on the proportion of mean_value_1 in the total
#         total_value = mean_value_1 + mean_value_2
#         alpha_value = mean_value_1 / total_value if total_value != 0 else 0  # Avoid division by zero

#             # Set the bar with the dynamically calculated alpha for facecolor
#         ax.bar(
#     cumulative_population + width / 2,
#     mean_value_1,
#     width=width,
#     edgecolor=(0.23, 0.30, 0.75, 0.5),  # Edgecolor with fixed alpha
#     facecolor='#3b4cc0',  # Face color
#     alpha=alpha_value*0.4,  # Dynamic alpha based on the ratio
#     linewidth=4
# )


#         # Stack the bar for the second item on top of the first
#         ax.bar(cumulative_population + width / 2, mean_value_2, width=width, edgecolor='black',
#                facecolor='white', alpha=0.8, bottom=mean_value_1, linewidth=4)
        
#         # Stack the bar for the third item on top of the second
#         ax.bar(cumulative_population + width / 2, mean_value_3, width=width, edgecolor="black", hatch= '.',
#                facecolor='white',  alpha=0.6, bottom=mean_value_1 + mean_value_2, linewidth=4)
        
#         # Plot statistical markers for the first item
#         min_value_1 = first_item_stats.loc[R11, 'min']
#         max_value_1 = first_item_stats.loc[R11, 'max']
#         med_value_1 = first_item_stats.loc[R11, 'median']
#         ax.plot(cumulative_population + width / 2, med_value_1, 'o', color='#3b4cc0', markersize=10, alpha=0.4)
    
#         ax.vlines(cumulative_population + width / 2, min_value_1, max_value_1, color='#3b4cc0', lw=4, alpha=0.4)
        
        
#         # Plot statistical markers for the second item
#         min_value_2 = stats_item2.loc[R11, 'min']
#         max_value_2 = stats_item2.loc[R11, 'max']
#         med_value_2 = stats_item2.loc[R11, 'median']
#         ax.plot(cumulative_population + width / 2 + (width / 8), med_value_2, 'o', color='black', markersize=10, alpha=0.4)
#         ax.vlines(cumulative_population + width / 2 + (width / 8), min_value_2, max_value_2, color='black', lw=4, alpha=0.4)
        
#         # Plot statistical markers for the third item
#         min_value_3 = stats_item3.loc[R11, 'min']
#         max_value_3 = stats_item3.loc[R11, 'max']
#         med_value_3 = stats_item3.loc[R11, 'median']
#         ax.plot(cumulative_population + width / 2 + (width / 4), med_value_3, 'o', color='black', markersize=10, alpha=0.4)
#         ax.vlines(cumulative_population + width / 2 + (width / 4), min_value_3, max_value_3, color='black', lw=4, linestyle = ':', alpha=0.8)
        
#         region_full_names_to_plot = {
#             'AFR': 'Subs.Africa',
#             'CPA': 'Centrally planned Asia',
#             'EEU': 'E.\nEU',
#             'FSU': 'F.Sov.\nUnion',
#             'LAM': 'L. America',
#             'MEA': 'N. Africa &\nM. East',
#             'NAM': 'N.Am-\nerica',
#             'PAO': 'JP, \nAU, NZ',
#             'PAS': 'Pacific\nAsia',
#             'SAS': 'South Asia',
#             'WEU': 'W.\nEU'
#             }
        
#         region_names_position = {
#             'AFR': -20,
#             'CPA': 4.5,
#             'EEU': -35,
#             'FSU': -35,
#             'LAM': 5.5,
#             'MEA': -35,
#             'NAM': 6.5,
#             'PAO': -35,
#             'PAS': -35,
#             'SAS': -20,
#             'WEU': 7.5
#             }
        
#         ax.text(cumulative_population + width / 2, region_names_position[R11], region_full_names_to_plot[R11],  
#                 ha='center', va='bottom', fontsize=20, color='black', rotation= 0)
        
#         cumulative_population += width
#         x_positions.append(cumulative_population - width / 2)  # Center position for x-ticks
    
#     #ax.set_title("Country variation of material stocks providing existing DLS-levels, to close DLS gaps & beyond-DLS", fontsize=32, pad=50, x=0.5)

#     ax.set_xlabel('Population\n(billions)', fontsize=20, labelpad=15)
#     ax.xaxis.set_label_coords(0.05, -0.05)  
#     ax.set_ylabel('tons / capita', fontsize=20)
    
#     ax.xaxis.set_major_formatter(plt.FuncFormatter(billion_formatter))
#     ax.set_xticklabels([f'{pop}' for pop in [0, 0, '', '', '', '', '', '', 7]])
#     ax.yaxis.set_tick_params(labelsize=18)
#     ax.xaxis.set_tick_params(labelsize=18)
    
#     ax.yaxis.set_tick_params(labelsize=20)
#     ax.xaxis.set_tick_params(labelsize=20)
    
#     # Turn off tick lines for positions with empty labels
#     for tick, label in zip(ax.xaxis.get_major_ticks(), ax.get_xticklabels()):
#         if label.get_text() == '':
#             tick.tick1line.set_visible(False)  # bottom tick line
#             tick.tick2line.set_visible(False)  # top tick line

    
#     # Add horizontal lines for reference
#     for y in [50, 100, 150, 200, 250, 300]:
#         ax.axhline(y=y, color='gray', linestyle='--', linewidth=1)
        
#     # Create the custom legend elements
#     white_patch_blue_line = mpatches.Patch(facecolor='blue', edgecolor='#3b4cc0', linewidth=4, alpha=0.3, label='Existing DLS stocks (regional mean)')
#     white_patch_red_line = mpatches.Patch(facecolor='white', edgecolor='black', linewidth=4, alpha=0.4, label='DLS stock gap: current practices (regional mean)')
#     white_patch_green_line = mpatches.Patch(facecolor='white', edgecolor='black', linewidth=4, hatch = '.',  alpha=0.4,label='Difference to existing economy-wide stocks (regional mean)')

#     line_with_marker = mlines.Line2D([], [], color='black', linestyle = ':', marker='o', markersize=8, label='Country variation within each region (min, med, max)')
    
#     legend = ax.legend(handles=[white_patch_blue_line, white_patch_red_line, white_patch_green_line, line_with_marker], 
#       loc='upper right', fontsize=20, ncol=1, title='Material stocks, Ø per capita (scale: countries entire population)',
#       title_fontsize=20)
    
#     legend.get_title().set_ha('left')  # Or 'right' or 'center'
    
    
#     ax.set_title("(c) Region & country variation of DLS material stocks against economy-wide stocks", fontsize=32, pad=25, x=0.5)
#     ax.set_xlim(0 - 50000, cumulative_population + 50000)
#     ax.set_ylim(0,300)
#     plt.tight_layout()
#     #fig.savefig("Fig_2c.pdf", bbox_inches='tight')
#     plt.show()
#     plt.close('all')
    
    
    
    
    


def plot_bars_horiz_gap_headroom_two_subplots_doubleGlob_mod(
    df1,
    df2,
    df3,
    df_add_beyond_stocks_regional,
    df_add_beyond_stocks_global,
    global_stocks,
    global_stock_prod,
    subcategory,
    line_y_values,
    line_labels,
):
    grouped_df1 = df1.groupby(["region", subcategory]).sum()
    grouped_df2 = df2.groupby(["region", subcategory]).sum()
    grouped_df3 = df3.set_index(["region", subcategory]).groupby("region").sum()

    grouped_df1 = pd.DataFrame(grouped_df1.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df2 = pd.DataFrame(grouped_df2.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df3 = pd.DataFrame(grouped_df3.sum(axis=1)).rename(columns={0: "line"})

    # Creating a single dataframe for plotting
    combined_df = pd.concat([grouped_df1, grouped_df2], axis=1, keys=["df1", "df2"])
    combined_df.columns = combined_df.columns.map("_".join)
    combined_df = combined_df / 1e3  # Convert to t/cap
    combined_df.reset_index(inplace=True)

    region_order = [
        "EEU",
        "WEU",
        "PAO",
        "NAM",
        "FSU",
        "CPA",
        "MEA",
        "LAM",
        "PAS",
        "SAS",
        "AFR",
    ]
    sector_order = [
        "res_buildings",
        "nonres_buildings",
        "other_construction",
        "road_rail",
        "transport_machinery",
        "machinery",
        "other",
    ]

    # Create a mapping of region to the custom order
    order_mapping = {region: i for i, region in enumerate(region_order)}
    # Create a mapping of sector to the custom order
    order_mapping_sector = {sector: i for i, sector in enumerate(sector_order)}

    # Map the 'region' column to a new column that represents its order
    combined_df["region_order"] = combined_df["region"].map(order_mapping)
    # Sort the DataFrame by the new 'region_order' column, then drop it

    # Map the 'sector' column to a new column that represents its order
    combined_df["sector_order"] = combined_df["sector"].map(order_mapping_sector)
    # Sort the DataFrame by the new 'sector_order' column, then drop it
    combined_df_sorted = combined_df.sort_values(
        by=["region_order", "sector_order"]
    ).drop(columns=["sector_order"])

    combined_df = combined_df_sorted.set_index(["region", "sector"])

    # also reorder global_stock_prod and global_stock
    global_stock_prod = global_stock_prod.reindex(sector_order)
    global_stocks = global_stocks.reindex(["minerals", "metals", "biomass", "fossils"])

    # Create the figure
    fig = plt.figure(figsize=(24, 24))
    gs = gridspec.GridSpec(2, 1, height_ratios=[2.5, 4])
    ax0 = fig.add_subplot(gs[0])
    ax2 = fig.add_subplot(gs[1])

    global_stocks_products = global_stock_prod.groupby("region", axis=1).sum()

    from matplotlib.patches import FancyBboxPatch

    # Materials and their colors
    materials = global_stocks.index
    sectors = global_stocks_products.index
    colors = [
        (0.6, 0.58, 0.58),
        (0.18, 0.46, 0.71),
        (0.44, 0.68, 0.28),
        (0.93, 0.40, 0.19),
    ]
    colors2 = plt.cm.get_cmap("Pastel2", len(sectors))
    # replace grey color in pastel2 so it is not confuesd with grey in variable colors
    colors_array = colors2(np.arange(len(sectors)))
    replacement_color = [1.0, 0.5, 0.5, 1.0]  # Replace with a red-like RGBA
    colors_array[-1] = replacement_color  # Replace gray with your chosen color
    from matplotlib.colors import ListedColormap

    custom_cmap = ListedColormap(colors_array)
    colors2 = custom_cmap
    # rearrange color order:
    # Extract the colors as an array
    colors3 = colors2(np.arange(colors2.N))
    # Define your custom order for the colors (example: swap the first and last color)
    custom_order = [4, 1, 6, 5, 3, 0, 2]  # Adjust based on the desired order
    # Reorder the colors based on the custom order
    reordered_colors = colors3[custom_order]
    # Create a new colormap with the reordered colors
    colors2 = ListedColormap(reordered_colors)

    # Adjust hatching patterns for keys
    key_hatch = {"DLS_prov": "", "beyond_DLS": "."}

    double_bar = 0.35

    # Function to create rounded horizontal bars
    def plot_rounded_barh(ax, y, width, left=0, color=None, hatch=None, alpha=None):
        boxstyle = "round,pad=0"
        bar_height = 0.275  # Consistent bar height across all bars

        # Draw an outline with a fully opaque black edge
        outline_bbox = FancyBboxPatch(
            (left, y - bar_height),
            width,
            bar_height * 2,
            boxstyle=boxstyle,
            ec="black",  # Opaque black edge
            fc="none",  # No face color for the outline
            hatch=hatch,  # Hatch pattern if provided
            lw=1.5,  # Line width for the edge
        )
        ax.add_patch(outline_bbox)

        # Draw the fill with the specified transparency (alpha)
        if color is not None:  # Only add fill if color is provided
            fill_bbox = FancyBboxPatch(
                (left, y - bar_height),
                width,
                bar_height * 2,
                boxstyle=boxstyle,
                ec="none",  # No edge for the fill
                fc=color,  # Face color with transparency
                hatch=hatch,  # Hatch pattern if provided
                lw=0,  # No line width for the fill
                alpha=alpha,  # Transparency for the fill
            )
            ax.add_patch(fill_bbox)

    # Plot bars for global stocks of products
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, sector in enumerate(sectors):
            value = global_stocks_products.loc[sector, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barh(
                    ax0,
                    -double_bar,
                    value,
                    left=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barh(
                    ax0,
                    -double_bar,
                    value,
                    left=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barh(
                    ax0,
                    -double_bar,
                    value,
                    left=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barh(
                    ax0,
                    -double_bar,
                    value,
                    left=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )  # Changed y=1 to y=0.5 for central alignment
            bottom += value

    # Plot bars for global stocks of materials
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, material in enumerate(materials):
            value = global_stocks.loc[material, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barh(
                    ax0,
                    double_bar,
                    value,
                    left=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barh(
                    ax0,
                    double_bar,
                    value,
                    left=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barh(
                    ax0,
                    double_bar,
                    value,
                    left=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barh(
                    ax0,
                    double_bar,
                    value,
                    left=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
                # Changed y=0 to y=-0.5 for central alignment
            bottom += value

    bottom_save = bottom
    # Add explicit legend entries for materials
    handles = [
        FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.1", fc=color, ec="black")
        for color in colors
    ]
    labels = list(materials)  # Convert to a list

    # Plot the gap keys
    keys_to_plot_target = ["gap_targeted", "gap_regional", "gap_trickle"]
    letters = ["i", "ii", "iii"]

    for idx, key in enumerate(keys_to_plot_target):
        value = global_stocks[key].sum()
        plot_rounded_barh(
            ax0, double_bar, value, left=bottom, color="white", hatch="None"
        )  # Changed y=0 to y=-0.5 for central alignment
        plot_rounded_barh(ax0, double_bar, value, left=bottom, color=None, hatch=None)
        ax0.text(
            bottom + value / 2,
            double_bar,
            letters[idx],
            ha="center",
            va="center",
            fontsize=24,
        )
        bottom += value

    keys_to_plot_all = ["gap_targeted", "gap_regional", "gap_trickle"]
    letters = ["i", "ii", "iii"]
    bottom = bottom_save
    for idx, key in enumerate(keys_to_plot_all):
        value = global_stocks[key].sum()
        plot_rounded_barh(
            ax0, -double_bar, value, left=bottom, color="white", hatch=None
        )
        plot_rounded_barh(
            ax0, -double_bar, value, left=bottom, color=None, hatch=None
        )  # Changed y=0 to y=-0.5 for central alignment
        ax0.text(
            bottom + value / 2,
            -double_bar,
            letters[idx],
            ha="center",
            va="center",
            fontsize=24,
        )
        bottom += value

    ax0.set_yticks([double_bar, -double_bar])
    ax0.set_yticklabels(
        [
            "Global, BY MATERIAL\n only build DLS\n stocks to close gap",
            "Global, BY PRODUCT\n also build beyond-DLS\n stocks to close gap",
        ],
        fontsize=24,
    )

    # Function to draw brackets with labels
    def draw_bracket_with_label(ax, left, right, y, height=0.1, label=""):
        ax.plot([left, left], [y, y + height], color="black")  # Left vertical line
        ax.plot([right, right], [y, y + height], color="black")  # Right vertical line
        ax.plot(
            [left, right], [y + height, y + height], color="black"
        )  # Connecting horizontal line
        ax.text(
            (left + right) / 2,
            y + height + 0.05,
            label,
            ha="center",
            va="bottom",
            fontsize=22,
        )

    def draw_bracket_with_label_low(ax, left, right, y, height=0.075, label=""):
        # Draw the bracket lines
        ax.plot(
            [left, left], [y, y - height], color="black"
        )  # Left vertical line, extending downwards
        ax.plot(
            [right, right], [y, y - height], color="black"
        )  # Right vertical line, extending downwards
        ax.plot(
            [left, right], [y - height, y - height], color="black"
        )  # Connecting horizontal line, at the bottom
        # Add label below the bracket
        ax.text(
            (left + right) / 2,
            y - height - 0.04,
            label,
            ha="left",
            va="top",
            fontsize=20,
        )  # Adjusted to put label below the bracket

    def draw_bracket_with_label_low_center(ax, left, right, y, height=0.075, label=""):
        # Draw the bracket lines
        ax.plot(
            [left, left], [y, y - height], color="black"
        )  # Left vertical line, extending downwards
        ax.plot(
            [right, right], [y, y - height], color="black"
        )  # Right vertical line, extending downwards
        ax.plot(
            [left, right], [y - height, y - height], color="black"
        )  # Connecting horizontal line, at the bottom
        # Add label below the bracket
        ax.text(
            (left + right) / 2,
            y - height - 0.1,
            label,
            ha="center",
            va="center",
            fontsize=20,
        )  # Adjusted to put label below the bracket

    def draw_bracket_with_label_low_low_center(
        ax, left, right, y, height=0.04, label=""
    ):
        # Draw the bracket lines
        ax.plot(
            [left, left], [y, y - height], color="black"
        )  # Left vertical line, extending downwards
        ax.plot(
            [right, right], [y, y - height], color="black"
        )  # Right vertical line, extending downwards
        ax.plot(
            [left, right], [y - height, y - height], color="black"
        )  # Connecting horizontal line, at the bottom
        # Add label below the bracket
        ax.text(
            (left + right) / 2,
            y - height - 0.15,
            label,
            ha="center",
            va="center",
            fontsize=20,
        )  # Adjusted to put label below the bracket

    # Draw brackets for each key
    y_bracket_position = 0.68  # Adjust as needed for visual clarity
    y_bracket_position_below = -0.67
    draw_bracket_with_label(
        ax0,
        0.9e10,
        2.092280e11 - 1.2e10,
        y_bracket_position,
        label="existing \n DLS stocks",
    )
    draw_bracket_with_label(
        ax0,
        2.092280e11 + 0.3e10,
        2.092280e11 + 4.880719e11 - 1e10,
        y_bracket_position,
        label="existing beyond-DLS stocks",
    )
    draw_bracket_with_label(
        ax0,
        2.092280e11 + 4.880719e11 + 0.5e10,
        1285591129139.884 - 1e10 + 12e10,
        y_bracket_position,
        label="stock additions to close DLS gap",
    )
    draw_bracket_with_label_low_low_center(
        ax0,
        2.092280e11 + 4.880719e11 + 0.8e10,
        2.092280e11 + 4.880719e11 + 7.866508e10 - 0.8e10,
        y_bracket_position_below,
        label="build\n DLS\n only",
    )
    draw_bracket_with_label_low_low_center(
        ax0,
        2.092280e11 + 4.880719e11 + 7.866508e10 - 1e10 + 2.6e10,
        957518380000.0 - 3.2e10,
        y_bracket_position_below,
        label="also build some\nbeyond-DLS:\n regional ratio",
    )
    draw_bracket_with_label_low_low_center(
        ax0,
        957518380000.0 - 1.3e10,
        957518380000.0 + 3.280728e11 - 1e10 + 12e10,
        y_bracket_position_below,
        label="also build many beyond-DLS:\n global trickle-down",
    )
    draw_bracket_with_label_low_low_center(
        ax0,
        0.9e10,
        2.092280e11 + 4.880719e11 - 1e10,
        y_bracket_position_below,
        label="existing\neconomy-wide stocks",
    )

    # Set axis limits and hide unnecessary parts
    ax0.set_xlim(0, bottom)
    ax0.set_ylim(-1, 1)

    # Set plot labels and title and tick params
    ax0.tick_params(axis="x", labelsize=22)
    ax0.tick_params(axis="y", labelsize=22)
    # Remove offset from x-axis
    from matplotlib.ticker import ScalarFormatter

    ax0.xaxis.set_major_formatter(ScalarFormatter(useOffset=False))
    x_ticks = ax0.get_xticks()
    x_labels = [int(x / 1e9) for x in x_ticks]
    ax0.set_xticklabels(x_labels)
    ax0.set_xlabel("[Gigatons = 10^9 tons]", fontsize=24, labelpad=20)
    ax0.set_title("..(a) GLOBAL, at scale", fontsize=36, loc="left", pad=35)

    # Unique regions and dimensions for plotting
    regions = combined_df.index.get_level_values(0).unique()
    dimensions = combined_df.index.get_level_values(1).unique()

    # Assigning distinct colors and hatch patterns to each dimension and creating legend handles
    colors = plt.cm.get_cmap("Pastel2", len(dimensions))
    colors = colors2
    hatches = ["", "", "", "", "", "", "", "", "", ""]
    hatches = [".", ".", ".", ".", ".", ".", ".", ".", ".", "."]
    legend_handles = []

    # Calculate cumulative population
    cumulative_population = np.cumsum(
        [0]
        + [
            121425.83099999999,
            497389.066,
            156532.16499999998,
            360286.49199999997,
            288469.298,
            1557155.8020000001,
            458693.972,
            618764.293,
            579768.6460000001,
            1749355.091,
            954919.91,
        ]
    )

    # Sort regions
    regions = combined_df.index.get_level_values(0).unique()

    # Plot each region
    for i, region in enumerate(regions):
        left_df1 = 0  # Left for df1 stack
        left_df2 = 0  # Left for df2 stack (starts at right of df1 stack)
        left_df3 = 0
        left_df4 = 0

        # Plot df1 and calculate total df1 value for the region
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df1_value = combined_df.loc[(region, dimension), "df1_stock"]
            ax2.barh(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df1_value,
                (cumulative_population[i + 1] - cumulative_population[i]),
                color=color,
                edgecolor="black",
                linewidth=1.5,
                label=f"{dimension} (df1)",
                left=left_df1,
                alpha=1,
            )
            left_df1 += df1_value

            # Add legend handle for this dimension if it's the first region
            if i == 0:
                legend_handles.append(
                    plt.Rectangle((0, 0), 1, 1, color=color, alpha=1, edgecolor="black")
                )

        # Plot df2 stacked by dimension starting slightly to the right of df1 for the region
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df2_value = combined_df.loc[(region, dimension), "df2_stock"]
            ax2.barh(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                (cumulative_population[i + 1] - cumulative_population[i]),
                color=color,
                edgecolor="none",
                linewidth=1.5,
                hatch=hatch,
                label=f"{dimension} (df2)",
                left=left_df1 + left_df2,
                alpha=0.6,
            )
            ax2.barh(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                (cumulative_population[i + 1] - cumulative_population[i]),
                color="none",
                edgecolor="black",
                linewidth=1.5,
                hatch=hatch,
                label=f"{dimension} (df2)",
                left=left_df1 + left_df2,
                alpha=1,
            )
            # Adjust 10 for the gap width
            left_df2 += df2_value

        # add additional stocks for gap, added_beyond_stocks_region, added_beyond_stocks_global
        df3_value = (
            grouped_df3.loc[region, "line"] / 1e3
            - combined_df.loc[(region), "df1_stock"].sum()
        )  # Convert to t/cap
        ax2.barh(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df3_value,
            (cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            label=f"{dimension} (df3)",
            left=left_df1 + left_df2,
        )
        # Center coordinates for the text and add text at the calculated position
        center_y = (cumulative_population[i] + cumulative_population[i + 1]) / 2
        center_x = left_df1 + left_df2 + df3_value / 2
        ax2.text(center_x, center_y, "i", ha="center", va="center", fontsize=20)

        left_df3 += df3_value
        ax2.barh(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_regional[region] / 1e3,
            (cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            left=left_df1 + left_df2 + left_df3,
        )
        center_y = (cumulative_population[i] + cumulative_population[i + 1]) / 2
        center_x = (
            left_df1
            + left_df2
            + left_df3
            + df_add_beyond_stocks_regional[region] / 1e3 / 2
        )
        ax2.text(center_x, center_y, "ii", ha="center", va="center", fontsize=20)

        left_df4 += df_add_beyond_stocks_regional[region] / 1e3
        ax2.barh(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_global[region] / 1e3,
            (cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            left=left_df1 + left_df2 + left_df3 + left_df4,
        )
        if df_add_beyond_stocks_global[region] > 0:
            center_y = (cumulative_population[i] + cumulative_population[i + 1]) / 2
            center_x = (
                left_df1
                + left_df2
                + left_df3
                + left_df4
                + df_add_beyond_stocks_global[region] / 1e3 / 2
            )
            ax2.text(center_x, center_y, "iii", ha="center", va="center", fontsize=20)

    # Legend labels
    legend_labels = [dim for dim in dimensions]

    # Dictionary mapping old labels to new labels
    replacement_dict = {
        "nonres_buildings": "nonresid. buildings",
        "other_construction": "civil engineering",
        "res_buildings": "resid. buildings",
        "road_rail": "road & rail",
        "transport_machinery": "transport vehicles",
    }

    # Replace labels using list comprehension
    legend_labels = [
        replacement_dict[label] if label in replacement_dict else label
        for label in legend_labels
    ]

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames = []
    legend_labels_frames = []

    custom_patch = Patch(facecolor="orange", alpha=1, edgecolor="black", linewidth=3)
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("DLS stocks")

    import matplotlib.patches as patches

    # custom_patch = Patch(facecolor='orange', alpha=0.4, edgecolor='black', linewidth=3, hatch='x')
    custom_patch = patches.Rectangle(
        (0, 0), 1, 1, facecolor="orange", edgecolor="black", linewidth=3, hatch="x"
    )

    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("beyond-DLS stocks")

    custom_patch = Patch(facecolor="none", edgecolor="black")
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("additions to close gap")

    # Create a custom patch with a hatch pattern
    custom_patch2 = Patch(
        facecolor="none", edgecolor=None, hatch=".", label="additions to close gap"
    )
    legend_handles.append(custom_patch2)
    legend_labels.append("beyond-DLS")

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames.append(custom_marker_a)
    legend_labels_frames.append("only DLS")
    legend_handles_frames.append(custom_marker_b)
    legend_labels_frames.append("beyond-DLS: regional")
    legend_handles_frames.append(custom_marker_c)
    legend_labels_frames.append("beyond-DLS: global")

    line_y_values = [164.13]  # Replace with actual y-values
    line_labels = ["scenario-iii: economy-wide stock requirement"]

    # Plot horizontal lines and add labels
    for y_value, label in zip(line_y_values, line_labels):
        ax2.axvline(x=y_value, color="gray", linestyle="--", linewidth=2)
        ax2.text(
            x=y_value,
            y=ax2.get_ylim()[0],
            s=label,
            verticalalignment="bottom",
            horizontalalignment="center",
            fontsize=20,
        )

    ax2.set_xlabel("[tons / capita]", fontsize=24, labelpad=20)

    titlea = "..(b) REGIONAL, per capita (y-axis height corresponds to population)"

    ax2.set_title(titlea, fontsize=36, loc="left", pad=35)

    # Calculate midpoints of the bars for the y-axis labels
    midpoints = (cumulative_population[:-1] + cumulative_population[1:]) / 2

    # Set y-ticks and y-tick labels to the midpoints of the bars
    ax2.set_yticks(midpoints)
    ax2.set_yticklabels(
        [
            "Eastern Europe",
            "Western Europe",
            " Japan, Australia, NZ",
            "North America",
            "Form. Soviet Union",
            "Centr. planned Asia",
            "N. Africa & M. East",
            "Latin America",
            "Pacific Asia",
            "South Asia",
            "Subsaharan Africa",
        ],
        fontsize=24,
    )

    # Optionally, adjust the font size of the tick labels on the x-axis as well
    ax2.tick_params(axis="x", labelsize=24)
    ax2.set_xlim(0, 240)
    ax2.axhspan(ymin=-1, ymax=3.8, color="yellow", alpha=0.2)
    ax2.axhspan(ymin=3.8, ymax=8.2, color="blue", alpha=0.2)
    ax2.axhspan(ymin=8.2, ymax=12, color="green", alpha=0.2)
    ax2.set_xticks([0, 20, 40, 60, 80, 100, 120, 140, 160, 180, 200, 220, 240])
    ax2.set_xticklabels(
        [0, 20, 40, 60, 80, 100, 120, 140, 160, 180, 200, 220, 240], fontsize=24
    )
    fig_title = (
        "Existings stocks providing DLS & beyond-DLS + additions to close DLS gaps..."
    )
    # fig.suptitle(fig_title, fontsize=40,  x=0.545, y=1.025)
    fig.suptitle(fig_title, fontsize=42, x=0.51, y=1.028)

    # reorder legend for product groups
    legend_handles = [
        handles[0],
        handles[1],
        handles[2],
        handles[3],
        legend_handles[0],
        legend_handles[1],
        legend_handles[2],
        legend_handles[3],
        legend_handles[4],
        legend_handles[5],
        legend_handles[6],
        legend_handles[7],
    ]
    legend_labels = [
        labels[0],
        labels[1],
        labels[2],
        labels[3],
        legend_labels[0],
        legend_labels[1],
        legend_labels[2],
        legend_labels[3],
        legend_labels[4],
        legend_labels[5],
        legend_labels[6],
        legend_labels[7],
    ]
    legend = fig.legend(
        legend_handles,
        legend_labels,
        loc="upper center",
        bbox_to_anchor=(0.58, 0.6),
        ncol=6,
        fontsize=20,
        title="  Stock of material groups          Stock of product groups",
        title_fontsize=20,
    )
    legend.get_frame().set_edgecolor("grey")
    legend.get_title().set_fontsize(22)
    legend._legend_box.align = "left"  # Align title text
    ax0.set_ylim([0 - 1.05, 1.05])
    plt.tight_layout()
    plt.subplots_adjust(hspace=0.6)
    plt.show()
    # fig.savefig('Fig2_DLSbeyond.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    plt.close("all")


def plot_bars_Vert_gap_headroom_two_subplots_doubleGlob_mod_eff(
    df1,
    df2,
    df3,
    df_add_beyond_stocks_regional,
    df_add_beyond_stocks_global,
    global_stocks,
    global_stock_prod,
    subcategory,
    line_y_values,
    line_labels,
    DLS_stock_thresh_eff_glob,
    DLS_stock_thresh_eff_cap,
):
    grouped_df1 = df1.groupby(["region", subcategory]).sum()
    grouped_df2 = df2.groupby(["region", subcategory]).sum()
    grouped_df3 = df3.set_index(["region", subcategory]).groupby("region").sum()

    grouped_df1 = pd.DataFrame(grouped_df1.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df2 = pd.DataFrame(grouped_df2.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df3 = pd.DataFrame(grouped_df3.sum(axis=1)).rename(columns={0: "line"})

    # Creating a single dataframe for plotting
    combined_df = pd.concat([grouped_df1, grouped_df2], axis=1, keys=["df1", "df2"])
    combined_df.columns = combined_df.columns.map("_".join)
    combined_df = combined_df / 1e3  # Convert to t/cap
    combined_df.reset_index(inplace=True)
    region_order = [
        "CPA",
        "EEU",
        "WEU",
        "FSU",
        "NAM",
        "PAO",
        "LAM",
        "AFR",
        "MEA",
        "PAS",
        "SAS",
    ]
    sector_order = [
        "res_buildings",
        "nonres_buildings",
        "other_construction",
        "road_rail",
        "transport_machinery",
        "machinery",
        "other",
    ]

    # Create a mapping of region to the custom order
    order_mapping = {region: i for i, region in enumerate(region_order)}
    # Create a mapping of sector to the custom order
    order_mapping_sector = {sector: i for i, sector in enumerate(sector_order)}

    # Map the 'region' column to a new column that represents its order
    combined_df["region_order"] = combined_df["region"].map(order_mapping)
    # Sort the DataFrame by the new 'region_order' column, then drop it

    # Map the 'sector' column to a new column that represents its order
    combined_df["sector_order"] = combined_df["sector"].map(order_mapping_sector)
    # Sort the DataFrame by the new 'sector_order' column, then drop it
    combined_df_sorted = combined_df.sort_values(
        by=["region_order", "sector_order"]
    ).drop(columns=["sector_order"])

    combined_df = combined_df_sorted.set_index(["region", "sector"])

    # also reorder global_stock_prod and global_stock
    global_stock_prod = global_stock_prod.reindex(sector_order)
    global_stocks = global_stocks.reindex(["minerals", "metals", "biomass", "fossils"])

    # Create the figure
    fig = plt.figure(figsize=(24, 28))

    # Create a 2-row, 2-column grid
    gs = gridspec.GridSpec(2, 2, height_ratios=[2, 2.1], width_ratios=[4.5, 2])

    # First panel spans only the first column of the first row (half-width)
    ax0 = fig.add_subplot(gs[0, 0])

    # Second panel spans both columns of the second row (full width)
    ax2 = fig.add_subplot(gs[1, :])

    # Optional: Hide the unused top-right space (gs[0, 1])
    fig.add_subplot(gs[0, 1]).axis("off")

    global_stocks_products = global_stock_prod.groupby("region", axis=1).sum()

    from matplotlib.patches import FancyBboxPatch

    # Materials and their colors
    materials = global_stocks.index
    sectors = global_stocks_products.index
    colors = [
        (0.6, 0.58, 0.58),
        (0.18, 0.46, 0.71),
        (0.44, 0.68, 0.28),
        (0.93, 0.40, 0.19),
    ]
    colors2 = plt.cm.get_cmap("Pastel2", len(sectors))
    # replace grey color in pastel2 so it is not confuesd with grey in variable colors
    colors_array = colors2(np.arange(len(sectors)))
    replacement_color = [1.0, 0.5, 0.5, 1.0]  # Replace with a red-like RGBA
    colors_array[-1] = replacement_color  # Replace gray with your chosen color
    from matplotlib.colors import ListedColormap

    custom_cmap = ListedColormap(colors_array)
    colors2 = custom_cmap
    # rearrange color order:
    # Extract the colors as an array
    colors3 = colors2(np.arange(colors2.N))
    # Define your custom order for the colors (example: swap the first and last color)
    custom_order = [4, 1, 6, 5, 3, 0, 2]  # Adjust based on the desired order
    # Reorder the colors based on the custom order
    reordered_colors = colors3[custom_order]
    # Create a new colormap with the reordered colors
    colors2 = ListedColormap(reordered_colors)

    # Adjust hatching patterns for keys
    key_hatch = {"DLS_prov": "", "beyond_DLS": "."}

    # adjust so that values are in Gigatons
    global_stocks_products = global_stocks_products / 1e9
    global_stocks = global_stocks / 1e9
    DLS_stock_thresh_eff_glob = DLS_stock_thresh_eff_glob / 1e9

    double_bar = 0.4

    # Function to create rounded vertical bars
    def plot_rounded_barv(ax, x, height, bottom=0, color=None, hatch=None, alpha=None):
        boxstyle = "round,pad=0"
        bar_width = 0.175  # Consistent width for all bars

        # Outline with opaque black edge
        outline_bbox = FancyBboxPatch(
            (x - bar_width, bottom),
            bar_width * 2,
            height,
            boxstyle=boxstyle,
            ec="black",
            fc="none",
            hatch=hatch,
            lw=1.5,
        )
        ax.add_patch(outline_bbox)

        # Filled bar with transparency if color is provided
        if color is not None:
            fill_bbox = FancyBboxPatch(
                (x - bar_width, bottom),
                bar_width * 2,
                height,
                boxstyle=boxstyle,
                ec="none",
                fc=color,
                hatch=hatch,
                lw=0,
                alpha=alpha,
            )
            ax.add_patch(fill_bbox)

    left_bar_x = -0.425  # shift further left
    right_bar_x = 0.075  # closer to center

    # Plot bars for global stocks of products (now vertical)
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, sector in enumerate(sectors):
            value = global_stocks_products.loc[sector, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            bottom += value

    # Plot bars for global stocks of materials (now vertical)
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, material in enumerate(materials):
            value = global_stocks.loc[material, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            bottom += value

    bottom_save = bottom

    # Add explicit legend entries for materials
    handles = [
        FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.1", fc=color, ec="black")
        for color in colors
    ]
    labels = list(materials)  # Convert to a list

    # Plot the gap keys (now vertical bars)
    keys_to_plot_target = ["gap_targeted", "gap_regional", "gap_trickle"]
    letters = ["i", "+ii", "+iii"]

    for idx, key in enumerate(keys_to_plot_target):
        value = global_stocks[key].sum()
        plot_rounded_barv(
            ax0, left_bar_x, value, bottom=bottom, color="white", hatch="None"
        )
        plot_rounded_barv(ax0, left_bar_x, value, bottom=bottom, color=None, hatch=None)
        ax0.text(
            left_bar_x,
            bottom + value / 2,
            letters[idx],
            ha="center",
            va="center",
            fontsize=26,
        )
        if key == "gap_targeted":
            ax0.plot(
                left_bar_x,
                bottom + DLS_stock_thresh_eff_glob,
                "+",
                color="red",
                markersize=24,
                markeredgewidth=2,
            )
        bottom += value

    # Plot for the convergence gap (also vertical)
    keys_to_plot_all = ["gap_targeted", "gap_regional", "gap_trickle"]
    bottom = bottom_save

    for idx, key in enumerate(keys_to_plot_all):
        value = global_stocks[key].sum()
        plot_rounded_barv(
            ax0, right_bar_x, value, bottom=bottom, color="white", hatch=None
        )
        plot_rounded_barv(
            ax0, right_bar_x, value, bottom=bottom, color=None, hatch=None
        )
        ax0.text(
            right_bar_x,
            bottom + value / 2,
            letters[idx],
            ha="center",
            va="center",
            fontsize=24,
        )
        if key == "gap_targeted":
            ax0.plot(
                right_bar_x,
                bottom + DLS_stock_thresh_eff_glob,
                "+",
                color="red",
                markersize=24,
                markeredgewidth=2,
            )
        bottom += value

    # Set x-ticks to indicate the two "positions" (product vs material logic)
    ax0.set_xticks([left_bar_x, right_bar_x])
    ax0.set_xticklabels(["by product", "by material"], fontsize=26)
    for label in ax0.get_xticklabels():
        label.set_rotation(0)
    # Vertical axis is now the value scale
    ax0.set_ylim(0, bottom)
    ax0.set_xlim(-1, 1)

    # Adjust tick sizes and labels
    ax0.tick_params(axis="y", labelsize=26)
    ax0.tick_params(axis="x", labelsize=26)

    # Y-axis label replaces former X-axis label
    ax0.set_ylabel("[Gigatons = 10^9 tons]", fontsize=26, labelpad=20)

    ax0.ticklabel_format(style="plain", axis="y")

    # Brackets now apply along y-axis (vertical stacking) — no change needed unless you want to flip the orientation
    y_bracket_position = 0.68  # Remains meaningful vertically
    y_bracket_position_below = (
        -0.67
    )  # These would be x-coordinates if you want to flip them horizontally
    axlim = ax0.get_ylim()

    def draw_vertical_bracket(
        ax, y_start, y_end, x_position, label, label_offset=0.1, bracket_width=0.1
    ):
        """
        Draws a vertical bracket from y_start to y_end at a given x_position.
        """
        # Draw vertical line (bracket spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Draw horizontal tips
        ax.plot(
            [x_position - bracket_width / 2, x_position + bracket_width / 2],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width / 2, x_position + bracket_width / 2],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Add label
        ax.text(
            x_position + bracket_width * 1.2,
            (y_start + y_end) / 2,
            label,
            ha="left",
            va="center",
            fontsize=20,
        )

    def draw_vertical_bracket_right_label(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point LEFT (towards left side)
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on right
        ax.text(
            x_position + label_offset,
            (y_start + y_end) / 2,
            label,
            ha="left",
            va="center",
            fontsize=26,
        )

    def draw_vertical_bracket_right_label_up(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point LEFT (towards left side)
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on right
        ax.text(
            x_position + label_offset,
            (y_start + y_end) / 1.8,
            label,
            ha="left",
            va="center",
            fontsize=26,
        )

    def draw_vertical_bracket_left_label(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point RIGHT (towards right side)
        ax.plot(
            [x_position, x_position + bracket_width],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position, x_position + bracket_width],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on left
        ax.text(
            x_position - label_offset,
            (y_start + y_end) / 2,
            label,
            ha="right",
            va="center",
            fontsize=26,
            rotation=90,
        )

    x_bracket_position = 0.5  # Adjust this to align with your bar or region

    space_keeper = 10
    draw_vertical_bracket_left_label(
        ax0,
        0.1 + space_keeper,
        699.4 - space_keeper,
        -0.7,
        label="existing economy-wide\n stocks in 2015",
    )
    # draw_vertical_bracket_left_label(ax0, 700 +space_keeper,  1399 -space_keeper, -0.7, label="net additions to stocks\n to close DLS gaps")

    draw_vertical_bracket_right_label(
        ax0, 0.1 + space_keeper, 193 - space_keeper, 0.35, label="existing\n DLS stocks"
    )
    draw_vertical_bracket_right_label(
        ax0,
        194.2 + space_keeper,
        699.4 - space_keeper,
        0.35,
        label="existing\n beyond-DLS\n stocks",
    )
    draw_vertical_bracket_right_label(
        ax0,
        699.4 + space_keeper,
        699.4 + 82.2 - space_keeper,
        0.5,
        label="DLS stock gap-\n scenario-i)",
    )
    draw_vertical_bracket_right_label_up(
        ax0,
        699.4 + space_keeper,
        699.4 + 82.2 + 146 - space_keeper,
        0.425,
        label="NAS- scenario-ii",
    )
    draw_vertical_bracket_right_label_up(
        ax0,
        699.4 + space_keeper,
        699.4 + 82.2 + 146 + 479.2 - space_keeper,
        0.35,
        label="net stock additions\n (NAS)- scenario-iii",
    )

    # Set final plot title
    ax0.set_title(
        "(a) Existings DLS & beyond-DLS stocks + additions closing DLS gaps:\n\n    ...GLOBAL, at scale",
        fontsize=38,
        loc="left",
        pad=35,
    )

    # Unique regions and dimensions for plotting
    regions = combined_df.index.get_level_values(0).unique()
    dimensions = sector_order

    # Assigning distinct colors and hatch patterns to each dimension and creating legend handles
    colors = plt.cm.get_cmap("Pastel2", len(dimensions))
    colors = colors2
    # hatches = ['x','x','x','x','x','x','x','x','x','x']
    hatches = ["", "", "", "", "", "", "", "", "", ""]
    hatches = [".", ".", ".", ".", ".", ".", ".", ".", ".", "."]
    legend_handles = []

    # Calculate cumulative population
    cumulative_population = np.cumsum(
        [0]
        + [
            1.55716e06,
            121426,
            497389,
            288469,
            360286,
            156532,
            618764,
            954920,
            469410,
            579769,
            1.74936e06,
        ]
    )

    # Sort regions
    regions = region_order

    # Plot each region
    for i, region in enumerate(regions):
        bottom_df1 = 0  # Bottom for df1 stack
        bottom_df2 = 0  # Bottom for df2 stack (starts at top of df1 stack)
        bottom_df3 = 0
        bottom_df4 = 0

        # Plot df1 and calculate total df1 value for the region
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df1_value = combined_df.loc[(region, dimension), "df1_stock"]
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1])
                / 2,  # x: center of region
                df1_value,  # height
                width=(
                    cumulative_population[i + 1] - cumulative_population[i]
                ),  # bar width = population share
                color=color,
                edgecolor="black",
                linewidth=1.5,
                label=f"{dimension} (df1)",
                bottom=bottom_df1,
                alpha=1,
            )
            bottom_df1 += df1_value

            # Add legend handle for the first region
            if i == 0:
                legend_handles.append(
                    plt.Rectangle((0, 0), 1, 1, color=color, alpha=1, edgecolor="black")
                )

        # Plot df2 stacked on top of df1
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df2_value = combined_df.loc[(region, dimension), "df2_stock"]
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                width=(cumulative_population[i + 1] - cumulative_population[i]),
                color=color,
                edgecolor="none",
                linewidth=1.5,
                hatch=hatch,
                label=f"{dimension} (df2)",
                bottom=bottom_df1 + bottom_df2,
                alpha=0.6,
            )
            # Overplot edges for hatches
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                width=(cumulative_population[i + 1] - cumulative_population[i]),
                color="none",
                edgecolor="black",
                linewidth=1.5,
                hatch=hatch,
                bottom=bottom_df1 + bottom_df2,
                alpha=1,
            )
            bottom_df2 += df2_value

        # df3 additional stock beyond df1+df2
        df3_value = (
            grouped_df3.loc[region, "line"] / 1e3
            - combined_df.loc[(region), "df1_stock"].sum()
        )
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df3_value,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            label=f"{dimension} (df3)",
            bottom=bottom_df1 + bottom_df2,
        )

        # Red cross for DLS threshold
        ax2.plot(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            bottom_df1 + bottom_df2 + DLS_stock_thresh_eff_cap.loc[region] / 1e3,
            "+",
            color="red",
            markersize=16,
            markeredgewidth=2,
        )

        # Annotate 'i'
        center_x = (cumulative_population[i] + cumulative_population[i + 1]) / 2
        center_y = bottom_df1 + bottom_df2 + df3_value / 2
        ax2.text(center_x, center_y, "i", ha="center", va="center", fontsize=24)

        bottom_df3 += df3_value

        # df4 - added beyond stocks regional
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_regional[region] / 1e3,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            bottom=bottom_df1 + bottom_df2 + bottom_df3,
        )

        center_y = (
            bottom_df1
            + bottom_df2
            + bottom_df3
            + df_add_beyond_stocks_regional[region] / 1e3 / 2
        )
        ax2.text(center_x, center_y, "+ii", ha="center", va="center", fontsize=24)

        bottom_df4 += df_add_beyond_stocks_regional[region] / 1e3

        # df5 - added beyond stocks global
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_global[region] / 1e3,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            bottom=bottom_df1 + bottom_df2 + bottom_df3 + bottom_df4,
        )

        if df_add_beyond_stocks_global[region] > 0:
            center_y = (
                bottom_df1
                + bottom_df2
                + bottom_df3
                + bottom_df4
                + df_add_beyond_stocks_global[region] / 1e3 / 2
            )
            ax2.text(center_x, center_y, "+iii", ha="center", va="center", fontsize=24)

    # Legend labels
    legend_labels = [dim for dim in dimensions]

    # Dictionary mapping old labels to new labels
    replacement_dict = {
        "nonres_buildings": "nonresid. buildings",
        "other_construction": "civil engineering",
        "res_buildings": "resid. buildings",
        "road_rail": "road & rail",
        "transport_machinery": "transport vehicles",
    }

    # Replace labels using list comprehension
    legend_labels = [
        replacement_dict[label] if label in replacement_dict else label
        for label in legend_labels
    ]

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames = []
    legend_labels_frames = []

    custom_patch = Patch(facecolor="orange", alpha=1, edgecolor="black", linewidth=3)
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("DLS stocks")

    import matplotlib.patches as patches

    # custom_patch = Patch(facecolor='orange', alpha=0.4, edgecolor='black', linewidth=3, hatch='x')
    custom_patch = patches.Rectangle(
        (0, 0), 1, 1, facecolor="orange", edgecolor="black", linewidth=3, hatch="x"
    )

    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("beyond-DLS stocks")

    custom_patch = Patch(facecolor="none", edgecolor="black")
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("additions to close gap")

    # Create a custom patch with a hatch pattern
    custom_patch2 = Patch(
        facecolor="none", edgecolor=None, hatch=".", label="additions to close gap"
    )
    legend_handles.append(custom_patch2)
    legend_labels.append("beyond-DLS")

    # Create custom markers using Unicode characters for converged cross
    custom_marker_cross = Line2D([0], [0], color="red", marker="$+$", markersize=12, 
    linestyle='None')
    legend_handles.append(custom_marker_cross)
    legend_labels.append("scenario-i: conver-\nged practices")

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames.append(custom_marker_a)
    legend_labels_frames.append("only DLS")
    legend_handles_frames.append(custom_marker_b)
    legend_labels_frames.append("beyond-DLS: regional")
    legend_handles_frames.append(custom_marker_c)
    legend_labels_frames.append("beyond-DLS: global")

    line_x_values = [177.36999999999998]  # Replace with actual x-values
    line_labels = ["scenario-iii: economy-wide stock requirement (global)"]

    # Plot horizontal lines and add labels (switched axes)
    for x_value, label in zip(line_x_values, line_labels):
        ax2.axhline(y=x_value, color="gray", linestyle="--", linewidth=2)
        ax2.text(
            x=ax2.get_xlim()[1] / 2,  # start of x-axis
            y=x_value + 6,  # aligned to the line's y-position
            s=label,
            verticalalignment="center",  # centered on the line
            horizontalalignment="left",  # aligned to the left of the plot
            fontsize=20,
        )

    titlea = "..(b) REGIONAL, Ø per capita (scale: regions' entire population')"

    ax2.set_title(titlea, fontsize=38, loc="left", pad=35)
    ax2.set_xlim(0 - 50000, cumulative_population[-1] + 50000)

    # Calculate midpoints of the bars for the y-axis labels
    midpoints = (cumulative_population[:-1] + cumulative_population[1:]) / 2

    # Set y-ticks and y-tick labels to the midpoints of the bars
    ax2.set_xticks(midpoints)
    # ax2.set_xticklabels(['Eastern Europe', 'Western Europe', ' Japan, Australia, NZ', 'North America', 'Form. Soviet Union', 'Centr. planned Asia', 'N. Africa & M. East', 'Latin America', 'Pacific Asia', 'South Asia', 'Subsaharan Africa'], fontsize=24)
    ax2.set_xticklabels(
        [
            "Centr.\n planned Asia",
            "Eastern Europe",
            "Western Europe",
            "Form.\n Soviet Union",
            "North America",
            " Japan,\n Australia, NZ",
            "Latin America",
            "Subsaharan\n Africa",
            "N. Africa &\n M. East",
            "Pacific Asia",
            "South Asia",
        ],
        fontsize=28,
    )
    for label in ax2.get_xticklabels():
        label.set_rotation(45)
    # Optionally, adjust the font size of the tick labels on the x-axis as well
    ax2.tick_params(axis="y", labelsize=26)
    ax2.set_ylim(0, 240)
    ax2.set_ylabel("tons / capita", fontsize=26)

    fig_title = (
        "Existings stocks providing DLS & beyond-DLS + additions to close DLS gaps..."
    )

    # reorder legend for product groups
    legend_handles1 = [
        legend_handles[0],
        legend_handles[1],
        legend_handles[2],
        legend_handles[3],
        legend_handles[4],
        legend_handles[5],
        legend_handles[6],
    ]
    legend_labels1 = [
        legend_labels[0],
        legend_labels[1],
        legend_labels[2],
        legend_labels[3],
        legend_labels[4],
        legend_labels[5],
        legend_labels[6],
    ]
    legend_handles2 = [handles[0], handles[1], handles[2], handles[3]]
    legend_labels2 = [labels[0], labels[1], labels[2], labels[3]]
    legend_handles3 = [legend_handles[7], legend_handles[8]]
    legend_labels3 = [legend_labels[7], legend_labels[8]]

    # legend = fig.legend(legend_handles, legend_labels, loc='upper center', bbox_to_anchor=(0.58, 0.6), ncol=6, fontsize=20, title='  Stock of material groups          Stock of product groups', title_fontsize=20)
    legend1 = fig.legend(
        legend_handles1,
        legend_labels1,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.91 + 0.035),
        ncol=1,
        fontsize=26,
        title="Stock of product groups",
        title_fontsize=26,
    )
    legend2 = fig.legend(
        legend_handles2,
        legend_labels2,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.75 + 0.03),
        ncol=1,
        fontsize=26,
        title="Stock of material groups",
        title_fontsize=26,
    )
    legend3 = fig.legend(
        legend_handles3,
        legend_labels3,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.65 + 0.025),
        ncol=1,
        fontsize=26,
        title="Other entries   ",
        title_fontsize=26,
    )

    for i in [legend1, legend2, legend3]:
        i.get_frame().set_edgecolor("grey")
        i.get_title().set_fontsize(26)
        i._legend_box.align = "left"  # Align title text
    plt.tight_layout()
    plt.subplots_adjust(hspace=0.3)
    plt.show()
    #fig.savefig("Fig_3.pdf", bbox_inches='tight')
    plt.close("all")


def plot_bars_Vert_gap_headroom_two_subplots_doubleGlob_mod_eff_ii(
    df1,
    df2,
    df3,
    df_add_beyond_stocks_regional,
    df_add_beyond_stocks_global,
    global_stocks,
    global_stock_prod,
    subcategory,
    line_y_values,
    line_labels,
    DLS_stock_thresh_eff_glob,
    DLS_stock_thresh_eff_cap,
):
    grouped_df1 = df1.groupby(["region", subcategory]).sum()
    grouped_df2 = df2.groupby(["region", subcategory]).sum()
    grouped_df3 = df3.set_index(["region", subcategory]).groupby("region").sum()

    grouped_df1 = pd.DataFrame(grouped_df1.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df2 = pd.DataFrame(grouped_df2.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df3 = pd.DataFrame(grouped_df3.sum(axis=1)).rename(columns={0: "line"})

    # Creating a single dataframe for plotting
    combined_df = pd.concat([grouped_df1, grouped_df2], axis=1, keys=["df1", "df2"])
    combined_df.columns = combined_df.columns.map("_".join)
    combined_df = combined_df / 1e3  # Convert to t/cap
    combined_df.reset_index(inplace=True)
    region_order = [
        "CPA",
        "EEU",
        "WEU",
        "FSU",
        "NAM",
        "PAO",
        "LAM",
        "AFR",
        "MEA",
        "PAS",
        "SAS",
    ]
    sector_order = [
        "res_buildings",
        "nonres_buildings",
        "other_construction",
        "road_rail",
        "transport_machinery",
        "machinery",
        "other",
    ]

    # Create a mapping of region to the custom order
    order_mapping = {region: i for i, region in enumerate(region_order)}
    # Create a mapping of sector to the custom order
    order_mapping_sector = {sector: i for i, sector in enumerate(sector_order)}

    # Map the 'region' column to a new column that represents its order
    combined_df["region_order"] = combined_df["region"].map(order_mapping)
    # Sort the DataFrame by the new 'region_order' column, then drop it

    # Map the 'sector' column to a new column that represents its order
    combined_df["sector_order"] = combined_df["sector"].map(order_mapping_sector)
    # Sort the DataFrame by the new 'sector_order' column, then drop it
    combined_df_sorted = combined_df.sort_values(
        by=["region_order", "sector_order"]
    ).drop(columns=["sector_order"])

    combined_df = combined_df_sorted.set_index(["region", "sector"])

    # also reorder global_stock_prod and global_stock
    global_stock_prod = global_stock_prod.reindex(sector_order)
    global_stocks = global_stocks.reindex(["minerals", "metals", "biomass", "fossils"])

    # Create the figure
    fig = plt.figure(figsize=(24, 28))

    # Create a 2-row, 2-column grid
    gs = gridspec.GridSpec(2, 2, height_ratios=[2, 2.1], width_ratios=[4.5, 2])

    # First panel spans only the first column of the first row (half-width)
    ax0 = fig.add_subplot(gs[0, 0])

    # Second panel spans both columns of the second row (full width)
    ax2 = fig.add_subplot(gs[1, :])

    # Optional: Hide the unused top-right space (gs[0, 1])
    fig.add_subplot(gs[0, 1]).axis("off")

    global_stocks_products = global_stock_prod.groupby("region", axis=1).sum()

    from matplotlib.patches import FancyBboxPatch

    # Materials and their colors
    materials = global_stocks.index
    sectors = global_stocks_products.index
    # colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99']
    colors = [
        (0.6, 0.58, 0.58),
        (0.18, 0.46, 0.71),
        (0.44, 0.68, 0.28),
        (0.93, 0.40, 0.19),
    ]
    colors2 = plt.cm.get_cmap("Pastel2", len(sectors))
    # replace grey color in pastel2 so it is not confuesd with grey in variable colors
    colors_array = colors2(np.arange(len(sectors)))
    replacement_color = [1.0, 0.5, 0.5, 1.0]  # Replace with a red-like RGBA
    colors_array[-1] = replacement_color  # Replace gray with your chosen color
    from matplotlib.colors import ListedColormap

    custom_cmap = ListedColormap(colors_array)
    colors2 = custom_cmap
    # rearrange color order:
    # Extract the colors as an array
    colors3 = colors2(np.arange(colors2.N))
    # Define your custom order for the colors (example: swap the first and last color)
    custom_order = [4, 1, 6, 5, 3, 0, 2]  # Adjust based on the desired order
    # Reorder the colors based on the custom order
    reordered_colors = colors3[custom_order]
    # Create a new colormap with the reordered colors
    colors2 = ListedColormap(reordered_colors)

    # Adjust hatching patterns for keys
    key_hatch = {"DLS_prov": "", "beyond_DLS": "."}

    # adjust so that values are in Gigatons
    global_stocks_products = global_stocks_products / 1e9
    global_stocks = global_stocks / 1e9
    DLS_stock_thresh_eff_glob = DLS_stock_thresh_eff_glob / 1e9

    double_bar = 0.4

    # Function to create rounded vertical bars
    def plot_rounded_barv(ax, x, height, bottom=0, color=None, hatch=None, alpha=None):
        boxstyle = "round,pad=0"
        bar_width = 0.175  # Consistent width for all bars

        # Outline with opaque black edge
        outline_bbox = FancyBboxPatch(
            (x - bar_width, bottom),
            bar_width * 2,
            height,
            boxstyle=boxstyle,
            ec="black",
            fc="none",
            hatch=hatch,
            lw=1.5,
        )
        ax.add_patch(outline_bbox)

        # Filled bar with transparency if color is provided
        if color is not None:
            fill_bbox = FancyBboxPatch(
                (x - bar_width, bottom),
                bar_width * 2,
                height,
                boxstyle=boxstyle,
                ec="none",
                fc=color,
                hatch=hatch,
                lw=0,
                alpha=alpha,
            )
            ax.add_patch(fill_bbox)

    left_bar_x = -0.425  # shift further left
    right_bar_x = 0.075  # closer to center

    # Plot bars for global stocks of products (now vertical)
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, sector in enumerate(sectors):
            value = global_stocks_products.loc[sector, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            bottom += value

    # Plot bars for global stocks of materials (now vertical)
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, material in enumerate(materials):
            value = global_stocks.loc[material, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            bottom += value

    bottom_save = bottom

    # Add explicit legend entries for materials
    handles = [
        FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.1", fc=color, ec="black")
        for color in colors
    ]
    labels = list(materials)  # Convert to a list

    # Plot the gap keys (now vertical bars)
    keys_to_plot_target = ["gap_targeted", "gap_regional"]
    letters = ["i", "+ii"]

    for idx, key in enumerate(keys_to_plot_target):
        value = global_stocks[key].sum()
        plot_rounded_barv(
            ax0, left_bar_x, value, bottom=bottom, color="white", hatch="None"
        )
        plot_rounded_barv(ax0, left_bar_x, value, bottom=bottom, color=None, hatch=None)
        ax0.text(
            left_bar_x,
            bottom + value / 2,
            letters[idx],
            ha="center",
            va="center",
            fontsize=26,
        )
        if key == "gap_targeted":
            ax0.plot(
                left_bar_x,
                bottom + DLS_stock_thresh_eff_glob,
                "+",
                color="red",
                markersize=24,
                markeredgewidth=2,
            )
        bottom += value

    # Plot for the convergence gap (also vertical)
    keys_to_plot_all = ["gap_targeted", "gap_regional"]
    bottom = bottom_save

    for idx, key in enumerate(keys_to_plot_all):
        value = global_stocks[key].sum()
        plot_rounded_barv(
            ax0, right_bar_x, value, bottom=bottom, color="white", hatch=None
        )
        plot_rounded_barv(
            ax0, right_bar_x, value, bottom=bottom, color=None, hatch=None
        )
        ax0.text(
            right_bar_x,
            bottom + value / 2,
            letters[idx],
            ha="center",
            va="center",
            fontsize=24,
        )
        if key == "gap_targeted":
            ax0.plot(
                right_bar_x,
                bottom + DLS_stock_thresh_eff_glob,
                "+",
                color="red",
                markersize=24,
                markeredgewidth=2,
            )
        bottom += value

    # Set x-ticks to indicate the two "positions" (product vs material logic)
    ax0.set_xticks([left_bar_x, right_bar_x])
    ax0.set_xticklabels(["by product", "by material"], fontsize=26)
    for label in ax0.get_xticklabels():
        label.set_rotation(0)
    # Vertical axis is now the value scale
    ax0.set_ylim(0, 1400)
    ax0.set_xlim(-1, 1)

    # Adjust tick sizes and labels
    ax0.tick_params(axis="y", labelsize=26)
    ax0.tick_params(axis="x", labelsize=26)

    # Y-axis label replaces former X-axis label
    ax0.set_ylabel("[Gigatons = 10^9 tons]", fontsize=26, labelpad=20)

    ax0.ticklabel_format(style="plain", axis="y")

    # Brackets now apply along y-axis (vertical stacking) — no change needed unless you want to flip the orientation
    y_bracket_position = 0.68  # Remains meaningful vertically
    y_bracket_position_below = (
        -0.67
    )  # These would be x-coordinates if you want to flip them horizontally
    axlim = ax0.get_ylim()

    def draw_vertical_bracket(
        ax, y_start, y_end, x_position, label, label_offset=0.1, bracket_width=0.1
    ):
        """
        Draws a vertical bracket from y_start to y_end at a given x_position.
        """
        # Draw vertical line (bracket spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Draw horizontal tips
        ax.plot(
            [x_position - bracket_width / 2, x_position + bracket_width / 2],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width / 2, x_position + bracket_width / 2],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Add label
        ax.text(
            x_position + bracket_width * 1.2,
            (y_start + y_end) / 2,
            label,
            ha="left",
            va="center",
            fontsize=20,
        )

    def draw_vertical_bracket_right_label(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point LEFT (towards left side)
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on right
        ax.text(
            x_position + label_offset,
            (y_start + y_end) / 2,
            label,
            ha="left",
            va="center",
            fontsize=26,
        )

    def draw_vertical_bracket_right_label_up(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point LEFT (towards left side)
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on right
        ax.text(
            x_position + label_offset,
            (y_start + y_end) / 1.8,
            label,
            ha="left",
            va="center",
            fontsize=26,
        )

    def draw_vertical_bracket_left_label(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point RIGHT (towards right side)
        ax.plot(
            [x_position, x_position + bracket_width],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position, x_position + bracket_width],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on left
        ax.text(
            x_position - label_offset,
            (y_start + y_end) / 2,
            label,
            ha="right",
            va="center",
            fontsize=26,
            rotation=90,
        )

    x_bracket_position = 0.5  # Adjust this to align with your bar or region

    space_keeper = 10
    draw_vertical_bracket_left_label(
        ax0,
        0.1 + space_keeper,
        699.4 - space_keeper,
        -0.7,
        label="existing economy-wide\n stocks in 2015",
    )
    # draw_vertical_bracket_left_label(ax0, 700 +space_keeper,  1399 -space_keeper, -0.7, label="net additions to stocks\n to close DLS gaps")

    draw_vertical_bracket_right_label(
        ax0, 0.1 + space_keeper, 193 - space_keeper, 0.35, label="existing\n DLS stocks"
    )
    draw_vertical_bracket_right_label(
        ax0,
        194.2 + space_keeper,
        699.4 - space_keeper,
        0.35,
        label="existing\n beyond-DLS\n stocks",
    )
    draw_vertical_bracket_right_label(
        ax0,
        699.4 + space_keeper,
        699.4 + 82.2 - space_keeper,
        0.5,
        label="DLS stock gap-\n scenario-i)",
    )
    draw_vertical_bracket_right_label_up(
        ax0,
        699.4 + space_keeper,
        699.4 + 82.2 + 146 - space_keeper,
        0.425,
        label="NAS- scenario-ii",
    )

    # Set final plot title
    ax0.set_title(
        "(a) Existings DLS & beyond-DLS stocks + additions closing DLS gaps:\n    ...GLOBAL, at scale",
        fontsize=38,
        loc="left",
        pad=35,
    )

    # Unique regions and dimensions for plotting
    regions = combined_df.index.get_level_values(0).unique()
    # dimensions = combined_df.index.get_level_values(1).unique()
    dimensions = sector_order

    # Assigning distinct colors and hatch patterns to each dimension and creating legend handles
    colors = plt.cm.get_cmap("Pastel2", len(dimensions))
    colors = colors2
    # hatches = ['x','x','x','x','x','x','x','x','x','x']
    hatches = ["", "", "", "", "", "", "", "", "", ""]
    hatches = [".", ".", ".", ".", ".", ".", ".", ".", ".", "."]
    legend_handles = []

    # Calculate cumulative population
    cumulative_population = np.cumsum(
        [0]
        + [
            1.55716e06,
            121426,
            497389,
            288469,
            360286,
            156532,
            618764,
            954920,
            469410,
            579769,
            1.74936e06,
        ]
    )

    # Sort regions
    regions = region_order

    # Plot each region
    for i, region in enumerate(regions):
        bottom_df1 = 0  # Bottom for df1 stack
        bottom_df2 = 0  # Bottom for df2 stack (starts at top of df1 stack)
        bottom_df3 = 0
        bottom_df4 = 0

        # Plot df1 and calculate total df1 value for the region
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df1_value = combined_df.loc[(region, dimension), "df1_stock"]
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1])
                / 2,  # x: center of region
                df1_value,  # height
                width=(
                    cumulative_population[i + 1] - cumulative_population[i]
                ),  # bar width = population share
                color=color,
                edgecolor="black",
                linewidth=1.5,
                label=f"{dimension} (df1)",
                bottom=bottom_df1,
                alpha=1,
            )
            bottom_df1 += df1_value

            # Add legend handle for the first region
            if i == 0:
                legend_handles.append(
                    plt.Rectangle((0, 0), 1, 1, color=color, alpha=1, edgecolor="black")
                )

        # Plot df2 stacked on top of df1
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df2_value = combined_df.loc[(region, dimension), "df2_stock"]
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                width=(cumulative_population[i + 1] - cumulative_population[i]),
                color=color,
                edgecolor="none",
                linewidth=1.5,
                hatch=hatch,
                label=f"{dimension} (df2)",
                bottom=bottom_df1 + bottom_df2,
                alpha=0.6,
            )
            # Overplot edges for hatches
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                width=(cumulative_population[i + 1] - cumulative_population[i]),
                color="none",
                edgecolor="black",
                linewidth=1.5,
                hatch=hatch,
                bottom=bottom_df1 + bottom_df2,
                alpha=1,
            )
            bottom_df2 += df2_value

        # df3 additional stock beyond df1+df2
        df3_value = (
            grouped_df3.loc[region, "line"] / 1e3
            - combined_df.loc[(region), "df1_stock"].sum()
        )
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df3_value,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            label=f"{dimension} (df3)",
            bottom=bottom_df1 + bottom_df2,
        )

        # Red cross for DLS threshold
        ax2.plot(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            bottom_df1 + bottom_df2 + DLS_stock_thresh_eff_cap.loc[region] / 1e3,
            "+",
            color="red",
            markersize=16,
            markeredgewidth=2,
        )

        # Annotate 'i'
        center_x = (cumulative_population[i] + cumulative_population[i + 1]) / 2
        center_y = bottom_df1 + bottom_df2 + df3_value / 2
        ax2.text(center_x, center_y, "i", ha="center", va="center", fontsize=24)

        bottom_df3 += df3_value

        # df4 - added beyond stocks regional
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_regional[region] / 1e3,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            bottom=bottom_df1 + bottom_df2 + bottom_df3,
        )

        center_y = (
            bottom_df1
            + bottom_df2
            + bottom_df3
            + df_add_beyond_stocks_regional[region] / 1e3 / 2
        )
        ax2.text(center_x, center_y, "+ii", ha="center", va="center", fontsize=24)

        bottom_df4 += df_add_beyond_stocks_regional[region] / 1e3

        # df5 - added beyond stocks global
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_global[region] / 1e3,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            bottom=bottom_df1 + bottom_df2 + bottom_df3 + bottom_df4,
        )

        if df_add_beyond_stocks_global[region] > 0:
            center_y = (
                bottom_df1
                + bottom_df2
                + bottom_df3
                + bottom_df4
                + df_add_beyond_stocks_global[region] / 1e3 / 2
            )
            ax2.text(center_x, center_y, "+iii", ha="center", va="center", fontsize=24)

    # Legend labels
    legend_labels = [dim for dim in dimensions]

    # Dictionary mapping old labels to new labels
    replacement_dict = {
        "nonres_buildings": "nonresid. buildings",
        "other_construction": "civil engineering",
        "res_buildings": "resid. buildings",
        "road_rail": "road & rail",
        "transport_machinery": "transport vehicles",
    }

    # Replace labels using list comprehension
    legend_labels = [
        replacement_dict[label] if label in replacement_dict else label
        for label in legend_labels
    ]

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames = []
    legend_labels_frames = []

    custom_patch = Patch(facecolor="orange", alpha=1, edgecolor="black", linewidth=3)
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("DLS stocks")

    import matplotlib.patches as patches

    # custom_patch = Patch(facecolor='orange', alpha=0.4, edgecolor='black', linewidth=3, hatch='x')
    custom_patch = patches.Rectangle(
        (0, 0), 1, 1, facecolor="orange", edgecolor="black", linewidth=3, hatch="x"
    )

    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("beyond-DLS stocks")

    custom_patch = Patch(facecolor="none", edgecolor="black")
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("additions to close gap")

    # Create a custom patch with a hatch pattern
    custom_patch2 = Patch(
        facecolor="none", edgecolor=None, hatch=".", label="additions to close gap"
    )
    legend_handles.append(custom_patch2)
    legend_labels.append("beyond-DLS")

    # Create custom markers using Unicode characters for converged cross
    custom_marker_cross = Line2D([0], [0], color="red", marker="$+$", markersize=12)
    legend_handles.append(custom_marker_cross)
    legend_labels.append("scenario-i: conver-\nged practices")

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames.append(custom_marker_a)
    legend_labels_frames.append("only DLS")
    legend_handles_frames.append(custom_marker_b)
    legend_labels_frames.append("beyond-DLS: regional")
    legend_handles_frames.append(custom_marker_c)
    legend_labels_frames.append("beyond-DLS: global")

    line_x_values = [177.36999999999998]  # Replace with actual x-values
    line_labels = ["scenario-iii: economy-wide stock requirement (global)"]

    # Plot horizontal lines and add labels (switched axes)
    for x_value, label in zip(line_x_values, line_labels):
        ax2.axhline(y=x_value, color="gray", linestyle="--", linewidth=2)
        ax2.text(
            x=ax2.get_xlim()[1] / 2,  # start of x-axis
            y=x_value + 6,  # aligned to the line's y-position
            s=label,
            verticalalignment="center",  # centered on the line
            horizontalalignment="left",  # aligned to the left of the plot
            fontsize=20,
        )

    titlea = "..(b) REGIONAL, Ø per capita (scale: regions' entire population')"

    ax2.set_title(titlea, fontsize=38, loc="left", pad=35)
    ax2.set_xlim(0 - 50000, cumulative_population[-1] + 50000)

    # Calculate midpoints of the bars for the y-axis labels
    midpoints = (cumulative_population[:-1] + cumulative_population[1:]) / 2

    # Set y-ticks and y-tick labels to the midpoints of the bars
    ax2.set_xticks(midpoints)
    # ax2.set_xticklabels(['Eastern Europe', 'Western Europe', ' Japan, Australia, NZ', 'North America', 'Form. Soviet Union', 'Centr. planned Asia', 'N. Africa & M. East', 'Latin America', 'Pacific Asia', 'South Asia', 'Subsaharan Africa'], fontsize=24)
    ax2.set_xticklabels(
        [
            "Centr.\n planned Asia",
            "Eastern Europe",
            "Western Europe",
            "Form.\n Soviet Union",
            "North America",
            " Japan,\n Australia, NZ",
            "Latin America",
            "Subsaharan\n Africa",
            "N. Africa &\n M. East",
            "Pacific Asia",
            "South Asia",
        ],
        fontsize=28,
    )
    for label in ax2.get_xticklabels():
        label.set_rotation(45)
    # Optionally, adjust the font size of the tick labels on the x-axis as well
    ax2.tick_params(axis="y", labelsize=26)
    ax2.set_ylim(0, 240)
    ax2.set_ylabel("tons / capita", fontsize=26)

    fig_title = (
        "Existings stocks providing DLS & beyond-DLS + additions to close DLS gaps..."
    )

    # reorder legend for product groups
    legend_handles1 = [
        legend_handles[0],
        legend_handles[1],
        legend_handles[2],
        legend_handles[3],
        legend_handles[4],
        legend_handles[5],
        legend_handles[6],
    ]
    legend_labels1 = [
        legend_labels[0],
        legend_labels[1],
        legend_labels[2],
        legend_labels[3],
        legend_labels[4],
        legend_labels[5],
        legend_labels[6],
    ]
    legend_handles2 = [handles[0], handles[1], handles[2], handles[3]]
    legend_labels2 = [labels[0], labels[1], labels[2], labels[3]]
    legend_handles3 = [legend_handles[7], legend_handles[8]]
    legend_labels3 = [legend_labels[7], legend_labels[8]]

    # legend = fig.legend(legend_handles, legend_labels, loc='upper center', bbox_to_anchor=(0.58, 0.6), ncol=6, fontsize=20, title='  Stock of material groups          Stock of product groups', title_fontsize=20)
    legend1 = fig.legend(
        legend_handles1,
        legend_labels1,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.91 + 0.035),
        ncol=1,
        fontsize=26,
        title="Stock of product groups",
        title_fontsize=26,
    )
    legend2 = fig.legend(
        legend_handles2,
        legend_labels2,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.75 + 0.03),
        ncol=1,
        fontsize=26,
        title="Stock of material groups",
        title_fontsize=26,
    )
    legend3 = fig.legend(
        legend_handles3,
        legend_labels3,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.65 + 0.025),
        ncol=1,
        fontsize=26,
        title="Other entries   ",
        title_fontsize=26,
    )

    for i in [legend1, legend2, legend3]:
        i.get_frame().set_edgecolor("grey")
        i.get_title().set_fontsize(26)
        i._legend_box.align = "left"  # Align title text
    plt.tight_layout()
    plt.subplots_adjust(hspace=0.3)
    plt.show()
    # fig.savefig('Fig2_DLSbeyond.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    plt.close("all")


def plot_bars_Vert_gap_headroom_two_subplots_doubleGlob_mod_eff_i(
    df1,
    df2,
    df3,
    df_add_beyond_stocks_regional,
    df_add_beyond_stocks_global,
    global_stocks,
    global_stock_prod,
    subcategory,
    line_y_values,
    line_labels,
    DLS_stock_thresh_eff_glob,
    DLS_stock_thresh_eff_cap,
):
    grouped_df1 = df1.groupby(["region", subcategory]).sum()
    grouped_df2 = df2.groupby(["region", subcategory]).sum()
    grouped_df3 = df3.set_index(["region", subcategory]).groupby("region").sum()

    grouped_df1 = pd.DataFrame(grouped_df1.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df2 = pd.DataFrame(grouped_df2.sum(axis=1)).rename(columns={0: "stock"})
    grouped_df3 = pd.DataFrame(grouped_df3.sum(axis=1)).rename(columns={0: "line"})

    # Creating a single dataframe for plotting
    combined_df = pd.concat([grouped_df1, grouped_df2], axis=1, keys=["df1", "df2"])
    combined_df.columns = combined_df.columns.map("_".join)
    combined_df = combined_df / 1e3  # Convert to t/cap
    combined_df.reset_index(inplace=True)
    region_order = [
        "CPA",
        "EEU",
        "WEU",
        "FSU",
        "NAM",
        "PAO",
        "LAM",
        "AFR",
        "MEA",
        "PAS",
        "SAS",
    ]
    sector_order = [
        "res_buildings",
        "nonres_buildings",
        "other_construction",
        "road_rail",
        "transport_machinery",
        "machinery",
        "other",
    ]

    # Create a mapping of region to the custom order
    order_mapping = {region: i for i, region in enumerate(region_order)}
    # Create a mapping of sector to the custom order
    order_mapping_sector = {sector: i for i, sector in enumerate(sector_order)}

    # Map the 'region' column to a new column that represents its order
    combined_df["region_order"] = combined_df["region"].map(order_mapping)
    # Sort the DataFrame by the new 'region_order' column, then drop it

    # Map the 'sector' column to a new column that represents its order
    combined_df["sector_order"] = combined_df["sector"].map(order_mapping_sector)
    # Sort the DataFrame by the new 'sector_order' column, then drop it
    combined_df_sorted = combined_df.sort_values(
        by=["region_order", "sector_order"]
    ).drop(columns=["sector_order"])

    combined_df = combined_df_sorted.set_index(["region", "sector"])

    # also reorder global_stock_prod and global_stock
    global_stock_prod = global_stock_prod.reindex(sector_order)
    global_stocks = global_stocks.reindex(["minerals", "metals", "biomass", "fossils"])

    # Create the figure
    fig = plt.figure(figsize=(24, 28))

    # Create a 2-row, 2-column grid
    gs = gridspec.GridSpec(2, 2, height_ratios=[2, 2.1], width_ratios=[4.5, 2])

    # First panel spans only the first column of the first row (half-width)
    ax0 = fig.add_subplot(gs[0, 0])

    # Second panel spans both columns of the second row (full width)
    ax2 = fig.add_subplot(gs[1, :])

    # Optional: Hide the unused top-right space (gs[0, 1])
    fig.add_subplot(gs[0, 1]).axis("off")

    global_stocks_products = global_stock_prod.groupby("region", axis=1).sum()

    from matplotlib.patches import FancyBboxPatch

    # Materials and their colors
    materials = global_stocks.index
    sectors = global_stocks_products.index
    colors = [
        (0.6, 0.58, 0.58),
        (0.18, 0.46, 0.71),
        (0.44, 0.68, 0.28),
        (0.93, 0.40, 0.19),
    ]
    colors2 = plt.cm.get_cmap("Pastel2", len(sectors))
    # replace grey color in pastel2 so it is not confuesd with grey in variable colors
    colors_array = colors2(np.arange(len(sectors)))
    replacement_color = [1.0, 0.5, 0.5, 1.0]  # Replace with a red-like RGBA
    colors_array[-1] = replacement_color  # Replace gray with your chosen color
    from matplotlib.colors import ListedColormap

    custom_cmap = ListedColormap(colors_array)
    colors2 = custom_cmap
    # rearrange color order:
    # Extract the colors as an array
    colors3 = colors2(np.arange(colors2.N))
    # Define your custom order for the colors (example: swap the first and last color)
    custom_order = [4, 1, 6, 5, 3, 0, 2]  # Adjust based on the desired order
    # Reorder the colors based on the custom order
    reordered_colors = colors3[custom_order]
    # Create a new colormap with the reordered colors
    colors2 = ListedColormap(reordered_colors)

    # Adjust hatching patterns for keys
    key_hatch = {"DLS_prov": "", "beyond_DLS": "."}

    # adjust so that values are in Gigatons
    global_stocks_products = global_stocks_products / 1e9
    global_stocks = global_stocks / 1e9
    DLS_stock_thresh_eff_glob = DLS_stock_thresh_eff_glob / 1e9

    double_bar = 0.4

    # Function to create rounded vertical bars
    def plot_rounded_barv(ax, x, height, bottom=0, color=None, hatch=None, alpha=None):
        boxstyle = "round,pad=0"
        bar_width = 0.175  # Consistent width for all bars

        # Outline with opaque black edge
        outline_bbox = FancyBboxPatch(
            (x - bar_width, bottom),
            bar_width * 2,
            height,
            boxstyle=boxstyle,
            ec="black",
            fc="none",
            hatch=hatch,
            lw=1.5,
        )
        ax.add_patch(outline_bbox)

        # Filled bar with transparency if color is provided
        if color is not None:
            fill_bbox = FancyBboxPatch(
                (x - bar_width, bottom),
                bar_width * 2,
                height,
                boxstyle=boxstyle,
                ec="none",
                fc=color,
                hatch=hatch,
                lw=0,
                alpha=alpha,
            )
            ax.add_patch(fill_bbox)

    left_bar_x = -0.425  # shift further left
    right_bar_x = 0.075  # closer to center

    # Plot bars for global stocks of products (now vertical)
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, sector in enumerate(sectors):
            value = global_stocks_products.loc[sector, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=colors2(mat_idx),
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barv(
                    ax0,
                    left_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            bottom += value

    # Plot bars for global stocks of materials (now vertical)
    bottom = 0
    for key in ["DLS_prov", "beyond_DLS"]:
        for mat_idx, material in enumerate(materials):
            value = global_stocks.loc[material, key]
            hatch = key_hatch.get(key)
            if key == "DLS_prov":
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=1,
                )
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            if key == "beyond_DLS":
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=colors[mat_idx],
                    hatch=hatch,
                    alpha=0.6,
                )
                plot_rounded_barv(
                    ax0,
                    right_bar_x,
                    value,
                    bottom=bottom,
                    color=None,
                    hatch=hatch,
                    alpha=1,
                )
            bottom += value

    bottom_save = bottom

    # Add explicit legend entries for materials
    handles = [
        FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.1", fc=color, ec="black")
        for color in colors
    ]
    labels = list(materials)  # Convert to a list

    # Plot the gap keys (now vertical bars)
    keys_to_plot_target = ["gap_targeted"]
    letters = ["i"]

    for idx, key in enumerate(keys_to_plot_target):
        value = global_stocks[key].sum()
        plot_rounded_barv(
            ax0, left_bar_x, value, bottom=bottom, color="white", hatch="None"
        )
        plot_rounded_barv(ax0, left_bar_x, value, bottom=bottom, color=None, hatch=None)
        ax0.text(
            left_bar_x,
            bottom + value / 2,
            letters[idx],
            ha="center",
            va="center",
            fontsize=26,
        )
        if key == "gap_targeted":
            ax0.plot(
                left_bar_x,
                bottom + DLS_stock_thresh_eff_glob,
                "+",
                color="red",
                markersize=24,
                markeredgewidth=2,
            )
        bottom += value

    # Plot for the convergence gap (also vertical)
    keys_to_plot_all = ["gap_targeted"]
    bottom = bottom_save

    for idx, key in enumerate(keys_to_plot_all):
        value = global_stocks[key].sum()
        plot_rounded_barv(
            ax0, right_bar_x, value, bottom=bottom, color="white", hatch=None
        )
        plot_rounded_barv(
            ax0, right_bar_x, value, bottom=bottom, color=None, hatch=None
        )
        ax0.text(
            right_bar_x,
            bottom + value / 2,
            letters[idx],
            ha="center",
            va="center",
            fontsize=24,
        )
        if key == "gap_targeted":
            ax0.plot(
                right_bar_x,
                bottom + DLS_stock_thresh_eff_glob,
                "+",
                color="red",
                markersize=24,
                markeredgewidth=2,
            )
        bottom += value

    # Set x-ticks to indicate the two "positions" (product vs material logic)
    ax0.set_xticks([left_bar_x, right_bar_x])
    ax0.set_xticklabels(["by product", "by material"], fontsize=26)
    for label in ax0.get_xticklabels():
        label.set_rotation(0)
    # Vertical axis is now the value scale
    ax0.set_ylim(0, 1400)
    ax0.set_xlim(-1, 1)

    # Adjust tick sizes and labels
    ax0.tick_params(axis="y", labelsize=26)
    ax0.tick_params(axis="x", labelsize=26)

    # Y-axis label replaces former X-axis label
    ax0.set_ylabel("[Gigatons = 10^9 tons]", fontsize=26, labelpad=20)

    ax0.ticklabel_format(style="plain", axis="y")

    # Brackets now apply along y-axis (vertical stacking) — no change needed unless you want to flip the orientation
    y_bracket_position = 0.68  # Remains meaningful vertically
    y_bracket_position_below = (
        -0.67
    )  # These would be x-coordinates if you want to flip them horizontally
    axlim = ax0.get_ylim()

    def draw_vertical_bracket(
        ax, y_start, y_end, x_position, label, label_offset=0.1, bracket_width=0.1
    ):
        """
        Draws a vertical bracket from y_start to y_end at a given x_position.
        """
        # Draw vertical line (bracket spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Draw horizontal tips
        ax.plot(
            [x_position - bracket_width / 2, x_position + bracket_width / 2],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width / 2, x_position + bracket_width / 2],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Add label
        ax.text(
            x_position + bracket_width * 1.2,
            (y_start + y_end) / 2,
            label,
            ha="left",
            va="center",
            fontsize=20,
        )

    def draw_vertical_bracket_right_label(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point LEFT (towards left side)
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on right
        ax.text(
            x_position + label_offset,
            (y_start + y_end) / 2,
            label,
            ha="left",
            va="center",
            fontsize=26,
        )

    def draw_vertical_bracket_right_label_up(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point LEFT (towards left side)
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position - bracket_width, x_position],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on right
        ax.text(
            x_position + label_offset,
            (y_start + y_end) / 1.8,
            label,
            ha="left",
            va="center",
            fontsize=26,
        )

    def draw_vertical_bracket_left_label(
        ax, y_start, y_end, x_position, label, label_offset=0.05, bracket_width=0.05
    ):
        # Vertical line (spine)
        ax.plot([x_position, x_position], [y_start, y_end], color="black", lw=1.5)
        # Tips point RIGHT (towards right side)
        ax.plot(
            [x_position, x_position + bracket_width],
            [y_start, y_start],
            color="black",
            lw=1.5,
        )
        ax.plot(
            [x_position, x_position + bracket_width],
            [y_end, y_end],
            color="black",
            lw=1.5,
        )
        # Label on left
        ax.text(
            x_position - label_offset,
            (y_start + y_end) / 2,
            label,
            ha="right",
            va="center",
            fontsize=26,
            rotation=90,
        )

    x_bracket_position = 0.5  # Adjust this to align with your bar or region

    space_keeper = 10
    draw_vertical_bracket_left_label(
        ax0,
        0.1 + space_keeper,
        699.4 - space_keeper,
        -0.7,
        label="existing economy-wide\n stocks in 2015",
    )
    # draw_vertical_bracket_left_label(ax0, 700 +space_keeper,  1399 -space_keeper, -0.7, label="net additions to stocks\n to close DLS gaps")

    draw_vertical_bracket_right_label(
        ax0, 0.1 + space_keeper, 193 - space_keeper, 0.35, label="existing\n DLS stocks"
    )
    draw_vertical_bracket_right_label(
        ax0,
        194.2 + space_keeper,
        699.4 - space_keeper,
        0.35,
        label="existing\n beyond-DLS\n stocks",
    )
    draw_vertical_bracket_right_label(
        ax0,
        699.4 + space_keeper,
        699.4 + 82.2 - space_keeper,
        0.5,
        label="DLS stock gap-\n scenario-i)",
    )

    # Set final plot title
    ax0.set_title(
        "(a) Existings DLS & beyond-DLS stocks + additions closing DLS gaps:\n    ...GLOBAL, at scale",
        fontsize=38,
        loc="left",
        pad=35,
    )

    # Unique regions and dimensions for plotting
    regions = combined_df.index.get_level_values(0).unique()
    # dimensions = combined_df.index.get_level_values(1).unique()
    dimensions = sector_order

    # Assigning distinct colors and hatch patterns to each dimension and creating legend handles
    colors = plt.cm.get_cmap("Pastel2", len(dimensions))
    colors = colors2
    # hatches = ['x','x','x','x','x','x','x','x','x','x']
    hatches = ["", "", "", "", "", "", "", "", "", ""]
    hatches = [".", ".", ".", ".", ".", ".", ".", ".", ".", "."]
    legend_handles = []

    # Calculate cumulative population
    cumulative_population = np.cumsum(
        [0]
        + [
            1.55716e06,
            121426,
            497389,
            288469,
            360286,
            156532,
            618764,
            954920,
            469410,
            579769,
            1.74936e06,
        ]
    )

    # Sort regions
    regions = region_order

    # Plot each region
    for i, region in enumerate(regions):
        bottom_df1 = 0  # Bottom for df1 stack
        bottom_df2 = 0  # Bottom for df2 stack (starts at top of df1 stack)
        bottom_df3 = 0
        bottom_df4 = 0

        # Plot df1 and calculate total df1 value for the region
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df1_value = combined_df.loc[(region, dimension), "df1_stock"]
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1])
                / 2,  # x: center of region
                df1_value,  # height
                width=(
                    cumulative_population[i + 1] - cumulative_population[i]
                ),  # bar width = population share
                color=color,
                edgecolor="black",
                linewidth=1.5,
                label=f"{dimension} (df1)",
                bottom=bottom_df1,
                alpha=1,
            )
            bottom_df1 += df1_value

            # Add legend handle for the first region
            if i == 0:
                legend_handles.append(
                    plt.Rectangle((0, 0), 1, 1, color=color, alpha=1, edgecolor="black")
                )

        # Plot df2 stacked on top of df1
        for j, dimension in enumerate(dimensions):
            color = colors(j)
            hatch = hatches[j]
            df2_value = combined_df.loc[(region, dimension), "df2_stock"]
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                width=(cumulative_population[i + 1] - cumulative_population[i]),
                color=color,
                edgecolor="none",
                linewidth=1.5,
                hatch=hatch,
                label=f"{dimension} (df2)",
                bottom=bottom_df1 + bottom_df2,
                alpha=0.6,
            )
            # Overplot edges for hatches
            ax2.bar(
                (cumulative_population[i] + cumulative_population[i + 1]) / 2,
                df2_value,
                width=(cumulative_population[i + 1] - cumulative_population[i]),
                color="none",
                edgecolor="black",
                linewidth=1.5,
                hatch=hatch,
                bottom=bottom_df1 + bottom_df2,
                alpha=1,
            )
            bottom_df2 += df2_value

        # df3 additional stock beyond df1+df2
        df3_value = (
            grouped_df3.loc[region, "line"] / 1e3
            - combined_df.loc[(region), "df1_stock"].sum()
        )
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df3_value,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            label=f"{dimension} (df3)",
            bottom=bottom_df1 + bottom_df2,
        )

        # Red cross for DLS threshold
        ax2.plot(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            bottom_df1 + bottom_df2 + DLS_stock_thresh_eff_cap.loc[region] / 1e3,
            "+",
            color="red",
            markersize=16,
            markeredgewidth=2,
        )

        # Annotate 'i'
        center_x = (cumulative_population[i] + cumulative_population[i + 1]) / 2
        center_y = bottom_df1 + bottom_df2 + df3_value / 2
        ax2.text(center_x, center_y, "i", ha="center", va="center", fontsize=24)

        bottom_df3 += df3_value

        # df4 - added beyond stocks regional
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_regional[region] / 1e3,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            bottom=bottom_df1 + bottom_df2 + bottom_df3,
        )

        center_y = (
            bottom_df1
            + bottom_df2
            + bottom_df3
            + df_add_beyond_stocks_regional[region] / 1e3 / 2
        )
        ax2.text(center_x, center_y, "+ii", ha="center", va="center", fontsize=24)

        bottom_df4 += df_add_beyond_stocks_regional[region] / 1e3

        # df5 - added beyond stocks global
        ax2.bar(
            (cumulative_population[i] + cumulative_population[i + 1]) / 2,
            df_add_beyond_stocks_global[region] / 1e3,
            width=(cumulative_population[i + 1] - cumulative_population[i]),
            color="none",
            edgecolor="black",
            linewidth=1.5,
            bottom=bottom_df1 + bottom_df2 + bottom_df3 + bottom_df4,
        )

        if df_add_beyond_stocks_global[region] > 0:
            center_y = (
                bottom_df1
                + bottom_df2
                + bottom_df3
                + bottom_df4
                + df_add_beyond_stocks_global[region] / 1e3 / 2
            )
            ax2.text(center_x, center_y, "+iii", ha="center", va="center", fontsize=24)

    # Legend labels
    legend_labels = [dim for dim in dimensions]

    # Dictionary mapping old labels to new labels
    replacement_dict = {
        "nonres_buildings": "nonresid. buildings",
        "other_construction": "civil engineering",
        "res_buildings": "resid. buildings",
        "road_rail": "road & rail",
        "transport_machinery": "transport vehicles",
    }

    # Replace labels using list comprehension
    legend_labels = [
        replacement_dict[label] if label in replacement_dict else label
        for label in legend_labels
    ]

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames = []
    legend_labels_frames = []

    custom_patch = Patch(facecolor="orange", alpha=1, edgecolor="black", linewidth=3)
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("DLS stocks")

    import matplotlib.patches as patches

    # custom_patch = Patch(facecolor='orange', alpha=0.4, edgecolor='black', linewidth=3, hatch='x')
    custom_patch = patches.Rectangle(
        (0, 0), 1, 1, facecolor="orange", edgecolor="black", linewidth=3, hatch="x"
    )

    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("beyond-DLS stocks")

    custom_patch = Patch(facecolor="none", edgecolor="black")
    legend_handles_frames.append(custom_patch)
    legend_labels_frames.append("additions to close gap")

    # Create a custom patch with a hatch pattern
    custom_patch2 = Patch(
        facecolor="none", edgecolor=None, hatch=".", label="additions to close gap"
    )
    legend_handles.append(custom_patch2)
    legend_labels.append("beyond-DLS")

    # Create custom markers using Unicode characters for converged cross
    custom_marker_cross = Line2D([0], [0], color="red", marker="$+$", markersize=12)
    legend_handles.append(custom_marker_cross)
    legend_labels.append("scenario-i: conver-\nged practices")

    # Create custom markers using Unicode characters
    custom_marker_a = Line2D([0], [0], color="black", marker="$i$", markersize=12)
    custom_marker_b = Line2D([0], [0], color="black", marker="$ii$", markersize=12)
    custom_marker_c = Line2D([0], [0], color="black", marker="$iii$", markersize=12)

    legend_handles_frames.append(custom_marker_a)
    legend_labels_frames.append("only DLS")
    legend_handles_frames.append(custom_marker_b)
    legend_labels_frames.append("beyond-DLS: regional")
    legend_handles_frames.append(custom_marker_c)
    legend_labels_frames.append("beyond-DLS: global")

    line_x_values = [177.36999999999998]  # Replace with actual x-values
    line_labels = ["scenario-iii: economy-wide stock requirement (global)"]

    # Plot horizontal lines and add labels (switched axes)
    for x_value, label in zip(line_x_values, line_labels):
        ax2.axhline(y=x_value, color="gray", linestyle="--", linewidth=2)
        ax2.text(
            x=ax2.get_xlim()[1] / 2,  # start of x-axis
            y=x_value + 6,  # aligned to the line's y-position
            s=label,
            verticalalignment="center",  # centered on the line
            horizontalalignment="left",  # aligned to the left of the plot
            fontsize=20,
        )

    titlea = "..(b) REGIONAL, Ø per capita (scale: regions' entire population')"

    ax2.set_title(titlea, fontsize=38, loc="left", pad=35)
    ax2.set_xlim(0 - 50000, cumulative_population[-1] + 50000)

    # Calculate midpoints of the bars for the y-axis labels
    midpoints = (cumulative_population[:-1] + cumulative_population[1:]) / 2

    # Set y-ticks and y-tick labels to the midpoints of the bars
    ax2.set_xticks(midpoints)
    # ax2.set_xticklabels(['Eastern Europe', 'Western Europe', ' Japan, Australia, NZ', 'North America', 'Form. Soviet Union', 'Centr. planned Asia', 'N. Africa & M. East', 'Latin America', 'Pacific Asia', 'South Asia', 'Subsaharan Africa'], fontsize=24)
    ax2.set_xticklabels(
        [
            "Centr.\n planned Asia",
            "Eastern Europe",
            "Western Europe",
            "Form.\n Soviet Union",
            "North America",
            " Japan,\n Australia, NZ",
            "Latin America",
            "Subsaharan\n Africa",
            "N. Africa &\n M. East",
            "Pacific Asia",
            "South Asia",
        ],
        fontsize=28,
    )
    for label in ax2.get_xticklabels():
        label.set_rotation(45)
    # Optionally, adjust the font size of the tick labels on the x-axis as well
    ax2.tick_params(axis="y", labelsize=26)
    ax2.set_ylim(0, 240)
    ax2.set_ylabel("tons / capita", fontsize=26)

    fig_title = (
        "Existings stocks providing DLS & beyond-DLS + additions to close DLS gaps..."
    )

    # reorder legend for product groups
    legend_handles1 = [
        legend_handles[0],
        legend_handles[1],
        legend_handles[2],
        legend_handles[3],
        legend_handles[4],
        legend_handles[5],
        legend_handles[6],
    ]
    legend_labels1 = [
        legend_labels[0],
        legend_labels[1],
        legend_labels[2],
        legend_labels[3],
        legend_labels[4],
        legend_labels[5],
        legend_labels[6],
    ]
    legend_handles2 = [handles[0], handles[1], handles[2], handles[3]]
    legend_labels2 = [labels[0], labels[1], labels[2], labels[3]]
    legend_handles3 = [legend_handles[7], legend_handles[8]]
    legend_labels3 = [legend_labels[7], legend_labels[8]]

    # legend = fig.legend(legend_handles, legend_labels, loc='upper center', bbox_to_anchor=(0.58, 0.6), ncol=6, fontsize=20, title='  Stock of material groups          Stock of product groups', title_fontsize=20)
    legend1 = fig.legend(
        legend_handles1,
        legend_labels1,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.91 + 0.035),
        ncol=1,
        fontsize=26,
        title="Stock of product groups",
        title_fontsize=26,
    )
    legend2 = fig.legend(
        legend_handles2,
        legend_labels2,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.75 + 0.03),
        ncol=1,
        fontsize=26,
        title="Stock of material groups",
        title_fontsize=26,
    )
    legend3 = fig.legend(
        legend_handles3,
        legend_labels3,
        loc="upper center",
        bbox_to_anchor=(0.85, 0.65 + 0.025),
        ncol=1,
        fontsize=26,
        title="Other entries   ",
        title_fontsize=26,
    )

    for i in [legend1, legend2, legend3]:
        i.get_frame().set_edgecolor("grey")
        i.get_title().set_fontsize(26)
        i._legend_box.align = "left"  # Align title text
    plt.tight_layout()
    plt.subplots_adjust(hspace=0.3)
    plt.show()
    # fig.savefig('Fig2_DLSbeyond.tif', format='tif', dpi =300, bbox_inches='tight', pad_inches=0)
    plt.close("all")


def share_plot(df, regions):
    # Define region and sector name mappings
    region_mapping = {
        "AFR": "Sub-Saharan Africa",
        "CPA": "Centrally planned Asia",
        "EEU": "Eastern Europe",
        "FSU": "Former Soviet Union",
        "LAM": "Latin America",
        "MEA": "N. Africa & Middle East",
        "NAM": "North America",
        "PAO": "Japan, Australia, New Zealand",
        "PAS": "Pacific Asia",
        "SAS": "South Asia",
        "WEU": "Western Europe",
    }

    sector_mapping = {
        "nonres_buildings": "nonresid. buildings",
        "other_construction": "civil engineering",
        "res_buildings": "resid. buildings",
        "road_rail": "road & rail",
        "transport_machinery": "transport vehicles",
    }

    # Replace region names in the list with their full names
    regions = [region_mapping.get(region, region) for region in regions]
    df = df.rename(region_mapping)
    df = df.rename(sector_mapping)

    # Define the custom colors for each sector from colors2 (using RGB tuples)
    colors2 = [
        [
            0.9882352941176471,
            0.9058823529411765,
            0.7470588235294118,
            1.0,
        ],  # Light yellow
        [0.9372549019607843, 0.6745098039215687, 0.7490196078431373, 1.0],  # Light rose
        [0.8666666666666667, 0.6784313725490196, 0.7725490196078432, 1.0],  # Light pink
        [
            0.7764705882352941,
            0.7294117647058823,
            0.5803921568627451,
            1.0,
        ],  # Pastel brown
        [0.7019607843137254, 0.8862745098039215, 0.803921568627451, 1.0],  # Light green
        [0.592156862745098, 0.9058823529411765, 0.8274509803921568, 1.0],  # Mint green
        [0.7490196078431373, 0.6823529411764706, 0.8313725490196079, 1.0],  # Lavender
        [
            0.9882352941176471,
            0.9058823529411765,
            0.7470588235294118,
            0,
        ],  # Light yellow with reduced alpha
        [
            0.9372549019607843,
            0.6745098039215687,
            0.7490196078431373,
            0,
        ],  # Light rose with reduced alpha
        [
            0.8666666666666667,
            0.6784313725490196,
            0.7725490196078432,
            0,
        ],  # Light pink with reduced alpha
        [
            0.7764705882352941,
            0.7294117647058823,
            0.5803921568627451,
            0,
        ],  # Pastel brown with reduced alpha
        [
            0.7019607843137254,
            0.8862745098039215,
            0.803921568627451,
            0,
        ],  # Light green with reduced alpha
        [
            0.592156862745098,
            0.9058823529411765,
            0.8274509803921568,
            0,
        ],  # Mint green with reduced alpha
        [
            0.7490196078431373,
            0.6823529411764706,
            0.8313725490196079,
            0,
        ],  # Lavender with reduced alpha
    ]

    custom_row_order = [
        "resid. buildings",
        "nonresid. buildings",
        "civil engineering",
        "road & rail",
        "transport vehicles",
        "machinery",
        "other",
    ]
    # Calculate the number of rows and columns for the subplots
    rows = (len(regions) // 3) + (1 if len(regions) % 3 != 0 else 0)
    cols = 4

    # Create subplots for each region
    fig, axes = plt.subplots(nrows=rows, ncols=cols, figsize=(20, rows * 3))

    # Flatten the axes array to handle both single and multiple subplots easily
    axes = axes.flatten()

    # Plot each region in a separate subplot
    for i, region in enumerate(regions):
        region_data = df.xs(region, level="region")
        region_data = region_data[["bDLS", "DLS gap"]]
        # Reorder the rows of the region_data according to custom_sector_order
        region_data = region_data.loc[custom_row_order]

        hatch_patterns = ["x", "", "", "", "x", "", ""]

        # Create the stacked bar plot for the current region
        region_plot = region_data.plot(
            kind="bar", stacked=True, ax=axes[i], width=0.8, legend=False
        )

        # Manually modify the colors of the bars (patches)
        for j, patch in enumerate(
            region_plot.patches
        ):  # region_plot.patches contains all the bars (patches)
            # Assign color based on the sector
            sector_idx = j % len(colors2)  # Use modulo to loop through colors2
            patch.set_facecolor(colors2[sector_idx])  # Set the color of the bar

        # Get the patches (each segment in the stacked bar chart)
        patches = region_plot.patches

        # Iterate over each patch (segment in the stacked bar)
        for j, patch in enumerate(patches):
            if j < 7:
                patch.set_hatch(".")  # Apply the hatch pattern to every even patch

        # Rename sectors according to sector_mapping
        region_data.columns = [
            sector_mapping.get(col, col) for col in region_data.columns
        ]

        axes[i].set_title(f"{region}", fontsize=16)
        axes[i].set_xticklabels(region_data.index, rotation=45, ha="right", fontsize=16)

        # Only show x-axis labels on the bottom plots
        if i < 4:
            axes[i].set_xticklabels([])
            axes[i].tick_params(
                axis="x", which="both", bottom=False, top=False, labelbottom=False
            )
        axes[i].set_xlabel("")

        axes[i].axhline(y=0.5, color="gray", linewidth=1.5, linestyle="--")
        axes[i].set_yticks([0, 0.25, 0.5, 0.75, 1])
        axes[i].set_yticklabels(["only\ngap", "1:3", "1:1", "3:1", "no\ngap"])

    # Remove unused subplots (if there are fewer regions than the subplot grid)
    for j in range(i + 1, len(axes)):
        axes[j].axis("off")

    # Add a common x-label for the entire plot
    axes[0].set_ylabel("Proportions\n bDLs: DLS gap", fontsize=14)

    # Add a title to the entire figure
    plt.suptitle(
        "Proportions of DLS stock gap and existing beyond-DLS stocks (regions with gaps) ",
        fontsize=24,
        x=0.45,
        y=1,
    )

    # Show the plot
    plt.show()


def share_plot_converge(df, df_converge, regions):
    # Define region and sector name mappings
    region_mapping = {
        "AFR": "Sub-Saharan Africa",
        "CPA": "Centrally planned Asia",
        "EEU": "Eastern Europe",
        "FSU": "Former Soviet Union",
        "LAM": "Latin America",
        "MEA": "N. Africa & Middle East",
        "NAM": "North America",
        "PAO": "Japan, Australia, New Zealand",
        "PAS": "Pacific Asia",
        "SAS": "South Asia",
        "WEU": "Western Europe",
    }

    sector_mapping = {
        "nonres_buildings": "nonresid. buildings",
        "other_construction": "civil engineering",
        "res_buildings": "resid. buildings",
        "road_rail": "road & rail",
        "transport_machinery": "transport vehicles",
    }

    # Replace region names in the list with their full names
    regions = [region_mapping.get(region, region) for region in regions]
    df = df.copy()
    df.rename(region_mapping, inplace=True)
    df.rename(sector_mapping, inplace=True)
    df_converge.rename(region_mapping, inplace=True)
    df_converge.rename(sector_mapping, inplace=True)

    # Define the custom colors for each sector from colors2 (using RGB tuples)
    colors2 = [
        [0.996, 0.878, 0.298, 0.4],  # Light yellow
        [0.996, 0.792, 0.682, 0.7],  # Light rose
        [0.933, 0.325, 0.286, 0.4],  # Light pink
        [0.933, 0.890, 0.729, 0.7],  # Pastel brown
        [0.945, 0.99, 0.678, 0.7],  # Light green
        [0.710, 0.867, 0.835, 0.6],  # Mint green
        [0.678, 0.800, 0.886, 0.6],  # Lavender
        [
            0.9882352941176471,
            0.9058823529411765,
            0.7470588235294118,
            0,
        ],  # Light yellow with reduced alpha
        [
            0.9372549019607843,
            0.6745098039215687,
            0.7490196078431373,
            0,
        ],  # Light rose with reduced alpha
        [
            0.8666666666666667,
            0.6784313725490196,
            0.7725490196078432,
            0,
        ],  # Light pink with reduced alpha
        [
            0.7764705882352941,
            0.7294117647058823,
            0.5803921568627451,
            0,
        ],  # Pastel brown with reduced alpha
        [
            0.7019607843137254,
            0.8862745098039215,
            0.803921568627451,
            0,
        ],  # Light green with reduced alpha
        [
            0.592156862745098,
            0.9058823529411765,
            0.8274509803921568,
            0,
        ],  # Mint green with reduced alpha
        [
            0.7490196078431373,
            0.6823529411764706,
            0.8313725490196079,
            0,
        ],  # Lavender with reduced alpha
    ]

    custom_row_order = [
        "resid. buildings",
        "nonresid. buildings",
        "civil engineering",
        "road & rail",
        "transport vehicles",
        "machinery",
        "other",
    ]
    # Calculate the number of rows and columns for the subplots
    rows = (len(regions) // 3) + (1 if len(regions) % 3 != 0 else 0)
    cols = 4

    # Create subplots for each region
    fig, axes = plt.subplots(nrows=rows, ncols=cols, figsize=(20, rows * 3))

    # Flatten the axes array to handle both single and multiple subplots easily
    axes = axes.flatten()

    # Plot each region in a separate subplot
    for i, region in enumerate(regions):
        region_data = df.xs(region, level="region")
        region_data = region_data[["bDLS", "DLS gap"]]
        # Reorder the rows of the region_data according to custom_sector_order
        region_data = region_data.loc[custom_row_order]

        hatch_patterns = ["x", "", "", "", "x", "", ""]

        # Create the stacked bar plot for the current region
        region_plot = region_data.plot(
            kind="bar", stacked=True, ax=axes[i], width=0.8, legend=False
        )

        # Manually modify the colors of the bars (patches)
        for j, patch in enumerate(
            region_plot.patches
        ):  # region_plot.patches contains all the bars (patches)
            # Assign color based on the sector
            sector_idx = j % len(colors2)  # Use modulo to loop through colors2
            patch.set_facecolor(colors2[sector_idx])  # Set the color of the bar

        # Get the patches (each segment in the stacked bar chart)
        patches = region_plot.patches

        # Iterate over each patch (segment in the stacked bar)
        for j, patch in enumerate(patches):
            if j < 7:
                patch.set_hatch(".")  # Apply the hatch pattern to every even patch

        # Rename sectors according to sector_mapping
        region_data.columns = [
            sector_mapping.get(col, col) for col in region_data.columns
        ]

        axes[i].set_title(f"{region}", fontsize=16)
        axes[i].set_xticklabels(region_data.index, rotation=45, ha="right", fontsize=16)

        # Only show x-axis labels on the bottom plots
        if i < 8:
            axes[i].set_xticklabels([])
            axes[i].tick_params(
                axis="x", which="both", bottom=False, top=False, labelbottom=False
            )
        axes[i].set_xlabel("")

        axes[i].axhline(y=0.5, color="gray", linewidth=1.5, linestyle="--")
        axes[i].set_yticks([0, 0.25, 0.5, 0.75, 1])
        axes[i].set_yticklabels(["only\ngap", "1:3", "1:1", "3:1", "no\ngap"])

        # Plot the red 'x' markers for bDLS targets from df_converge
        bDLS_vals = df_converge.xs(region, level="region").loc[custom_row_order, "bDLS"]
        axes[i].plot(
            range(len(custom_row_order)),  # x-position = bar index
            bDLS_vals.values,  # y-position = bDLS target value
            marker="x",
            linestyle="none",
            color="red",
            markersize=8,
            label="bDLS target" if i == 0 else None,  # Add label only once for legend
        )

    # Remove unused subplots (if there are fewer regions than the subplot grid)
    for j in range(i + 1, len(axes)):
        axes[j].axis("off")

    # Add a common x-label for the entire plot
    axes[0].set_ylabel("Proportions\n bDLs: DLS gap", fontsize=14)

    import matplotlib.lines as mlines

    # Create a custom red cross legend handle
    red_cross = mlines.Line2D(
        [],
        [],
        color="red",
        marker="x",
        linestyle="None",
        markersize=8,
        label="proportions for converged practices",
    )

    # Add the legend below the entire figure
    fig.legend(
        handles=[red_cross],
        loc="lower center",
        bbox_to_anchor=(0.2, 0.11),  # adjust for spacing
        frameon=False,
        fontsize=14,
    )

    # Add a title to the entire figure
    plt.suptitle(
        "Proportions of DLS stock gap and existing beyond-DLS stocks",
        fontsize=24,
        x=0.35,
        y=0.98,
    )
    # Save plot
    #fig.savefig("Fig_4.pdf", bbox_inches='tight')
    # Show the plot
    plt.show()


def logistic_func_bounded(x, k, x0):
    L = 1  # Maximum value
    return L / (1 + np.exp(-k * (x - x0)))


def find_x_for_y(target_y, k, x0):
    # Define the function to find the root for
    func = lambda x: logistic_func_bounded(x, k, x0) - target_y
    # Use fsolve to find the x that gives the target_y
    x_target = fsolve(func, x0)  # Start searching from x0
    return x_target[0]


def singleplot_predef_data_satur_bounded(
    x, y_select, y_label, x_label, title, labels=None
):
    fig, axs = plt.subplots(nrows=1, ncols=1, figsize=(6, 6))
    axs.scatter(x, y_select, zorder=3)

    if labels is not None:
        for i, label in enumerate(labels):
            axs.annotate(label, (x[i], y_select[i]), zorder=3)

    # Fit the data with modified logistic function
    popt, _ = curve_fit(logistic_func_bounded, x, y_select)

    x_fit = np.linspace(min(x), max(x), 1000)
    y_fit = logistic_func_bounded(x_fit, *popt)
    axs.plot(
        x_fit,
        y_fit,
        "r-",
        label=f"y = 1 / (1 + e^{{-{popt[0]:.2f} (x - {popt[1]:.2f})}})",
        zorder=3,
    )

    # Color areas above and below the regression line
    color1 = "lightblue"  # Color above the line
    color2 = "lightgreen"  # Color below the line
    axs.fill_between(x_fit, y_fit, max(y_select), color=color1, zorder=1)
    axs.fill_between(x_fit, y_fit, min(y_select), color=color2, zorder=1)

    legend_handles = [
        Patch(facecolor=color1, label="High Performers"),
        Patch(facecolor=color2, label="Low Performers"),
    ]

    residuals = y_select - logistic_func_bounded(x, *popt)
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((y_select - np.mean(y_select)) ** 2)
    r2 = 1 - (ss_res / ss_tot)

    axs.text(0.1, 0.9, f"$R^2$ = {r2:.2f}", transform=axs.transAxes)

    x_90 = find_x_for_y(0.9, *popt)
    x_95 = find_x_for_y(0.95, *popt)
    x_99 = find_x_for_y(0.99, *popt)

    # Add vertical lines and text annotations for x_90 and x_95
    axs.axvline(x=x_90, color="green", linestyle="--")
    axs.axvline(x=x_95, color="blue", linestyle="--")
    axs.axvline(x=x_99, color="purple", linestyle="--")
    axs.text(
        x_90,
        0.5,
        f"x_90: {x_90:.2f}",
        rotation=90,
        verticalalignment="center",
        color="green",
    )
    axs.text(
        x_95,
        0.5,
        f"x_95: {x_95:.2f}",
        rotation=90,
        verticalalignment="center",
        color="blue",
    )
    axs.text(
        x_99,
        0.5,
        f"x_99: {x_99:.2f}",
        rotation=90,
        verticalalignment="center",
        color="purple",
    )

    axs.set_title(title, fontsize=12)
    axs.set_xlabel(x_label, fontsize=10)
    axs.set_ylabel(y_label, fontsize=10)
    axs.legend(loc="best")
    axs.legend(handles=legend_handles + [axs.lines[0]], loc="best")
    fig.tight_layout()

    plt.show()

    return x_90, x_95


def singleplot_predef_data_satur_bounded_popweight(
    x, y_select, y_label, x_label, title, populations, labels=None
):
    fig, axs = plt.subplots(nrows=1, ncols=1, figsize=(6, 6))

    # Scale dot sizes by population (square root of populations for better visual scaling)
    dot_sizes = (
        np.sqrt(populations) * 0.75
    )  # Scaling factor for visual size, adjust as needed

    # Scatter plot with population-scaled dot sizes
    scatter = axs.scatter(x, y_select, s=dot_sizes, zorder=3)

    # Annotate labels if provided
    if labels is not None:
        for i, label in enumerate(labels):
            axs.annotate(label, (x[i], y_select[i]), zorder=3)

    # Ensure populations is a numpy array and has the correct shape
    populations = np.array(populations)

    if populations.shape != y_select.shape:
        raise ValueError("Populations array must have the same shape as y_select")

    # Use population weights (larger populations = more influence)
    weights = 1 / populations  # Inverse of population size as weights

    # Perform weighted curve fitting
    popt, _ = curve_fit(logistic_func_bounded, x, y_select, sigma=weights)

    # Plot the fitted logistic curve
    x_fit = np.linspace(min(x), max(x), 1000)
    y_fit = logistic_func_bounded(x_fit, *popt)
    axs.plot(
        x_fit,
        y_fit,
        "r-",
        label=f"y = 1 / (1 + e^{{-{popt[0]:.2f} (x - {popt[1]:.2f})}})",
        zorder=3,
    )

    # Color areas above and below the regression line
    color1 = "lightblue"  # Color above the line
    color2 = "lightgreen"  # Color below the line
    axs.fill_between(x_fit, y_fit, max(y_select), color=color1, zorder=1)
    axs.fill_between(x_fit, y_fit, min(y_select), color=color2, zorder=1)

    legend_handles = [
        Patch(facecolor=color1, label="High Performers"),
        Patch(facecolor=color2, label="Low Performers"),
    ]

    # Calculate residuals and R^2
    residuals = y_select - logistic_func_bounded(x, *popt)
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((y_select - np.mean(y_select)) ** 2)
    r2 = 1 - (ss_res / ss_tot)

    axs.text(0.1, 0.9, f"$R^2$ = {r2:.2f}", transform=axs.transAxes)

    # Find x-values for specific y-values (90%, 95%, 99%)
    x_90 = find_x_for_y(0.9, *popt)
    x_95 = find_x_for_y(0.95, *popt)
    x_99 = find_x_for_y(0.99, *popt)

    # Add vertical lines and text annotations for x_90, x_95, x_99
    axs.axvline(x=x_90, color="green", linestyle="--")
    axs.axvline(x=x_95, color="blue", linestyle="--")
    axs.axvline(x=x_99, color="purple", linestyle="--")
    axs.text(
        x_90,
        0.5,
        f"x_90: {x_90:.2f}",
        rotation=90,
        verticalalignment="center",
        color="green",
    )
    axs.text(
        x_95,
        0.5,
        f"x_95: {x_95:.2f}",
        rotation=90,
        verticalalignment="center",
        color="blue",
    )
    axs.text(
        x_99,
        0.5,
        f"x_99: {x_99:.2f}",
        rotation=90,
        verticalalignment="center",
        color="purple",
    )

    # Set title, labels, and legend
    axs.set_title(title, fontsize=12)
    axs.set_xlabel(x_label, fontsize=10)
    axs.set_ylabel(y_label, fontsize=10)
    axs.legend(loc="best")
    axs.legend(handles=legend_handles + [axs.lines[0]], loc="best")
    fig.tight_layout()

    plt.show()

    return x_90, x_95


def singleplot_predef_data_2col(x, y_select, y_label, x_label, title):
    fig, axs = plt.subplots(nrows=1, ncols=1, figsize=(6, 6))
    # x is a DataFrame with two columns
    # Scatter plot for each column in x
    colors = ["blue", "red"]  # Define colors for the scatter plots
    for i, col in enumerate(x.columns):
        axs.scatter(x[col], y_select, color=colors[i], label=col)
    axs.set_title(title, fontsize=12)
    axs.set_xlabel(x_label, fontsize=10)
    axs.set_ylabel(y_label, fontsize=10)
    axs.legend()  # Add a legend to distinguish the columns
    fig.tight_layout()
    plt.show()


def singleplot_predef_data_3col(x, y_select, y_label, x_label, title):
    fig, axs = plt.subplots(nrows=1, ncols=1, figsize=(6, 6))
    # x is a DataFrame with two columns
    # Scatter plot for each column in x
    colors = ["blue", "red", "green"]  # Define colors for the scatter plots
    for i, col in enumerate(x.columns):
        axs.scatter(x[col], y_select, color=colors[i], label=col)
    axs.set_title(title, fontsize=12)
    axs.set_xlabel(x_label, fontsize=10)
    axs.set_ylabel(y_label, fontsize=10)
    axs.legend()  # Add a legend to distinguish the columns
    fig.tight_layout()
    plt.show()


def extrapolate_NAS(df, additional_years):
    # Ensure the index and values are numeric
    df = df.apply(pd.to_numeric, errors="coerce")
    df.index = df.index.set_levels(
        [df.index.levels[0], pd.to_numeric(df.index.levels[1])]
    )

    # Get unique regions
    regions = df.index.get_level_values("region").unique()

    # Prepare a list to collect extrapolated data
    extrapolated_data = []

    for region in regions:
        # Select the data for the region
        region_data = df.xs(region, level="region")

        # Perform extrapolation for each value column
        for col in df.columns:
            if region == "CPA":
                # Special handling for 'CPA'
                # Reduce values by 1% per year from 2016 onwards
                for year in additional_years:
                    if year >= 2016:
                        # Calculate the number of years since 2016
                        years_since_2016 = year - 2016
                        # Apply the 1% reduction per year
                        extrapolated_value = region_data[col].iloc[-1] * (
                            0.99**years_since_2016
                        )
                    else:
                        # For years before 2016, perform linear extrapolation
                        x = region_data.index.values
                        y = region_data[col].values
                        coefficients = np.polyfit(x, y, 1)
                        trend = np.poly1d(coefficients)
                        extrapolated_value = trend(year)

                    extrapolated_data.append((region, year, col, extrapolated_value))
            else:
                # Linear extrapolation for other regions
                x = region_data.index.values
                y = region_data[col].values
                coefficients = np.polyfit(x, y, 1)
                trend = np.poly1d(coefficients)

                # Extrapolate for additional years
                for year in additional_years:
                    extrapolated_value = trend(year)
                    extrapolated_data.append((region, year, col, extrapolated_value))

    # Create a DataFrame for the extrapolated data
    extrapolated_df = pd.DataFrame(
        extrapolated_data, columns=["region", "year", "variable", "value"]
    )
    extrapolated_df = extrapolated_df.pivot_table(
        index=["region", "year"], columns="variable", values="value"
    )

    # Combine the original DataFrame with the extrapolated DataFrame
    result_df = pd.concat([df, extrapolated_df])

    return result_df


def plot_timing_close_DLS_gaps(
    NAS_all_histprosp, NAS_all_hist, NAS_all_prosp, scenario_summary, material_as_str
):
    # Helper function to format billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e9)}B"

    # Set up the 2x2 grid
    fig, axs = plt.subplots(2, 2, figsize=(18, 14))

    # Plot 1: Historical NAS by world regions and world average
    df = NAS_all_hist[[material_as_str]]
    df = df[df.index.get_level_values(1) != 2016]
    df.index = df.index.set_levels(
        df.index.levels[0]
        .to_series()
        .replace(
            {
                "AFR": "Subs. Africa",
                "CPA": "Centr. Asia",
                "EEU": "Eastern Europe",
                "FSU": "Former Sov.Union",
                "LAM": "Latin America",
                "MEA": "N. Africa & M.East",
                "NAM": "North America",
                "PAO": "JP,AU,NZ",
                "PAS": "Pacific Asia",
                "SAS": "South Asia",
                "WEU": "Western Europe",
                "Global": "Global av.",
            }
        ),
        level=0,
    )

    regions = df.index.get_level_values("region").unique()
    columns = df.columns

    ax = axs[0, 0]
    for region in regions:
        df_region = df.xs(region, level="region")
        for column in columns:
            if region == "Global av.":
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=8,
                )
            if region in [
                "Centr. Asia",
                "Subs. Africa",
                "South Asia",
                "Pacific Asia",
                "Latin America",
                "N. Africa & M.East",
            ]:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="purple",
                )
            else:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="grey",
                )

            if region in ["Centr. Asia", "Pacific Asia", "Subs. Africa", "South Asia"]:
                last_year = df_region.index[-7] + 0.1
                last_value = df_region[column].iloc[-1] - 0.1
                ax.text(last_year, last_value, region, fontsize=15, ha="left")
            if region in ["N. Africa & M.East"]:
                last_year = df_region.index[-12] + 0.15
                last_value = df_region[column].iloc[-2] + 0.15
                ax.text(last_year, last_value, region, fontsize=15, ha="left")
            if region in ["Former Sov.Union"]:
                last_year = df_region.index[-17] + 0.1
                last_value = df_region[column].iloc[-17] + 0.1
                ax.text(last_year, last_value, region, fontsize=15, ha="left")

    ax.set_ylabel("tons / capita/year", fontsize=22)
    ax.set_ylim(0, 13)
    ax.set_title(
        r"(a) $\bf{Historical}$ speed of new construction", fontsize=25, pad=10
    )
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)

    # Plot 2: Prospective NAS by world regions and world average
    NAS_all_prosp = NAS_all_histprosp[
        NAS_all_histprosp.index.get_level_values(1).astype(int) > 2016
    ].copy()
    #! set negative NAS to 0 (assuming that no negative NAS exist)
    NAS_all_prosp[NAS_all_prosp < 0] = 0

    df = NAS_all_prosp[[material_as_str]]
    df.index = df.index.set_levels(
        df.index.levels[0]
        .to_series()
        .replace(
            {
                "AFR": "Subs. Africa",
                "CPA": "Centr. Asia",
                "EEU": "Eastern Europe",
                "FSU": "Former Sov.Union",
                "LAM": "Latin America",
                "MEA": "N. Africa & M.East",
                "NAM": "North America",
                "PAO": "JP,AU,NZ",
                "PAS": "Pacific Asia",
                "SAS": "South Asia",
                "WEU": "Western Europe",
                "Global": "Global av.",
            }
        ),
        level=0,
    )

    regions = df.index.get_level_values(0).unique()
    columns = df.columns

    ax = axs[0, 1]
    for region in regions:
        df_region = df.xs(region, level="region")
        for column in columns:
            if region == "Global av.":
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=8,
                )
            if region in [
                "Subs. Africa",
                "South Asia",
                "Pacific Asia",
                "Latin America",
                "N. Africa & M.East",
                "Centr. Asia",
            ]:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="purple",
                )
            else:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="grey",
                )

            if region in ["Subs. Africa", "South Asia", "Pacific Asia"]:
                last_year = df_region.index[-7] + 0.1
                last_value = df_region[column].iloc[-1]
                ax.text(last_year, last_value, region, fontsize=15, ha="left")

            if region in ["Latin America"]:
                last_year = df_region.index[-7] + 0.1
                last_value = df_region[column].iloc[-1] - 0.6
                ax.text(last_year, last_value, region, fontsize=15, ha="left")

            if region in ["North America"]:
                last_year = df_region.index[-21] + 0.1
                first_value = 0.55
                ax.text(
                    last_year,
                    first_value,
                    "North America, Europe, F.Sov.Union, JP/AU/NZ",
                    fontsize=15,
                    ha="left",
                )

            if region in ["Centr. Asia"]:
                last_year = df_region.index[-34] + 0.1
                first_value = 12
                ax.text(
                    last_year,
                    first_value,
                    "Centr. Asia (exception: -1%/a)",
                    fontsize=15,
                    ha="left",
                )

            if region in ["N. Africa & M.East"]:
                last_year = df_region.index[-34] + 0.1
                first_value = df_region[column].iloc[1] + 0.8
                ax.text(last_year, first_value, region, fontsize=15, ha="left")

    ax.set_title(r"(b) $\bf{Prospective}$ speed (extrapolated)", fontsize=25, pad=10)
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)
    ax.set_ylim(0, 13)
    ax.set_yticks([])

    # Plot 3: DLS scenarios
    ax = axs[1, 1]
    scenario_summary[
        [col for col in scenario_summary.columns if material_as_str + "_" in col]
    ].plot(
        ax=ax,
        lw=3,
        kind="line",
        color=["g", "orange", "r", "g", "orange", "r"],
        style=["-", "-", "-", "--", "--", "--"],
        markersize=15,
        fontsize=22,
    )

    ax.set_ylabel("DLS stocks achieved", fontsize=22)
    ax.set_title("(c) Global DLS achieved (several scenarios)", fontsize=25, pad=10)
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)

    from matplotlib.ticker import FuncFormatter

    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{int(x * 100)}%"))

    from matplotlib.lines import Line2D

    custom_lines = [
        Line2D([0], [0], color="g", lw=3, label="build DLS stocks only (scenario-i)"),
        Line2D(
            [0],
            [0],
            color="orange",
            lw=3,
            label="build some beyond-DLS \n stocks (scenario-ii)",
        ),
        Line2D(
            [0],
            [0],
            color="r",
            lw=3,
            label="build many beyond-DLS \n stocks (scenario-iii)",
        ),
        Line2D(
            [0],
            [0],
            color="black",
            linestyle="-",
            markersize=15,
            label="global speed of construction",
        ),
        Line2D(
            [0],
            [0],
            color="black",
            linestyle="--",
            lw=2,
            label="regional speed of construction",
        ),
    ]

    ax.legend(
        handles=custom_lines,
        title="Legend panel (c)",
        title_fontsize=20,
        loc="best",
        fontsize=20,
        bbox_to_anchor=(-0.225, 0.85),
    )

    custom_lines_panelA = [
        Line2D(
            [0], [0], color="purple", lw=2, label="speed of regions with large DLS gaps"
        ),
        Line2D([0], [0], color="grey", lw=2, label="speed of other regions"),
        Line2D([0], [0], color="blue", lw=4, label="global average speed"),
    ]

    axs[0, 0].legend(handles=custom_lines_panelA, loc="best", fontsize=20)

    ax.set_xlabel("")
    # Leave the bottom-left panel empty
    axs[1, 0].axis("off")

    # legend = ax.legend(fontsize=22)
    # legend.set_bbox_to_anchor((-0.225, 0.85), transform=ax.transAxes)

    fig_title = "Closing DLS gaps by extrapolating historical construction speed"
    fig.suptitle(fig_title, fontsize=30, x=0.5, y=1.04)

    plt.tight_layout()
    plt.subplots_adjust(wspace=0, hspace=0.25)
    plt.show()
    plt.close("all")


def plot_timing_close_DLS_gaps_converged(
    NAS_all_histprosp,
    NAS_all_hist,
    NAS_all_prosp,
    scenario_summary,
    scenario_summary_converged,
    material_as_str,
):
    # Helper function to format billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e9)}B"

    # Set up the 2x2 grid
    fig, axs = plt.subplots(2, 2, figsize=(18, 14))

    # Plot 1: Historical NAS by world regions and world average
    df = NAS_all_hist[[material_as_str]]
    df = df[df.index.get_level_values(1) != 2016]
    df.index = df.index.set_levels(
        df.index.levels[0]
        .to_series()
        .replace(
            {
                "AFR": "Subs. Africa",
                "CPA": "Centr. Asia",
                "EEU": "Eastern Europe",
                "FSU": "Former Sov.Union",
                "LAM": "Latin America",
                "MEA": "N. Africa & M.East",
                "NAM": "North America",
                "PAO": "JP,AU,NZ",
                "PAS": "Pacific Asia",
                "SAS": "South Asia",
                "WEU": "Western Europe",
                "Global": "Global av.",
            }
        ),
        level=0,
    )

    regions = df.index.get_level_values("region").unique()
    columns = df.columns

    ax = axs[0, 0]
    for region in regions:
        df_region = df.xs(region, level="region")
        for column in columns:
            if region == "Global av.":
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=8,
                )
            if region in [
                "Centr. Asia",
                "Subs. Africa",
                "South Asia",
                "Pacific Asia",
                "Latin America",
                "N. Africa & M.East",
            ]:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="purple",
                )
            else:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="grey",
                )

            if region in ["Centr. Asia", "Pacific Asia", "Subs. Africa", "South Asia"]:
                last_year = df_region.index[-7] + 0.1
                last_value = df_region[column].iloc[-1] - 0.1
                ax.text(last_year, last_value, region, fontsize=15, ha="left")
            if region in ["N. Africa & M.East"]:
                last_year = df_region.index[-12] + 0.15
                last_value = df_region[column].iloc[-2] + 0.15
                ax.text(last_year, last_value, region, fontsize=15, ha="left")
            if region in ["Former Sov.Union"]:
                last_year = df_region.index[-17] + 0.1
                last_value = df_region[column].iloc[-17] + 0.1
                ax.text(last_year, last_value, region, fontsize=15, ha="left")

    ax.set_ylabel("tons / capita/year", fontsize=22)
    ax.set_ylim(0, 13)
    ax.set_title(
        r"(a) $\bf{Historical}$ speed of new construction", fontsize=25, pad=10
    )
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)

    # Plot 2: Prospective NAS by world regions and world average
    NAS_all_prosp = NAS_all_histprosp[
        NAS_all_histprosp.index.get_level_values(1).astype(int) > 2016
    ]
    #! set negative NAS to 0 (assuming that no negative NAS exist)
    NAS_all_prosp[NAS_all_prosp < 0] = 0

    df = NAS_all_prosp[[material_as_str]]
    df.index = df.index.set_levels(
        df.index.levels[0]
        .to_series()
        .replace(
            {
                "AFR": "Subs. Africa",
                "CPA": "Centr. Asia",
                "EEU": "Eastern Europe",
                "FSU": "Former Sov.Union",
                "LAM": "Latin America",
                "MEA": "N. Africa & M.East",
                "NAM": "North America",
                "PAO": "JP,AU,NZ",
                "PAS": "Pacific Asia",
                "SAS": "South Asia",
                "WEU": "Western Europe",
                "Global": "Global av.",
            }
        ),
        level=0,
    )

    regions = df.index.get_level_values(0).unique()
    columns = df.columns

    ax = axs[0, 1]
    for region in regions:
        df_region = df.xs(region, level="region")
        for column in columns:
            if region == "Global av.":
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=8,
                )
            if region in [
                "Subs. Africa",
                "South Asia",
                "Pacific Asia",
                "Latin America",
                "N. Africa & M.East",
                "Centr. Asia",
            ]:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="purple",
                )
            else:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="grey",
                )

            if region in ["Subs. Africa", "South Asia", "Pacific Asia"]:
                last_year = df_region.index[-7] + 0.1
                last_value = df_region[column].iloc[-1]
                ax.text(last_year, last_value, region, fontsize=15, ha="left")

            if region in ["Latin America"]:
                last_year = df_region.index[-7] + 0.1
                last_value = df_region[column].iloc[-1] - 0.6
                ax.text(last_year, last_value, region, fontsize=15, ha="left")

            if region in ["North America"]:
                last_year = df_region.index[-21] + 0.1
                first_value = 0.55
                ax.text(
                    last_year,
                    first_value,
                    "North America, Europe, F.Sov.Union, JP/AU/NZ",
                    fontsize=15,
                    ha="left",
                )

            if region in ["Centr. Asia"]:
                last_year = df_region.index[-34] + 0.1
                first_value = 12
                ax.text(
                    last_year,
                    first_value,
                    "Centr. Asia (exception: -1%/a)",
                    fontsize=15,
                    ha="left",
                )

            if region in ["N. Africa & M.East"]:
                last_year = df_region.index[-34] + 0.1
                first_value = df_region[column].iloc[1] + 0.8
                ax.text(last_year, first_value, region, fontsize=15, ha="left")

    ax.set_title(r"(b) $\bf{Prospective}$ speed (extrapolated)", fontsize=25, pad=10)
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)
    ax.set_ylim(0, 13)
    ax.set_yticks([])

    # Plot 3: DLS scenarios
    ax = axs[1, 1]
    scenario_summary[
        [col for col in scenario_summary.columns if material_as_str + "_" in col]
    ].plot(
        ax=ax,
        lw=3,
        kind="line",
        color=["g", "orange", "r", "g", "orange", "r"],
        style=["-", "-", "-", "--", "--", "--"],
        markersize=15,
        fontsize=22,
    )

    ## plot converged practice trajectories with lightened colors
    lightened_colors = [
        [0.5, 0.75, 0.5, 0.4],
        [1.0, 0.8235, 0.5, 0.4],
        [1.0, 0.5, 0.5, 0.4],
        [0.5, 0.75, 0.5, 0.4],
        [1.0, 0.8235, 0.5, 0.4],
        [1.0, 0.5, 0.5, 0.4],
    ]
    scenario_summary_converged[
        [col for col in scenario_summary.columns if material_as_str + "_" in col]
    ].plot(
        ax=ax,
        lw=3,
        kind="line",
        color=lightened_colors,
        style=["-", "-", "-", "--", "--", "--"],
        markersize=15,
        fontsize=22,
    )

    ax.set_ylabel("DLS stock threshold achieved by:", fontsize=22)
    ax.set_title(
        "(c) Global DLS stock threshold achieved (scenarios)", fontsize=25, pad=10
    )
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)

    from matplotlib.ticker import FuncFormatter

    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{int(x * 100)}%"))

    from matplotlib.lines import Line2D

    custom_lines = [
        Line2D([0], [0], color="g", lw=3, label="build DLS stocks only (scenario-i)"),
        Line2D(
            [0],
            [0],
            color="orange",
            lw=3,
            label="build some beyond-DLS \n stocks (scenario-ii)",
        ),
        Line2D(
            [0],
            [0],
            color="r",
            lw=3,
            label="build many beyond-DLS \n stocks (scenario-iii)",
        ),
        Line2D(
            [0],
            [0],
            color="black",
            linestyle="-",
            markersize=15,
            label="global speed of construction",
        ),
        Line2D(
            [0],
            [0],
            color="black",
            linestyle="--",
            lw=2,
            label="regional speed of construction",
        ),
        Line2D([0], [0], color="g", lw=3, label="saturated colors = current practices"),
        Line2D(
            [0],
            [0],
            color="g",
            alpha=0.2,
            lw=3,
            label="light colors = converged practices",
        ),
    ]

    ax.legend(
        handles=custom_lines,
        title="Trajectories closing DLS stock gaps:",
        title_fontsize=20,
        loc="best",
        fontsize=20,
        bbox_to_anchor=(-0.225, 0.85),
    )

    custom_lines_panelA = [
        Line2D(
            [0], [0], color="purple", lw=2, label="speed of regions with large DLS gaps"
        ),
        Line2D([0], [0], color="grey", lw=2, label="speed of other regions"),
        Line2D([0], [0], color="blue", lw=4, label="global average speed"),
    ]

    axs[0, 0].legend(handles=custom_lines_panelA, loc="best", fontsize=20)

    ax.set_xlabel("")
    # Leave the bottom-left panel empty
    axs[1, 0].axis("off")

    fig_title = "Closing DLS gaps by extrapolating historical construction speed"
    fig.suptitle(fig_title, fontsize=30, x=0.5, y=1.04)

    plt.tight_layout()
    plt.subplots_adjust(wspace=0, hspace=0.25)
    plt.show()
    #fig.savefig("Fig_5.pdf", bbox_inches='tight')
    plt.close("all")


def plot_timing_close_DLS_gaps_noYlim(
    NAS_all_histprosp, NAS_all_hist, NAS_all_prosp, scenario_summary, material_as_str
):
    # Helper function to format billions
    def billion_formatter(x, pos):
        return f"{int(x / 1e9)}B"

    # Set up the 2x2 grid
    fig, axs = plt.subplots(2, 2, figsize=(18, 14))

    # Plot 1: Historical NAS by world regions and world average
    df = NAS_all_hist[[material_as_str]]
    df = df[df.index.get_level_values(1) != 2016]
    df.index = df.index.set_levels(
        df.index.levels[0]
        .to_series()
        .replace(
            {
                "AFR": "Subs. Africa",
                "CPA": "Centr. Asia",
                "EEU": "Eastern Europe",
                "FSU": "Former Sov.Union",
                "LAM": "Latin America",
                "MEA": "N. Africa & M.East",
                "NAM": "North America",
                "PAO": "JP,AU,NZ",
                "PAS": "Pacific Asia",
                "SAS": "South Asia",
                "WEU": "Western Europe",
                "Global": "Global av.",
            }
        ),
        level=0,
    )

    regions = df.index.get_level_values("region").unique()
    columns = df.columns

    # global_North = ['EEU', 'FSU',  'NAM', 'PAO', 'WEU']
    # global_South_noCPA = ['AFR', 'LAM', 'MEA', 'PAS', 'SAS']

    ax = axs[0, 0]
    for region in regions:
        df_region = df.xs(region, level="region")
        for column in columns:
            if region == "Global av.":
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=8,
                )
            else:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="grey",
                )

    ax.set_ylabel("tons / capita/year", fontsize=22)
    ax.set_title(
        r"(a) $\bf{Historical}$ speed of new construction", fontsize=25, pad=10
    )
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)
    # ax.set_xticks([2015,2020,2025,2030,2035,2040,2045,2050], fontsize=22)

    # Plot 2: Prospective NAS by world regions and world average
    NAS_all_prosp = NAS_all_histprosp[
        NAS_all_histprosp.index.get_level_values(1).astype(int) > 2016
    ].copy()
    #! set negative NAS to 0 (assuming that no negative NAS exist)
    NAS_all_prosp[NAS_all_prosp < 0] = 0

    df = NAS_all_prosp[[material_as_str]]
    df.index = df.index.set_levels(
        df.index.levels[0]
        .to_series()
        .replace(
            {
                "AFR": "Subs. Africa",
                "CPA": "Centr. Asia",
                "EEU": "Eastern Europe",
                "FSU": "Former Sov.Union",
                "LAM": "Latin America",
                "MEA": "N. Africa & M.East",
                "NAM": "North America",
                "PAO": "JP,AU,NZ",
                "PAS": "Pacific Asia",
                "SAS": "South Asia",
                "WEU": "Western Europe",
                "Global": "Global av.",
            }
        ),
        level=0,
    )

    regions = df.index.get_level_values(0).unique()
    columns = df.columns
    first_subplot_ylim = ax.get_ylim()

    ax = axs[0, 1]
    for region in regions:
        df_region = df.xs(region, level="region")
        for column in columns:
            if region == "Global av.":
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=8,
                )
            else:
                ax.plot(
                    df_region.index,
                    df_region[column],
                    label=f"{region} - {column}",
                    linewidth=2,
                    color="grey",
                )

    ax.set_title(r"(b) $\bf{Prospective}$ speed (extrapolated)", fontsize=25, pad=10)
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)
    ax.set_ylim(first_subplot_ylim[0], first_subplot_ylim[1])
    # ax.set_ylim(0,12)
    ax.set_yticks([])

    # Plot 3: DLS scenarios
    ax = axs[1, 1]
    scenario_summary[
        [col for col in scenario_summary.columns if material_as_str + "_" in col]
    ].plot(
        ax=ax,
        lw=3,
        kind="line",
        color=["g", "orange", "r", "g", "orange", "r"],
        style=["-", "-", "-", "--", "--", "--"],
        markersize=15,
        fontsize=22,
    )

    ax.set_ylabel("DLS achieved", fontsize=22)
    ax.set_title("(c) Global DLS achieved (several scenarios)", fontsize=25, pad=10)
    ax.tick_params(axis="y", labelsize=22)
    ax.tick_params(axis="x", labelsize=22)

    from matplotlib.ticker import FuncFormatter

    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{int(x * 100)}%"))

    from matplotlib.lines import Line2D

    custom_lines = [
        Line2D([0], [0], color="g", lw=3, label="build DLS stocks only (scenario-i)"),
        Line2D(
            [0],
            [0],
            color="orange",
            lw=3,
            label="build some beyond-DLS \n stocks (scenario-ii)",
        ),
        Line2D(
            [0],
            [0],
            color="r",
            lw=3,
            label="build many beyond-DLS \n stocks (scenario-iii)",
        ),
        Line2D(
            [0],
            [0],
            color="black",
            linestyle="-",
            markersize=15,
            label="global speed of construction",
        ),
        Line2D(
            [0],
            [0],
            color="black",
            linestyle="--",
            lw=2,
            label="regional speed of construction",
        ),
    ]

    ax.legend(
        handles=custom_lines,
        title="Legend panel (c)",
        title_fontsize=20,
        loc="best",
        fontsize=20,
        bbox_to_anchor=(-0.225, 0.85),
    )

    custom_lines_panelA = [
        Line2D([0], [0], color="grey", lw=2, label="speed of other regions"),
        Line2D([0], [0], color="blue", lw=4, label="global average speed"),
    ]

    axs[0, 0].legend(handles=custom_lines_panelA, loc="best", fontsize=20)

    ax.set_xlabel("")
    # Leave the bottom-left panel empty
    axs[1, 0].axis("off")

    fig_title = (
        "Closing DLS gaps at past construction speed ("
        + material_as_str
        + " net additions to stock)"
    )
    fig.suptitle(fig_title, fontsize=30, x=0.5, y=1.04)

    plt.tight_layout()
    plt.subplots_adjust(wspace=0, hspace=0.25)
    plt.show()
    plt.close("all")
